from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, session
import json
import sqlite3
import re
import unicodedata
import secrets
import tempfile
from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, timedelta
import io
import time
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

BASE = Path(__file__).resolve().parent
DB = BASE / "due.db"
app = Flask(__name__)
app.secret_key = "troque-esta-chave-em-producao"

CONTRACT_IMPORT_STAGE_TTL = 1800
CONTRACT_IMPORT_STAGE_PREFIX = "duecontrol_contract_import_"
INVOICE_IMPORT_STAGE_TTL = 1800
INVOICE_IMPORT_STAGE_PREFIX = "duecontrol_invoice_import_"
INVOICE_CONTRACT_SCHEMA_VERSION = 1
INVOICE_SCHEMA_VERSION = 3

SALDO_TOLERANCE = Decimal("0.005")
STATUS_PENDENTE = "PENDENTE"
STATUS_CONCLUIDO = "CONCLUÍDO"
STATUS_PARCIAL = "PARCIAL"
INVOICE_STATUS_AGUARDANDO_RECEBIMENTO = "AGUARDANDO_RECEBIMENTO"
# Mantido como alias de compatibilidade com o estado calculado antigo, que
# possuía uma opção PARCIAL separada.
INVOICE_STATUS_PARCIAL = INVOICE_STATUS_AGUARDANDO_RECEBIMENTO
INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO = "RECEBIDA_AGUARDANDO_CAMBIO"
INVOICE_STATUS_LIQUIDADA = "LIQUIDADA"
INVOICE_STATUS_LABELS = {
    INVOICE_STATUS_AGUARDANDO_RECEBIMENTO: "AGUARDANDO RECEBIMENTO",
    INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO: "RECEBIDO AGUARDANDO CAMBIO",
    INVOICE_STATUS_LIQUIDADA: "LIQUIDADA",
}
INVOICE_STATUS_OPTIONS = tuple(INVOICE_STATUS_LABELS.values())

def normalize_invoice_status(value, default=None):
    """Normaliza o status da Invoice e aceita valores das versões anteriores."""
    if value is None or str(value).strip() == "":
        return default
    text_value = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    key = re.sub(r"[^A-Z0-9]+", "_", text_value.upper()).strip("_")
    aliases = {
        "AGUARDANDO_RECEBIMENTO": INVOICE_STATUS_AGUARDANDO_RECEBIMENTO,
        "AGUARDANDO_RECEBIMENTO_INVOICE": INVOICE_STATUS_AGUARDANDO_RECEBIMENTO,
        "PARCIAL": INVOICE_STATUS_AGUARDANDO_RECEBIMENTO,
        "RECEBIDO": INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO,
        "RECEBIDO_AGUARDANDO_CAMBIO": INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO,
        "RECEBIDA_AGUARDANDO_CAMBIO": INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO,
        "LIQUIDADA": INVOICE_STATUS_LIQUIDADA,
    }
    normalized = aliases.get(key)
    if normalized is None:
        raise ValueError(
            "Status de Invoice invalido. Use: " + ", ".join(INVOICE_STATUS_OPTIONS) + "."
        )
    return normalized
INVOICE_STATUS_TYPES = (
    "PROFORMA", "COMMERCIAL_INVOICE", "SERVICE_INVOICE", "DEBIT_NOTE"
)
NDF_STATUS_ATIVA = "ATIVA"
NDF_STATUS_LIQUIDADA = "LIQUIDADA"
NDF_STATUS_CANCELADA = "CANCELADA"
NDF_STATUS_VENCIDA = "VENCIDA"
NDF_STATUSES = (NDF_STATUS_ATIVA, NDF_STATUS_LIQUIDADA, NDF_STATUS_CANCELADA)
NDF_TIPOS = ("BANCO", "TRADING")
NDF_POSICOES = ("COMPRA", "VENDA")
PTAX_MOEDAS = ("USD", "EUR")
PTAX_API_URL = "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/CotacaoMoedaPeriodo"
PTAX_API_TIMEOUT = 20
CLIENTE_PAISES = (
    ("AD", "Andorra"), ("AE", "Emirados Árabes Unidos"), ("AF", "Afeganistão"),
    ("AG", "Antígua e Barbuda"), ("AI", "Anguilla"), ("AL", "Albânia"),
    ("AM", "Armênia"), ("AO", "Angola"), ("AQ", "Antártida"), ("AR", "Argentina"),
    ("AS", "Samoa Americana"), ("AT", "Áustria"), ("AU", "Austrália"), ("AW", "Aruba"),
    ("AX", "Ilhas Åland"), ("AZ", "Azerbaijão"), ("BA", "Bósnia e Herzegovina"),
    ("BB", "Barbados"), ("BD", "Bangladesh"), ("BE", "Bélgica"), ("BF", "Burkina Faso"),
    ("BG", "Bulgária"), ("BH", "Bahrein"), ("BI", "Burundi"), ("BJ", "Benim"),
    ("BL", "São Bartolomeu"), ("BM", "Bermudas"), ("BN", "Brunei"), ("BO", "Bolívia"),
    ("BQ", "Bonaire, Santo Eustáquio e Saba"), ("BR", "Brasil"), ("BS", "Bahamas"),
    ("BT", "Butão"), ("BV", "Ilha Bouvet"), ("BW", "Botsuana"), ("BY", "Belarus"),
    ("BZ", "Belize"), ("CA", "Canadá"), ("CC", "Ilhas Cocos"),
    ("CD", "República Democrática do Congo"), ("CF", "República Centro-Africana"),
    ("CG", "República do Congo"), ("CH", "Suíça"), ("CI", "Costa do Marfim"),
    ("CK", "Ilhas Cook"), ("CL", "Chile"), ("CM", "Camarões"), ("CN", "China"),
    ("CO", "Colômbia"), ("CR", "Costa Rica"), ("CU", "Cuba"), ("CV", "Cabo Verde"),
    ("CW", "Curaçau"), ("CX", "Ilha Christmas"), ("CY", "Chipre"), ("CZ", "Tchéquia"),
    ("DE", "Alemanha"), ("DJ", "Djibuti"), ("DK", "Dinamarca"), ("DM", "Dominica"),
    ("DO", "República Dominicana"), ("DZ", "Argélia"), ("EC", "Equador"), ("EE", "Estônia"),
    ("EG", "Egito"), ("EH", "Saara Ocidental"), ("ER", "Eritreia"), ("ES", "Espanha"),
    ("ET", "Etiópia"), ("FI", "Finlândia"), ("FJ", "Fiji"), ("FK", "Ilhas Malvinas"),
    ("FM", "Micronésia"), ("FO", "Ilhas Faroé"), ("FR", "França"), ("GA", "Gabão"),
    ("GB", "Reino Unido"), ("GD", "Granada"), ("GE", "Geórgia"), ("GF", "Guiana Francesa"),
    ("GG", "Guernsey"), ("GH", "Gana"), ("GI", "Gibraltar"), ("GL", "Groenlândia"),
    ("GM", "Gâmbia"), ("GN", "Guiné"), ("GP", "Guadalupe"), ("GQ", "Guiné Equatorial"),
    ("GR", "Grécia"), ("GS", "Geórgia do Sul e Ilhas Sandwich do Sul"), ("GT", "Guatemala"),
    ("GU", "Guam"), ("GW", "Guiné-Bissau"), ("GY", "Guiana"), ("HK", "Hong Kong"),
    ("HM", "Ilhas Heard e McDonald"), ("HN", "Honduras"), ("HR", "Croácia"), ("HT", "Haiti"),
    ("HU", "Hungria"), ("ID", "Indonésia"), ("IE", "Irlanda"), ("IL", "Israel"),
    ("IM", "Ilha de Man"), ("IN", "Índia"), ("IO", "Território Britânico do Oceano Índico"),
    ("IQ", "Iraque"), ("IR", "Irã"), ("IS", "Islândia"), ("IT", "Itália"),
    ("JE", "Jersey"), ("JM", "Jamaica"), ("JO", "Jordânia"), ("JP", "Japão"),
    ("KE", "Quênia"), ("KG", "Quirguistão"), ("KH", "Camboja"), ("KI", "Kiribati"),
    ("KM", "Comores"), ("KN", "São Cristóvão e Névis"), ("KP", "Coreia do Norte"),
    ("KR", "Coreia do Sul"), ("KW", "Kuwait"), ("KY", "Ilhas Cayman"), ("KZ", "Cazaquistão"),
    ("LA", "Laos"), ("LB", "Líbano"), ("LC", "Santa Lúcia"), ("LI", "Liechtenstein"),
    ("LK", "Sri Lanka"), ("LR", "Libéria"), ("LS", "Lesoto"), ("LT", "Lituânia"),
    ("LU", "Luxemburgo"), ("LV", "Letônia"), ("LY", "Líbia"), ("MA", "Marrocos"),
    ("MC", "Mônaco"), ("MD", "Moldávia"), ("ME", "Montenegro"), ("MF", "São Martinho"),
    ("MG", "Madagascar"), ("MH", "Ilhas Marshall"), ("MK", "Macedônia do Norte"),
    ("ML", "Mali"), ("MM", "Mianmar"), ("MN", "Mongólia"), ("MO", "Macau"),
    ("MP", "Ilhas Marianas do Norte"), ("MQ", "Martinica"), ("MR", "Mauritânia"),
    ("MS", "Montserrat"), ("MT", "Malta"), ("MU", "Maurício"), ("MV", "Maldivas"),
    ("MW", "Malaui"), ("MX", "México"), ("MY", "Malásia"), ("MZ", "Moçambique"),
    ("NA", "Namíbia"), ("NC", "Nova Caledônia"), ("NE", "Níger"), ("NF", "Ilha Norfolk"),
    ("NG", "Nigéria"), ("NI", "Nicarágua"), ("NL", "Países Baixos"), ("NO", "Noruega"),
    ("NP", "Nepal"), ("NR", "Nauru"), ("NU", "Niue"), ("NZ", "Nova Zelândia"),
    ("OM", "Omã"), ("PA", "Panamá"), ("PE", "Peru"), ("PF", "Polinésia Francesa"),
    ("PG", "Papua-Nova Guiné"), ("PH", "Filipinas"), ("PK", "Paquistão"), ("PL", "Polônia"),
    ("PM", "São Pedro e Miquelão"), ("PN", "Pitcairn"), ("PR", "Porto Rico"),
    ("PS", "Palestina"), ("PT", "Portugal"), ("PW", "Palau"), ("PY", "Paraguai"),
    ("QA", "Catar"), ("RE", "Reunião"), ("RO", "Romênia"), ("RS", "Sérvia"),
    ("RU", "Rússia"), ("RW", "Ruanda"), ("SA", "Arábia Saudita"), ("SB", "Ilhas Salomão"),
    ("SC", "Seicheles"), ("SD", "Sudão"), ("SE", "Suécia"), ("SG", "Singapura"),
    ("SH", "Santa Helena, Ascensão e Tristão da Cunha"), ("SI", "Eslovênia"),
    ("SJ", "Svalbard e Jan Mayen"), ("SK", "Eslováquia"), ("SL", "Serra Leoa"),
    ("SM", "San Marino"), ("SN", "Senegal"), ("SO", "Somália"), ("SR", "Suriname"),
    ("SS", "Sudão do Sul"), ("ST", "São Tomé e Príncipe"), ("SV", "El Salvador"),
    ("SX", "Sint Maarten"), ("SY", "Síria"), ("SZ", "Essuatíni"), ("TC", "Ilhas Turks e Caicos"),
    ("TD", "Chade"), ("TF", "Terras Austrais e Antárticas Francesas"), ("TG", "Togo"),
    ("TH", "Tailândia"), ("TJ", "Tajiquistão"), ("TK", "Tokelau"), ("TL", "Timor-Leste"),
    ("TM", "Turcomenistão"), ("TN", "Tunísia"), ("TO", "Tonga"), ("TR", "Turquia"),
    ("TT", "Trinidad e Tobago"), ("TV", "Tuvalu"), ("TW", "Taiwan"), ("TZ", "Tanzânia"),
    ("UA", "Ucrânia"), ("UG", "Uganda"), ("UM", "Ilhas Menores Distantes dos Estados Unidos"),
    ("US", "Estados Unidos"), ("UY", "Uruguai"), ("UZ", "Uzbequistão"), ("VA", "Santa Sé"),
    ("VC", "São Vicente e Granadinas"), ("VE", "Venezuela"), ("VG", "Ilhas Virgens Britânicas"),
    ("VI", "Ilhas Virgens Americanas"), ("VN", "Vietnã"), ("VU", "Vanuatu"),
    ("WF", "Wallis e Futuna"), ("WS", "Samoa"), ("YE", "Iêmen"), ("YT", "Mayotte"),
    ("ZA", "África do Sul"), ("ZM", "Zâmbia"), ("ZW", "Zimbábue"),
)
CLIENTE_PAISES_MAP = dict(CLIENTE_PAISES)

def pais_sort_key(item):
    nome = unicodedata.normalize("NFKD", item[1])
    nome = "".join(caractere for caractere in nome if not unicodedata.combining(caractere))
    return nome.casefold()

CLIENTE_PAISES_ORDENADOS = tuple(sorted(CLIENTE_PAISES, key=pais_sort_key))

def db():
    conn = sqlite3.connect(DB, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA busy_timeout = 30000")
    return conn

def init_db():
    conn = db()
    try:
        schema_version = conn.execute("PRAGMA user_version").fetchone()[0]
        conn.executescript("""
    CREATE TABLE IF NOT EXISTS empresas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        razao_social TEXT NOT NULL,
        cnpj TEXT NOT NULL UNIQUE,
        apelido TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS competencias (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        empresa_id INTEGER NOT NULL,
        descricao TEXT NOT NULL,
        data_inicial TEXT NOT NULL,
        data_final TEXT NOT NULL,
        status TEXT NOT NULL DEFAULT 'ABERTA',
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE CASCADE,
        UNIQUE(empresa_id, descricao)
    );

    CREATE TABLE IF NOT EXISTS clientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL COLLATE NOCASE,
        pais TEXT NOT NULL,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(nome, pais)
    );

    CREATE TABLE IF NOT EXISTS contrapartes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL COLLATE NOCASE UNIQUE,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS dues (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero_due TEXT NOT NULL UNIQUE,
        chave_acesso TEXT,
        data_due TEXT,
        cnpj TEXT,
        cliente TEXT,
        moeda TEXT NOT NULL DEFAULT 'USD',
        valor_original REAL NOT NULL DEFAULT 0,
        status TEXT NOT NULL DEFAULT 'PENDENTE',
        observacao TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        ,competencia_id INTEGER
    );

    CREATE TABLE IF NOT EXISTS contratos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero_contrato TEXT NOT NULL UNIQUE,
        banco TEXT,
        banco_credito TEXT,
        banco_liquidacao TEXT,
        data_contrato TEXT,
        data_recebimento TEXT,
        data_liquidacao TEXT,
        cnpj TEXT,
        cliente TEXT,
        moeda TEXT NOT NULL DEFAULT 'USD',
        valor_moeda REAL NOT NULL DEFAULT 0,
        taxa_cambio REAL,
        valor_reais REAL,
        status TEXT NOT NULL DEFAULT 'PENDENTE',
        saldo_zerado_manual INTEGER NOT NULL DEFAULT 0,
        observacao TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        ,competencia_id INTEGER
    );

    CREATE TABLE IF NOT EXISTS ndfs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero_operacao TEXT NOT NULL UNIQUE,
        cnpj TEXT NOT NULL,
        contraparte TEXT NOT NULL,
        tipo TEXT NOT NULL,
        moeda TEXT NOT NULL DEFAULT 'USD',
        valor_contratado REAL NOT NULL DEFAULT 0,
        taxa_contratada REAL NOT NULL,
        data_contratacao TEXT NOT NULL,
        data_vencimento TEXT NOT NULL,
        posicao TEXT NOT NULL,
        finalidade TEXT,
        observacao TEXT,
        status TEXT NOT NULL DEFAULT 'ATIVA',
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        ,competencia_id INTEGER
    );

    CREATE TABLE IF NOT EXISTS ptax_cotacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data_cotacao TEXT NOT NULL,
        moeda TEXT NOT NULL,
        ptax_compra REAL NOT NULL,
        ptax_venda REAL NOT NULL,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(moeda, data_cotacao)
    );

    CREATE TABLE IF NOT EXISTS due_movimentacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        due_id INTEGER NOT NULL,
        contrato_id INTEGER,
        due_contrato_id INTEGER,
        data_movimentacao TEXT NOT NULL,
        tipo TEXT NOT NULL DEFAULT 'UTILIZACAO',
        documento TEXT,
        valor REAL NOT NULL DEFAULT 0,
        observacao TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (due_id) REFERENCES dues(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS due_contratos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        due_id INTEGER NOT NULL,
        contrato_id INTEGER NOT NULL,
        valor_vinculado REAL NOT NULL DEFAULT 0,
        observacao TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(due_id, contrato_id),
        FOREIGN KEY (due_id) REFERENCES dues(id) ON DELETE CASCADE,
        FOREIGN KEY (contrato_id) REFERENCES contratos(id) ON DELETE CASCADE
    );
        """)
        if schema_version < INVOICE_CONTRACT_SCHEMA_VERSION:
            contrato_count = conn.execute("SELECT COUNT(*) FROM contratos").fetchone()[0]
            if contrato_count:
                raise RuntimeError(
                    "A atualização do fluxo de Invoices exige uma base de contratos vazia."
                )
            conn.execute("DROP TABLE IF EXISTS due_contratos")
            conn.execute("DROP TABLE contratos")
            conn.execute("""
                CREATE TABLE contratos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    numero_contrato TEXT NOT NULL UNIQUE,
                    banco_liquidacao_id INTEGER,
                    banco_liquidacao TEXT,
                    data_fechamento TEXT,
                    data_liquidacao TEXT,
                    moeda TEXT NOT NULL DEFAULT 'USD',
                    taxa_cambio REAL,
                    status TEXT NOT NULL DEFAULT 'PENDENTE',
                    observacao TEXT,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    valor_moeda REAL NOT NULL DEFAULT 0,
                    valor_reais REAL,
                    banco TEXT,
                    banco_id INTEGER,
                    banco_credito TEXT,
                    data_contrato TEXT,
                    data_recebimento TEXT,
                    cnpj TEXT,
                    cliente TEXT,
                    cliente_id INTEGER,
                    competencia_id INTEGER,
                    saldo_zerado_manual INTEGER NOT NULL DEFAULT 0,
                    FOREIGN KEY (banco_liquidacao_id) REFERENCES contrapartes(id) ON DELETE SET NULL
                )
            """)
            conn.execute("""
                CREATE TABLE due_contratos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    due_id INTEGER NOT NULL,
                    contrato_id INTEGER NOT NULL,
                    valor_vinculado REAL NOT NULL DEFAULT 0,
                    observacao TEXT,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(due_id, contrato_id),
                    FOREIGN KEY (due_id) REFERENCES dues(id) ON DELETE CASCADE,
                    FOREIGN KEY (contrato_id) REFERENCES contratos(id) ON DELETE CASCADE
                )
            """)
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                empresa_id INTEGER NOT NULL,
                numero_invoice TEXT NOT NULL,
                tipo_documento TEXT NOT NULL CHECK(tipo_documento IN
                    ('PROFORMA','COMMERCIAL_INVOICE','SERVICE_INVOICE','DEBIT_NOTE')),
                competencia_id INTEGER,
                cliente_id INTEGER,
                data_emissao TEXT,
                data_credito TEXT,
                moeda TEXT NOT NULL DEFAULT 'USD',
                valor_moeda REAL NOT NULL DEFAULT 0 CHECK(valor_moeda >= 0),
                status TEXT NOT NULL DEFAULT 'AGUARDANDO_RECEBIMENTO' CHECK(status IN
                    ('AGUARDANDO_RECEBIMENTO','RECEBIDA_AGUARDANDO_CAMBIO','LIQUIDADA')),
                status_manual INTEGER NOT NULL DEFAULT 0,
                observacao TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE RESTRICT,
                FOREIGN KEY (cliente_id) REFERENCES clientes(id) ON DELETE SET NULL,
                UNIQUE(empresa_id, numero_invoice, tipo_documento)
            );

            CREATE TABLE IF NOT EXISTS recebimentos_invoice (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER NOT NULL,
                banco_credito_id INTEGER,
                data_credito TEXT NOT NULL,
                moeda TEXT NOT NULL,
                valor_moeda REAL NOT NULL CHECK(valor_moeda > 0),
                documento TEXT,
                observacao TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (invoice_id) REFERENCES invoices(id) ON DELETE CASCADE,
                FOREIGN KEY (banco_credito_id) REFERENCES contrapartes(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS invoice_contrato_cambio (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER NOT NULL,
                contrato_id INTEGER NOT NULL,
                valor_alocado REAL NOT NULL CHECK(valor_alocado > 0),
                observacao TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(invoice_id, contrato_id),
                FOREIGN KEY (invoice_id) REFERENCES invoices(id) ON DELETE CASCADE,
                FOREIGN KEY (contrato_id) REFERENCES contratos(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS due_invoice (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                due_id INTEGER NOT NULL,
                invoice_id INTEGER NOT NULL,
                valor_vinculado REAL NOT NULL CHECK(valor_vinculado > 0),
                observacao TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(due_id, invoice_id),
                FOREIGN KEY (due_id) REFERENCES dues(id) ON DELETE CASCADE,
                FOREIGN KEY (invoice_id) REFERENCES invoices(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_invoices_empresa ON invoices(empresa_id);
            CREATE INDEX IF NOT EXISTS idx_invoices_status ON invoices(status);
            CREATE INDEX IF NOT EXISTS idx_invoices_emissao ON invoices(data_emissao);
            CREATE INDEX IF NOT EXISTS idx_recebimentos_invoice ON recebimentos_invoice(invoice_id);
            CREATE INDEX IF NOT EXISTS idx_recebimentos_banco ON recebimentos_invoice(banco_credito_id);
            CREATE INDEX IF NOT EXISTS idx_invoice_contrato_invoice ON invoice_contrato_cambio(invoice_id);
            CREATE INDEX IF NOT EXISTS idx_invoice_contrato_contrato ON invoice_contrato_cambio(contrato_id);
            CREATE INDEX IF NOT EXISTS idx_due_invoice_due ON due_invoice(due_id);
            CREATE INDEX IF NOT EXISTS idx_due_invoice_invoice ON due_invoice(invoice_id);
            CREATE INDEX IF NOT EXISTS idx_contratos_fechamento ON contratos(data_fechamento);
        """)
        if schema_version < INVOICE_SCHEMA_VERSION:
            invoice_columns = {row[1] for row in conn.execute("PRAGMA table_info(invoices)")}
            if "contrato_comercial" not in invoice_columns:
                conn.execute("ALTER TABLE invoices ADD COLUMN contrato_comercial TEXT")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_invoices_contrato_comercial ON invoices(contrato_comercial)")
            if schema_version < 3 and "competencia_id" not in invoice_columns:
                conn.execute("ALTER TABLE invoices ADD COLUMN competencia_id INTEGER")
        invoice_columns = {row[1] for row in conn.execute("PRAGMA table_info(invoices)")}
        if "contrato_comercial" not in invoice_columns:
            conn.execute("ALTER TABLE invoices ADD COLUMN contrato_comercial TEXT")
            invoice_columns.add("contrato_comercial")
        if "competencia_id" not in invoice_columns:
            conn.execute("ALTER TABLE invoices ADD COLUMN competencia_id INTEGER")
        if "status_manual" not in invoice_columns:
            conn.execute("ALTER TABLE invoices ADD COLUMN status_manual INTEGER NOT NULL DEFAULT 0")
        if "data_credito" not in invoice_columns:
            conn.execute("ALTER TABLE invoices ADD COLUMN data_credito TEXT")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_invoices_contrato_comercial ON invoices(contrato_comercial)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_invoices_competencia ON invoices(competencia_id)")
        for invoice in conn.execute("SELECT id, status FROM invoices").fetchall():
            try:
                normalized_status = normalize_invoice_status(
                    invoice["status"], default=INVOICE_STATUS_AGUARDANDO_RECEBIMENTO
                )
            except ValueError:
                normalized_status = INVOICE_STATUS_AGUARDANDO_RECEBIMENTO
            if normalized_status != invoice["status"]:
                conn.execute("UPDATE invoices SET status=? WHERE id=?", (normalized_status, invoice["id"]))
        conn.execute("""
            UPDATE invoices
            SET data_credito=(SELECT MIN(r.data_credito)
                              FROM recebimentos_invoice r WHERE r.invoice_id=invoices.id)
            WHERE data_credito IS NULL AND status IN (?, ?)
        """, (INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO, INVOICE_STATUS_LIQUIDADA))
        conn.execute(f"PRAGMA user_version = {INVOICE_SCHEMA_VERSION}")
        columns = {row[1] for row in conn.execute("PRAGMA table_info(dues)")}
        if "chave_acesso" not in columns:
            conn.execute("ALTER TABLE dues ADD COLUMN chave_acesso TEXT")
        if "created_at" not in columns:
            conn.execute("ALTER TABLE dues ADD COLUMN created_at TEXT")
            conn.execute("UPDATE dues SET created_at=COALESCE(data_due, CURRENT_TIMESTAMP) WHERE created_at IS NULL")
        contrato_columns = {row[1] for row in conn.execute("PRAGMA table_info(contratos)")}
        if "banco" not in contrato_columns:
            conn.execute("ALTER TABLE contratos ADD COLUMN banco TEXT")
        if "banco_id" not in contrato_columns:
            conn.execute("ALTER TABLE contratos ADD COLUMN banco_id INTEGER")
        if "banco_credito" not in contrato_columns:
            conn.execute("ALTER TABLE contratos ADD COLUMN banco_credito TEXT")
        if "banco_liquidacao" not in contrato_columns:
            conn.execute("ALTER TABLE contratos ADD COLUMN banco_liquidacao TEXT")
        if "data_recebimento" not in contrato_columns:
            conn.execute("ALTER TABLE contratos ADD COLUMN data_recebimento TEXT")
        if "data_liquidacao" not in contrato_columns:
            conn.execute("ALTER TABLE contratos ADD COLUMN data_liquidacao TEXT")
        conn.execute("UPDATE contratos SET banco_credito=COALESCE(NULLIF(banco_credito,''),banco), banco_liquidacao=COALESCE(NULLIF(banco_liquidacao,''),banco_credito,banco)")
        if "cliente_id" not in contrato_columns:
            conn.execute("ALTER TABLE contratos ADD COLUMN cliente_id INTEGER")
        if "saldo_zerado_manual" not in contrato_columns:
            conn.execute("ALTER TABLE contratos ADD COLUMN saldo_zerado_manual INTEGER NOT NULL DEFAULT 0")
        if "competencia_id" not in contrato_columns:
            conn.execute("ALTER TABLE contratos ADD COLUMN competencia_id INTEGER")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_contratos_data_contrato ON contratos(data_contrato)")
        due_columns = {row[1] for row in conn.execute("PRAGMA table_info(dues)")}
        if "competencia_id" not in due_columns:
            conn.execute("ALTER TABLE dues ADD COLUMN competencia_id INTEGER")
        ndf_columns = {row[1] for row in conn.execute("PRAGMA table_info(ndfs)")}
        if "cliente_id" not in ndf_columns:
            conn.execute("ALTER TABLE ndfs ADD COLUMN cliente_id INTEGER")
        if "contraparte_id" not in ndf_columns:
            conn.execute("ALTER TABLE ndfs ADD COLUMN contraparte_id INTEGER")
        if "competencia_id" not in ndf_columns:
            conn.execute("ALTER TABLE ndfs ADD COLUMN competencia_id INTEGER")
        conn.execute("""CREATE UNIQUE INDEX IF NOT EXISTS idx_dues_chave_acesso
                       ON dues(chave_acesso)
                       WHERE chave_acesso IS NOT NULL AND chave_acesso <> ''""")
        movimentacao_columns = {row[1] for row in conn.execute("PRAGMA table_info(due_movimentacoes)")}
        if "contrato_id" not in movimentacao_columns:
            conn.execute("ALTER TABLE due_movimentacoes ADD COLUMN contrato_id INTEGER")
        if "due_contrato_id" not in movimentacao_columns:
            conn.execute("ALTER TABLE due_movimentacoes ADD COLUMN due_contrato_id INTEGER")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_mov_due_tipo ON due_movimentacoes(due_id, tipo)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_mov_contrato_tipo ON due_movimentacoes(contrato_id, tipo)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_mov_due_contrato ON due_movimentacoes(due_contrato_id)")
        for link in conn.execute("""
            SELECT v.id, v.due_id, v.contrato_id, v.valor_vinculado
            FROM due_contratos v
            LEFT JOIN due_movimentacoes m
              ON m.due_contrato_id=v.id
              OR (m.due_id=v.due_id AND m.contrato_id=v.contrato_id AND m.tipo='VINCULACAO')
            WHERE m.id IS NULL
        """).fetchall():
            conn.execute("""INSERT INTO due_movimentacoes
                (due_id,contrato_id,due_contrato_id,data_movimentacao,tipo,documento,valor,observacao)
                VALUES (?,?,?,?,?,?,?,?)""",
                (link["due_id"], link["contrato_id"], link["id"], date.today().isoformat(),
                 "VINCULACAO", f"VINCULO:{link['id']}", link["valor_vinculado"],
                 "Movimentação criada na migração do vínculo existente."))
        recalculate_statuses(conn)
        conn.commit()
    finally:
        conn.close()

@app.template_filter("money")
def money(v):
    try:
        return f"{float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,00"

@app.template_filter("rate")
def rate(v):
    try:
        return f"{float(v):,.4f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,0000"

@app.template_filter("date_br")
def date_br(v):
    normalized = normalize_date(v)
    if not normalized:
        return ""
    return normalized[8:10] + "/" + normalized[5:7] + "/" + normalized[0:4]

def normalize_date(value):
    if value is None or str(value).strip() in {"", "NaT", "nan"}:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
        except (OverflowError, ValueError):
            return str(value)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    try:
        return (datetime(1899, 12, 30) + timedelta(days=float(text))).strftime("%Y-%m-%d")
    except (OverflowError, ValueError):
        return text

def parse_date(value):
    normalized = normalize_date(value)
    if not normalized:
        return None
    try:
        datetime.strptime(normalized, "%Y-%m-%d")
        return normalized
    except ValueError:
        raise ValueError("Data inválida. Use o formato dd/mm/aaaa.")

def normalize_contract_import_columns(columns):
    """Normaliza cabeçalhos do Excel para os nomes canônicos dos contratos."""
    aliases = {
        "numero_contrato": "numero_contrato",
        "numero_do_contrato": "numero_contrato",
        "numero_de_contrato": "numero_contrato",
        "banco": "banco",
        "banco_credito": "banco_credito",
        "banco_de_credito": "banco_credito",
        "banco_liquidacao": "banco_liquidacao",
        "banco_de_liquidacao": "banco_liquidacao",
        "data_contrato": "data_contrato",
        "data_do_contrato": "data_contrato",
        "data_de_contrato": "data_contrato",
        "data_recebimento": "data_recebimento",
        "data_de_recebimento": "data_recebimento",
        "data_liquidacao": "data_liquidacao",
        "data_de_liquidacao": "data_liquidacao",
        "cnpj": "cnpj",
        "cliente": "cliente",
        "moeda": "moeda",
        "valor_moeda": "valor_moeda",
        "valor_na_moeda": "valor_moeda",
        "taxa_cambio": "taxa_cambio",
        "taxa_de_cambio": "taxa_cambio",
        "valor_reais": "valor_reais",
        "valor_em_reais": "valor_reais",
    }
    normalized = []
    for column in columns:
        text = unicodedata.normalize("NFKD", str(column)).encode("ascii", "ignore").decode("ascii")
        key = re.sub(r"[^a-zA-Z0-9]+", "_", text.strip().lower()).strip("_")
        normalized.append(aliases.get(key, key))
    duplicates = sorted({column for column in normalized if normalized.count(column) > 1})
    if duplicates:
        raise ValueError("O Excel contém colunas duplicadas após a normalização: " + ", ".join(duplicates) + ".")
    return normalized

def launch_timestamp():
    """Gera no backend a data e hora do lançamento de uma nova DU-E."""
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")

def parse_number(value):
    if value is None or str(value).strip() == "":
        return 0.0
    text = str(value).strip().replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        raise ValueError("Valor inválido. Use o formato 1.234,56.")

def selected_record_ids(form):
    """Validate and deduplicate IDs sent by a batch deletion form."""
    raw_ids = form.getlist("selected_ids")
    if not raw_ids:
        raise ValueError("Selecione pelo menos um registro para excluir.")
    ids = []
    for raw_id in raw_ids:
        value = str(raw_id or "").strip()
        if not re.fullmatch(r"[1-9]\d*", value):
            raise ValueError("A seleção contém um identificador inválido.")
        record_id = int(value)
        if record_id not in ids:
            ids.append(record_id)
    return ids

def ensure_existing_record_ids(conn, table, ids, label):
    """Prevent partial deletion when one of the selected IDs no longer exists."""
    placeholders = ",".join("?" for _ in ids)
    found = {
        row[0] for row in conn.execute(
            f"SELECT id FROM {table} WHERE id IN ({placeholders})", ids
        ).fetchall()
    }
    if len(found) != len(ids):
        raise ValueError(f"Um ou mais {label} selecionados não foram encontrados.")

def redirect_batch_result(endpoint):
    """Preserve filters, ordering and pagination sent as hidden form fields."""
    args = {
        key: value for key, value in request.form.items()
        if key != "selected_ids" and value not in (None, "")
    }
    return redirect(url_for(endpoint, **args))

def parse_chave_acesso(value):
    chave = (value or "").strip().upper()
    if not chave:
        raise ValueError("A Chave de Acesso é obrigatória.")
    if not re.fullmatch(r"[A-Z0-9]{14}", chave):
        raise ValueError("A Chave de Acesso deve conter exatamente 14 caracteres alfanuméricos.")
    return chave

def normalize_cnpj(value):
    """Valida e armazena o CNPJ somente com seus 14 dígitos."""
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) != 14 or len(set(digits)) == 1:
        raise ValueError("Informe um CNPJ válido no formato 00.000.000/0000-00.")
    first_weights = (5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2)
    second_weights = (6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2)
    first = sum(int(digit) * weight for digit, weight in zip(digits[:12], first_weights))
    first_check = 0 if first % 11 < 2 else 11 - first % 11
    second = sum(int(digit) * weight for digit, weight in zip(digits[:12] + str(first_check), second_weights))
    second_check = 0 if second % 11 < 2 else 11 - second % 11
    if digits[-2:] != f"{first_check}{second_check}":
        raise ValueError("Informe um CNPJ válido no formato 00.000.000/0000-00.")
    return digits

def normalize_import_cnpj(value):
    """Normaliza CNPJs vindos do Excel, inclusive números sem zeros à esquerda."""
    if value is None:
        return None
    text_value = str(value).strip()
    if not text_value:
        return None
    # O Excel costuma converter CNPJs para número e o pandas pode expor o
    # valor como ``1234567890123.0``. Remova apenas a parte decimal nula.
    if re.fullmatch(r"\d+\.0+", text_value):
        text_value = text_value.split(".", 1)[0]
    if not re.fullmatch(r"[\d\s./-]+", text_value):
        return None
    digits = re.sub(r"\D", "", text_value)
    if len(digits) < 14:
        # CNPJs iniciados por zero perdem esse(s) zero(s) quando a célula é
        # numérica. O preenchimento só é aceito se o CNPJ passar a validação.
        digits = digits.zfill(14)
    return normalize_cnpj(digits)

def format_cnpj(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) != 14:
        return str(value or "")
    return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"

@app.template_filter("cnpj")
def cnpj_filter(value):
    return format_cnpj(value)

def empresas_for_form(conn, current_cnpj=None, selected_id=None):
    empresas = conn.execute("""
        SELECT id, razao_social, cnpj, apelido
        FROM empresas
        ORDER BY CASE WHEN TRIM(COALESCE(apelido, '')) <> '' THEN 0 ELSE 1 END,
                 apelido, razao_social
    """).fetchall()
    selected = None
    if selected_id not in (None, ""):
        try:
            selected = int(selected_id)
        except (TypeError, ValueError):
            selected = None
    if selected is None and current_cnpj:
        current_digits = re.sub(r"\D", "", str(current_cnpj))
        selected = next((empresa["id"] for empresa in empresas
                         if empresa["cnpj"] == current_digits), None)
    return empresas, selected

def cnpj_da_empresa(conn, empresa_id, current_cnpj=None):
    if empresa_id not in (None, ""):
        try:
            empresa_id = int(empresa_id)
        except (TypeError, ValueError):
            raise ValueError("Selecione uma empresa cadastrada.")
        empresa = conn.execute("SELECT cnpj FROM empresas WHERE id=?", (empresa_id,)).fetchone()
        if not empresa:
            raise ValueError("A empresa selecionada não foi encontrada.")
        return normalize_cnpj(empresa["cnpj"])
    if current_cnpj:
        return normalize_cnpj(current_cnpj)
    raise ValueError("Selecione uma empresa em Configurações > Cadastrar minhas empresas.")

COMPETENCIA_STATUS_ABERTA = "ABERTA"
COMPETENCIA_STATUS_ENCERRADA = "ENCERRADA"
COMPETENCIA_STATUSES = (COMPETENCIA_STATUS_ABERTA, COMPETENCIA_STATUS_ENCERRADA)

def competencia_data(form, conn, current=None):
    descricao = (form.get("descricao") or "").strip()
    if not descricao:
        raise ValueError("A descrição da competência é obrigatória.")
    try:
        empresa_id = int(form.get("empresa_id"))
    except (TypeError, ValueError):
        empresa_id = current["empresa_id"] if current else None
    if not empresa_id or not conn.execute("SELECT id FROM empresas WHERE id=?", (empresa_id,)).fetchone():
        raise ValueError("Selecione uma empresa cadastrada.")
    data_inicial = parse_date(form.get("data_inicial"))
    data_final = parse_date(form.get("data_final"))
    if not data_inicial or not data_final:
        raise ValueError("As datas inicial e final são obrigatórias.")
    if data_inicial > data_final:
        raise ValueError("A data inicial não pode ser posterior à data final.")
    status = (form.get("status") or COMPETENCIA_STATUS_ABERTA).strip().upper()
    if status not in COMPETENCIA_STATUSES:
        raise ValueError("Selecione um status válido para a competência.")
    return {"empresa_id": empresa_id, "descricao": descricao, "data_inicial": data_inicial, "data_final": data_final, "status": status}

def salvar_competencia(conn, data, competencia_id=None):
    """Persiste uma competência em transação própria e confirma a linha gravada."""
    conn.execute("BEGIN IMMEDIATE")
    try:
        values = (
            data["empresa_id"], data["descricao"], data["data_inicial"],
            data["data_final"], data["status"],
        )
        if competencia_id:
            cursor = conn.execute("""UPDATE competencias
                SET empresa_id=?, descricao=?, data_inicial=?, data_final=?, status=?
                WHERE id=?""", values + (competencia_id,))
            if cursor.rowcount != 1:
                raise ValueError("A competência selecionada não foi encontrada.")
            saved_id = competencia_id
        else:
            cursor = conn.execute("""INSERT INTO competencias
                (empresa_id, descricao, data_inicial, data_final, status)
                VALUES (?,?,?,?,?)""", values)
            saved_id = cursor.lastrowid
        saved = conn.execute("""SELECT id, empresa_id, descricao, data_inicial, data_final, status
            FROM competencias WHERE id=?""", (saved_id,)).fetchone()
        if not saved:
            raise sqlite3.DatabaseError("A competência não foi encontrada após a gravação.")
        conn.commit()
        return saved
    except Exception:
        conn.rollback()
        raise

def sugerir_competencia(conn, empresa_id, data_recebimento):
    """Busca uma competência aberta que contenha a data de recebimento.

    Contratos ainda não possuem data_recebimento nem competencia_id; a função
    fica pronta para a integração futura sem alterar a regra atual.
    """
    data = parse_date(data_recebimento)
    if not data or not empresa_id:
        return None
    return conn.execute("""SELECT id, empresa_id, descricao, data_inicial, data_final, status
        FROM competencias WHERE empresa_id=? AND status=? AND data_inicial <= ? AND data_final >= ?
        ORDER BY data_inicial DESC, id DESC LIMIT 1""",
        (empresa_id, COMPETENCIA_STATUS_ABERTA, data, data)).fetchone()

def empresa_id_por_cnpj(conn, cnpj):
    if not cnpj:
        return None
    row = conn.execute("SELECT id FROM empresas WHERE cnpj=?", (re.sub(r"\D", "", str(cnpj)),)).fetchone()
    return row["id"] if row else None

def competencia_da_operacao(conn, raw_id, empresa_id, data_referencia, current_id=None):
    """Valida a competência da operação e sugere uma aberta pela data."""
    competencia_id = form_record_id(raw_id, current_id)
    if not competencia_id:
        sugerida = sugerir_competencia(conn, empresa_id, data_referencia)
        competencia_id = sugerida["id"] if sugerida else None
    if not competencia_id:
        raise ValueError("Selecione uma competência para a operação.")
    competencia = conn.execute("SELECT id FROM competencias WHERE id=? AND empresa_id=?", (competencia_id, empresa_id)).fetchone()
    if not competencia:
        raise ValueError("A competência selecionada não pertence à empresa da operação.")
    return competencia_id

def competencias_for_empresa(conn, empresa_id):
    if empresa_id:
        return conn.execute("""SELECT c.id, c.empresa_id, c.descricao, c.data_inicial, c.data_final, c.status,
            e.razao_social, e.apelido AS empresa_apelido
            FROM competencias c JOIN empresas e ON e.id=c.empresa_id
            WHERE c.empresa_id=? ORDER BY c.data_inicial DESC, c.descricao""", (empresa_id,)).fetchall()
    return conn.execute("""SELECT c.id, c.empresa_id, c.descricao, c.data_inicial, c.data_final, c.status,
        e.razao_social, e.apelido AS empresa_apelido
        FROM competencias c JOIN empresas e ON e.id=c.empresa_id
        ORDER BY c.data_inicial DESC, c.descricao""").fetchall()

COMPETENCIA_IMPORT_PROXIMIDADE_DIAS = 31
COMPETENCIA_MESES = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
    "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

def normalize_competencia_import(value):
    text_value = " ".join(str(value or "").split()).strip()
    if any(unicodedata.category(character) == "Cc" for character in text_value):
        raise ValueError("A competência não pode conter caracteres de controle.")
    if len(text_value) > 120:
        raise ValueError("A competência deve ter no máximo 120 caracteres.")
    return text_value or None

def competencia_import_key(value):
    text_value = normalize_competencia_import(value) or ""
    text_value = unicodedata.normalize("NFKD", text_value)
    text_value = "".join(character for character in text_value if not unicodedata.combining(character))
    text_value = re.sub(r"[^a-zA-Z0-9]+", " ", text_value.casefold())
    return " ".join(text_value.split())

def competencia_import_period(value):
    """Interpreta competências mensais comuns para comparar períodos."""
    key = competencia_import_key(value)
    if not key:
        return None
    month = year = None
    match = re.search(r"\b(20\d{2})[\s._/-]+(0?[1-9]|1[0-2])\b", key)
    if match:
        year, month = int(match.group(1)), int(match.group(2))
    else:
        match = re.search(r"\b(0?[1-9]|1[0-2])[\s._/-]+(20\d{2})\b", key)
        if match:
            month, year = int(match.group(1)), int(match.group(2))
    if month is None:
        for name, number in COMPETENCIA_MESES.items():
            if re.search(rf"\b{name}\b", key):
                month = number
                break
        match = re.search(r"\b(20\d{2})\b", key)
        if month is not None and match:
            year = int(match.group(1))
    if not month or not year:
        return None
    start = date(year, month, 1)
    next_month = date(year + (month == 12), 1 if month == 12 else month + 1, 1)
    return start.isoformat(), (next_month - timedelta(days=1)).isoformat()

def competencia_import_distance(reference_date, competencia):
    if not reference_date:
        return None
    reference = date.fromisoformat(reference_date)
    start = date.fromisoformat(competencia["data_inicial"])
    end = date.fromisoformat(competencia["data_final"])
    if start <= reference <= end:
        return 0
    return (start - reference).days if reference < start else (reference - end).days

def resolve_import_competencia(conn, empresa_id, descricao, data_emissao=None):
    """Retorna uma competência segura, inclusive quando a descrição varia."""
    competencias = conn.execute("""SELECT id, empresa_id, descricao, data_inicial, data_final, status
        FROM competencias WHERE empresa_id=? ORDER BY data_inicial DESC, id DESC""", (empresa_id,)).fetchall()
    key = competencia_import_key(descricao)
    exact = [item for item in competencias if competencia_import_key(item["descricao"]) == key]
    if exact:
        return exact[0]
    requested_period = competencia_import_period(descricao)
    relevant = []
    for item in competencias:
        distance = competencia_import_distance(data_emissao, item) if data_emissao else None
        period_match = bool(requested_period and not (
            requested_period[1] < item["data_inicial"] or requested_period[0] > item["data_final"]
        ))
        if period_match or distance == 0:
            relevant.append((0, 0 if item["status"] == COMPETENCIA_STATUS_ABERTA else 1, item))
            continue
        if distance is not None and distance <= COMPETENCIA_IMPORT_PROXIMIDADE_DIAS:
            relevant.append((distance, 0 if item["status"] == COMPETENCIA_STATUS_ABERTA else 1, item))
    return min(relevant, key=lambda value: (value[0], value[1], -value[2]["id"]))[2] if relevant else None

def default_import_competencia_period(rows):
    dates = sorted(row.get("data_emissao") for row in rows if row.get("data_emissao"))
    raw_period = competencia_import_period(rows[0].get("competencia")) if rows else None
    if raw_period:
        return raw_period
    if dates:
        first = date.fromisoformat(dates[0])
        last = date.fromisoformat(dates[-1])
        return first.isoformat(), last.isoformat()
    today = date.today()
    next_month = date(today.year + (today.month == 12), 1 if today.month == 12 else today.month + 1, 1)
    return today.replace(day=1).isoformat(), (next_month - timedelta(days=1)).isoformat()

def resolve_invoice_import_competencies(conn, rows):
    """Vincula competências existentes e prepara sugestões por empresa."""
    groups = {}
    for row in rows:
        empresa = conn.execute("SELECT id, razao_social, apelido, cnpj FROM empresas WHERE cnpj=?",
                               (row["cnpj"],)).fetchone()
        if not empresa:
            raise ValueError(f"A empresa com CNPJ {row['cnpj']} não está cadastrada.")
        row["empresa_id"] = empresa["id"]
        row["competencia"] = normalize_competencia_import(row.get("competencia"))
        key = f"{empresa['id']}|{competencia_import_key(row['competencia'])}"
        row["competencia_import_key"] = key
        groups.setdefault(key, []).append(row)
    suggestions = []
    for key, group in groups.items():
        first = group[0]
        competencia = resolve_import_competencia(
            conn, first["empresa_id"], first["competencia"], first.get("data_emissao")
        )
        if competencia:
            for row in group:
                row["competencia_id"] = competencia["id"]
            continue
        data_inicial, data_final = default_import_competencia_period(group)
        suggestions.append({
            "suggestion_id": f"p{len(suggestions) + 1}", "key": key,
            "empresa_id": first["empresa_id"], "empresa_nome": first.get("empresa_nome") or "",
            "empresa_apelido": first.get("empresa_apelido") or "", "cnpj": first["cnpj"],
            "descricao": first["competencia"], "data_inicial": data_inicial,
            "data_final": data_final, "linhas": [row["source_row"] for row in group],
        })
    for suggestion in suggestions:
        empresa = conn.execute("SELECT razao_social, apelido FROM empresas WHERE id=?",
                               (suggestion["empresa_id"],)).fetchone()
        suggestion["empresa_nome"] = empresa["razao_social"] if empresa else ""
        suggestion["empresa_apelido"] = empresa["apelido"] if empresa else ""
    return suggestions

def ensure_invoice_import_competencies(conn, rows, new_competency_overrides=None):
    """Usa competências resolvidas na prévia ou cadastra as aprovadas pelo usuário."""
    overrides = new_competency_overrides or {}
    grouped = {}
    for row in rows:
        grouped.setdefault(row.get("competencia_import_key"), []).append(row)
    for key, group in grouped.items():
        first = group[0]
        competencia_id = first.get("competencia_id")
        if competencia_id:
            valid = conn.execute("SELECT id FROM competencias WHERE id=? AND empresa_id=?",
                                 (competencia_id, first["empresa_id"])).fetchone()
            if not valid:
                raise ValueError("Uma competência selecionada não pertence mais à empresa da Invoice.")
            for row in group:
                row["competencia_id"] = competencia_id
            continue
        override = overrides.get(key)
        if not override:
            raise ValueError(
                f"A competência {first['competencia']} não está cadastrada para a empresa. "
                "Confirme o cadastro sugerido na prévia."
            )
        descricao = normalize_competencia_import(override.get("descricao"))
        data_inicial = parse_date(override.get("data_inicial"))
        data_final = parse_date(override.get("data_final"))
        if not descricao or not data_inicial or not data_final:
            raise ValueError(f"Preencha descrição e período da nova competência {first['competencia']}.")
        if data_inicial > data_final:
            raise ValueError("A data inicial da competência não pode ser posterior à data final.")
        existing = conn.execute("""SELECT id FROM competencias WHERE empresa_id=? AND descricao=? COLLATE NOCASE""",
                                (first["empresa_id"], descricao)).fetchone()
        if existing:
            competencia_id = existing["id"]
        else:
            try:
                competencia_id = conn.execute("""INSERT INTO competencias
                    (empresa_id, descricao, data_inicial, data_final, status)
                    VALUES (?,?,?,?,?)""", (first["empresa_id"], descricao, data_inicial, data_final,
                                             COMPETENCIA_STATUS_ABERTA)).lastrowid
            except sqlite3.IntegrityError:
                existing = conn.execute("""SELECT id FROM competencias
                    WHERE empresa_id=? AND descricao=? COLLATE NOCASE""", (first["empresa_id"], descricao)).fetchone()
                if not existing:
                    raise
                competencia_id = existing["id"]
        for row in group:
            row["competencia_id"] = competencia_id
    return grouped

def resolve_invoice_import_companies(conn, rows):
    """Converte a coluna Empresa (CNPJ, razão social ou apelido) para CNPJ."""
    companies = conn.execute("SELECT id, razao_social, apelido, cnpj FROM empresas ORDER BY id").fetchall()
    for row in rows:
        if row.get("cnpj"):
            company = next((item for item in companies if item["cnpj"] == row["cnpj"]), None)
            if not company:
                raise ValueError(f"A empresa com CNPJ {row['cnpj']} não está cadastrada.")
            row["empresa_id"] = company["id"]
            continue
        company_text = normalize_client_name_display(row.get("empresa"))
        company_key = normalize_client_name_key(company_text)
        candidates = [item for item in companies if company_key in {
            normalize_client_name_key(item["razao_social"]),
            normalize_client_name_key(item["apelido"]),
        }]
        if not candidates:
            raise ValueError(
                f"A empresa {company_text or '-'} não está cadastrada. "
                "Informe o CNPJ ou o nome/apelido exatamente como cadastrado."
            )
        if len(candidates) > 1:
            raise ValueError(f"A empresa {company_text} corresponde a mais de uma empresa cadastrada.")
        company = candidates[0]
        row["cnpj"] = company["cnpj"]
        row["empresa_id"] = company["id"]
    return rows

@app.template_filter("pais_nome")
def pais_nome(value):
    codigo = str(value or "").strip().upper()
    return CLIENTE_PAISES_MAP.get(codigo, value or "-")

def normalize_pais(value):
    pais = (value or "").strip().upper()
    if pais not in CLIENTE_PAISES_MAP:
        raise ValueError("Selecione um país válido.")
    return pais

def normalize_client_name_display(value):
    """Normaliza apenas a apresentação do nome, preservando sua grafia."""
    return " ".join(str(value or "").split()).strip() or None

def normalize_client_name_key(value):
    """Gera uma chave tolerante a acentos, caixa, pontuação e espaços."""
    text_value = normalize_client_name_display(value) or ""
    text_value = unicodedata.normalize("NFKD", text_value)
    text_value = "".join(character for character in text_value if not unicodedata.combining(character))
    text_value = "".join(character if character.isalnum() else " "
                           for character in text_value.casefold())
    return " ".join(text_value.split())

def normalize_import_client_country(value):
    """Aceita código ISO ou nome do país na coluna opcional da planilha."""
    text_value = normalize_client_name_display(value)
    if not text_value:
        return None
    code = text_value.upper()
    if code in CLIENTE_PAISES_MAP:
        return code
    country_key = normalize_client_name_key(text_value)
    for country_code, country_name in CLIENTE_PAISES:
        if normalize_client_name_key(country_name) == country_key:
            return country_code
    raise ValueError(f"País do cliente inválido: {text_value}.")

def import_client_key(name, country=None):
    name_key = normalize_client_name_key(name)
    if not name_key:
        return None
    return f"{name_key}|{(country or '').upper()}"

def import_client_candidates(conn, name):
    name_key = normalize_client_name_key(name)
    if not name_key:
        return []
    clients = conn.execute("SELECT id, nome, pais FROM clientes ORDER BY id").fetchall()
    exact = [client for client in clients if normalize_client_name_key(client["nome"]) == name_key]
    if exact:
        return exact
    name_parts = name_key.split()
    if len(name_parts) < 2:
        return []
    prefix = " ".join(name_parts[:2])
    return [client for client in clients
            if " ".join(normalize_client_name_key(client["nome"]).split()[:2]) == prefix]

def resolve_import_client(conn, name, country=None):
    """Retorna um cliente somente quando a correspondência é segura."""
    candidates = import_client_candidates(conn, name)
    if country:
        candidates = [client for client in candidates if client["pais"] == country]
    if not candidates:
        return None, []
    countries = {client["pais"] for client in candidates}
    if len(countries) == 1:
        return min(candidates, key=lambda client: client["id"]), candidates
    return None, candidates

def resolve_invoice_import_clients(conn, rows):
    """Anexa IDs encontrados e devolve sugestões para nomes não reconhecidos."""
    groups = {}
    for row in rows:
        name = normalize_client_name_display(row.get("cliente"))
        country = row.get("cliente_pais")
        row["cliente"] = name
        row["cliente_import_key"] = import_client_key(name, country)
        if name:
            groups.setdefault(row["cliente_import_key"], []).append(row)
    suggestions = []
    for key, group in groups.items():
        first = group[0]
        client, candidates = resolve_import_client(conn, first["cliente"], first.get("cliente_pais"))
        if client:
            for row in group:
                row["cliente_id"] = client["id"]
            continue
        suggestions.append({
            "suggestion_id": f"c{len(suggestions) + 1}",
            "key": key,
            "nome": first["cliente"],
            "pais": first.get("cliente_pais"),
            "linhas": [row["source_row"] for row in group],
            "candidatos": [{"nome": candidate["nome"], "pais": candidate["pais"]}
                           for candidate in candidates],
        })
    return suggestions

def ensure_invoice_import_clients(conn, rows, country_overrides=None, client_overrides=None):
    """Resolve clientes, usando a seleção da prévia ou criando os novos confirmados."""
    country_overrides = country_overrides or {}
    client_overrides = client_overrides or {}
    groups = {}
    for row in rows:
        if row.get("cliente"):
            groups.setdefault(row.get("cliente_import_key") or import_client_key(
                row["cliente"], row.get("cliente_pais")), []).append(row)
    for key, group in groups.items():
        first = group[0]
        selected_client_id = client_overrides.get(key)
        if selected_client_id not in (None, ""):
            try:
                selected_client_id = int(selected_client_id)
            except (TypeError, ValueError):
                raise ValueError("O cliente selecionado na prévia é inválido.")
            client = conn.execute("SELECT id, nome, pais FROM clientes WHERE id=?",
                                  (selected_client_id,)).fetchone()
            if not client:
                raise ValueError("O cliente selecionado na prévia não foi encontrado.")
            for row in group:
                row["cliente_id"] = client["id"]
                row["cliente"] = client["nome"]
            continue
        country = country_overrides.get(key) or first.get("cliente_pais")
        client, _ = resolve_import_client(conn, first["cliente"], country)
        if not client:
            if not country:
                raise ValueError(
                    f"O cliente {first['cliente']} não foi identificado. "
                    "Informe o País na prévia para sugerir seu cadastro."
                )
            try:
                conn.execute("INSERT INTO clientes (nome, pais) VALUES (?, ?)", (first["cliente"], country))
            except sqlite3.IntegrityError:
                pass
            client, _ = resolve_import_client(conn, first["cliente"], country)
            if not client:
                raise ValueError(f"Não foi possível associar ou cadastrar o cliente {first['cliente']}.")
        for row in group:
            row["cliente_id"] = client["id"]
            row["cliente"] = client["nome"]
    return groups

def clientes_for_form(conn):
    return conn.execute("""
        SELECT id, nome, pais
        FROM clientes
        ORDER BY nome, pais
    """).fetchall()

def contrapartes_for_form(conn):
    return conn.execute("""
        SELECT id, nome
        FROM contrapartes
        ORDER BY nome
    """).fetchall()

def form_record_id(value, fallback=None):
    if value in (None, ""):
        return fallback
    try:
        return int(value)
    except (TypeError, ValueError):
        return None

def cliente_da_selecao(conn, raw_id, current_id=None, required=False):
    if raw_id in (None, ""):
        if current_id:
            cliente = conn.execute("SELECT id, nome, pais FROM clientes WHERE id=?", (current_id,)).fetchone()
            if cliente:
                return cliente
        if required:
            raise ValueError("Selecione um cliente cadastrado em Configurações.")
        return None
    try:
        cliente_id = int(raw_id)
    except (TypeError, ValueError):
        raise ValueError("Selecione um cliente cadastrado.")
    cliente = conn.execute("SELECT id, nome, pais FROM clientes WHERE id=?", (cliente_id,)).fetchone()
    if not cliente:
        raise ValueError("O cliente selecionado não foi encontrado.")
    return cliente

def contraparte_da_selecao(conn, raw_id, current_id=None, required=False):
    if raw_id in (None, ""):
        if current_id:
            contraparte = conn.execute("SELECT id, nome FROM contrapartes WHERE id=?", (current_id,)).fetchone()
            if contraparte:
                return contraparte
        if required:
            raise ValueError("Selecione um Banco / Contraparte cadastrado em Configurações.")
        return None
    try:
        contraparte_id = int(raw_id)
    except (TypeError, ValueError):
        raise ValueError("Selecione um Banco / Contraparte cadastrado.")
    contraparte = conn.execute("SELECT id, nome FROM contrapartes WHERE id=?", (contraparte_id,)).fetchone()
    if not contraparte:
        raise ValueError("O Banco / Contraparte selecionado não foi encontrado.")
    return contraparte

def banco_do_contrato(conn, raw_id, current=None):
    current_id = current["banco_id"] if current and "banco_id" in current.keys() else None
    contraparte = contraparte_da_selecao(conn, raw_id, current_id=current_id)
    if contraparte:
        return contraparte["id"], contraparte["nome"]
    return current_id, (current["banco"] if current else None)

def banco_id_for_form(conn, contrato=None, selected_id=None):
    if selected_id not in (None, ""):
        return form_record_id(selected_id)
    if not contrato:
        return None
    if "banco_id" in contrato.keys() and contrato["banco_id"]:
        return contrato["banco_id"]
    if contrato["banco"]:
        registro = conn.execute(
            "SELECT id FROM contrapartes WHERE nome=? COLLATE NOCASE",
            (contrato["banco"],),
        ).fetchone()
        return registro["id"] if registro else None
    return None

def banco_ids_for_contrato_form(conn, contrato=None, credito_id=None, liquidacao_id=None):
    """Resolve os dois bancos para o formulário, mantendo contratos legados."""
    credito_id = banco_id_for_form(conn, contrato, credito_id)
    credito_nome = ((contrato["banco_credito"] if "banco_credito" in contrato.keys() else None) if contrato else None) or ((contrato["banco"] if contrato and "banco" in contrato.keys() else None) if contrato else None)
    liquidacao_nome = ((contrato["banco_liquidacao"] if "banco_liquidacao" in contrato.keys() else None) if contrato else None) or credito_nome
    if liquidacao_id in (None, "") and liquidacao_nome:
        row = conn.execute("SELECT id FROM contrapartes WHERE nome=? COLLATE NOCASE", (liquidacao_nome,)).fetchone()
        liquidacao_id = row["id"] if row else None
    return credito_id, liquidacao_id

def bancos_do_contrato(conn, form, current=None):
    """Obtém crédito/liquidação sem perder o banco legado nem uma escolha manual."""
    credito_raw = form.get("banco_credito_id")
    if credito_raw in (None, ""):
        credito_raw = form.get("banco_id")
    credito_id, credito = banco_do_contrato(conn, credito_raw, current=current)
    anterior_credito = ((current["banco_credito"] if "banco_credito" in current.keys() else None) or
                        (current["banco"] if current else None)) if current else None
    anterior_liquidacao = ((current["banco_liquidacao"] if "banco_liquidacao" in current.keys() else None) or
                           anterior_credito) if current else None
    liquidacao_raw = form.get("banco_liquidacao_id")
    if liquidacao_raw in (None, ""):
        liquidacao_id, liquidacao = credito_id, credito
    else:
        liquidacao_id, liquidacao = banco_do_contrato(conn, liquidacao_raw)
        if current and anterior_liquidacao == anterior_credito and liquidacao == anterior_credito and credito != anterior_credito:
            liquidacao_id, liquidacao = credito_id, credito
    return credito_id, credito, liquidacao_id, liquidacao

def cliente_do_contrato(conn, raw_id, current=None, legacy_name=None):
    current_id = current["cliente_id"] if current and "cliente_id" in current.keys() else None
    cliente = cliente_da_selecao(conn, raw_id, current_id=current_id)
    if cliente:
        return cliente["id"], cliente["nome"]
    if legacy_name is not None:
        return current_id, (legacy_name or "").strip() or None
    return current_id, (current["cliente"] if current else None)

def cliente_id_for_form(contrato=None, selected_id=None):
    if selected_id not in (None, ""):
        return form_record_id(selected_id)
    if contrato and "cliente_id" in contrato.keys() and contrato["cliente_id"]:
        return contrato["cliente_id"]
    return None

def decimal_value(value):
    """Converte valores monetários para Decimal sem perder a regra de tolerância."""
    if value is None or str(value).strip() in {"", "nan", "NaT"}:
        return Decimal("0")
    return Decimal(str(value))

def normalize_balance(balance):
    balance = decimal_value(balance)
    return Decimal("0") if abs(balance) <= SALDO_TOLERANCE else balance

def status_from_balance(balance, amortized=0):
    """Calcula o status pelo saldo restante e pelo valor já amortizado/vinculado."""
    saldo = normalize_balance(balance)
    if saldo <= 0:
        return STATUS_CONCLUIDO
    return STATUS_PARCIAL if normalize_balance(amortized) > 0 else STATUS_PENDENTE

@app.template_filter("status_class")
def status_class(status):
    if status == STATUS_PARCIAL:
        return "status-parcial"
    return "status-pendente" if status == STATUS_PENDENTE else "status-concluido"

def movement_effect_sql(alias="m"):
    return f"CASE WHEN {alias}.tipo IN ('UTILIZACAO','VINCULACAO') THEN {alias}.valor ELSE -{alias}.valor END"

def due_effect(conn, due_id):
    return conn.execute(f"""SELECT COALESCE(SUM({movement_effect_sql()}),0)
                            FROM due_movimentacoes m WHERE m.due_id=?""", (due_id,)).fetchone()[0]

def due_balance(valor_original, utilizado):
    return normalize_balance(decimal_value(valor_original) - decimal_value(utilizado))

def contract_balance(valor_moeda, vinculado):
    return normalize_balance(decimal_value(valor_moeda) - decimal_value(vinculado))

def ensure_non_negative_balance(balance, message):
    if decimal_value(balance) < -SALDO_TOLERANCE:
        raise ValueError(message)

def update_due_status(conn, due_id):
    row = conn.execute(f"""
        SELECT d.valor_original,
               COALESCE(SUM({movement_effect_sql()}), 0) AS utilizado
        FROM dues d
        LEFT JOIN due_movimentacoes m ON m.due_id=d.id
        WHERE d.id=?
        GROUP BY d.id
    """, (due_id,)).fetchone()
    if not row:
        return None
    saldo = due_balance(row["valor_original"], row["utilizado"])
    status = status_from_balance(saldo, row["utilizado"])
    conn.execute("UPDATE dues SET status=? WHERE id=?", (status, due_id))
    return status

def update_contract_status(conn, contrato_id):
    row = conn.execute("""
        SELECT c.valor_moeda, c.saldo_zerado_manual,
               COALESCE((SELECT SUM(v.valor_alocado) FROM invoice_contrato_cambio v
                         WHERE v.contrato_id=c.id), c.valor_moeda) AS valor_moeda_consolidado,
               COALESCE(SUM(CASE WHEN m.tipo='VINCULACAO' THEN m.valor ELSE 0 END), 0) AS vinculado
        FROM contratos c
        LEFT JOIN due_movimentacoes m ON m.contrato_id=c.id
        WHERE c.id=?
        GROUP BY c.id
    """, (contrato_id,)).fetchone()
    if not row:
        return None
    total = decimal_value(row["valor_moeda_consolidado"])
    saldo = Decimal("0") if row["saldo_zerado_manual"] else contract_balance(total, row["vinculado"])
    status = STATUS_CONCLUIDO if row["saldo_zerado_manual"] else status_from_balance(saldo, row["vinculado"])
    conn.execute("UPDATE contratos SET status=? WHERE id=?", (status, contrato_id))
    return status

def recalculate_statuses(conn, due_ids=None, contrato_ids=None):
    if due_ids is None:
        due_ids = [row[0] for row in conn.execute("SELECT id FROM dues")]
    if contrato_ids is None:
        contrato_ids = [row[0] for row in conn.execute("SELECT id FROM contratos")]
    for due_id in set(due_ids):
        update_due_status(conn, due_id)
    for contrato_id in set(contrato_ids):
        update_contract_status(conn, contrato_id)

def decorate_due(row):
    data = dict(row)
    data["valor_original"] = decimal_value(data.get("valor_original"))
    data["utilizado"] = decimal_value(data.get("utilizado"))
    data["saldo"] = due_balance(data.get("valor_original"), data["utilizado"])
    data["status"] = status_from_balance(data["saldo"], data["utilizado"])
    return data

def decorate_contract(row):
    data = dict(row)
    if data.get("valor_moeda_consolidado") is not None:
        data["valor_moeda"] = data["valor_moeda_consolidado"]
    data["valor_moeda"] = decimal_value(data.get("valor_moeda"))
    data["vinculado"] = decimal_value(data.get("vinculado"))
    data["saldo_zerado_manual"] = bool(data.get("saldo_zerado_manual"))
    data["saldo"] = Decimal("0") if data["saldo_zerado_manual"] else contract_balance(data.get("valor_moeda"), data["vinculado"])
    data["status"] = STATUS_CONCLUIDO if data["saldo_zerado_manual"] else status_from_balance(data["saldo"], data["vinculado"])
    return data

@app.template_filter("ndf_status_class")
def ndf_status_class(status):
    return {
        NDF_STATUS_ATIVA: "status-ndf-ativa",
        NDF_STATUS_VENCIDA: "status-ndf-vencida",
        NDF_STATUS_LIQUIDADA: "status-ndf-liquidada",
        NDF_STATUS_CANCELADA: "status-ndf-cancelada",
    }.get(status, "status-ndf-cancelada")

def ndf_display_status(status, data_vencimento):
    if status == NDF_STATUS_ATIVA and data_vencimento and data_vencimento < date.today().isoformat():
        return NDF_STATUS_VENCIDA
    return status

def decorate_ndf(row):
    data = dict(row)
    data["valor_contratado"] = decimal_value(data.get("valor_contratado"))
    data["taxa_contratada"] = decimal_value(data.get("taxa_contratada"))
    data["status_exibicao"] = ndf_display_status(data.get("status"), data.get("data_vencimento"))
    return data

def parse_ndf_form(form, conn, current=None):
    numero_operacao = (form.get("numero_operacao") or "").strip()
    if not numero_operacao:
        raise ValueError("O número/ID da operação é obrigatório.")

    cnpj = cnpj_da_empresa(conn, form.get("empresa_id"), current["cnpj"] if current else None)
    empresa_id = form_record_id(form.get("empresa_id"), empresa_id_por_cnpj(conn, cnpj))
    current_cliente_id = current["cliente_id"] if current and "cliente_id" in current.keys() else None
    current_contraparte_id = current["contraparte_id"] if current and "contraparte_id" in current.keys() else None
    cliente = cliente_da_selecao(
        conn, form.get("cliente_id"), current_id=current_cliente_id, required=current is None
    )
    contraparte_registro = contraparte_da_selecao(
        conn, form.get("contraparte_id"), current_id=current_contraparte_id, required=current is None
    )
    contraparte = contraparte_registro["nome"] if contraparte_registro else (current["contraparte"] if current else "")
    if not contraparte:
        raise ValueError("A contraparte é obrigatória.")

    tipo = (form.get("tipo") or "").strip().upper()
    if tipo not in NDF_TIPOS:
        raise ValueError("Selecione um tipo válido: Banco ou Trading.")

    moeda = (form.get("moeda") or "").strip().upper()
    if not re.fullmatch(r"[A-Z]{3}", moeda):
        raise ValueError("A moeda deve conter exatamente 3 letras.")

    valor_bruto = form.get("valor_contratado")
    if not str(valor_bruto or "").strip():
        raise ValueError("O valor contratado é obrigatório.")
    valor_contratado = parse_number(valor_bruto)
    ensure_non_negative_balance(valor_contratado, "O valor contratado não pode ser negativo.")

    taxa_bruta = form.get("taxa_contratada")
    if not str(taxa_bruta or "").strip():
        raise ValueError("A taxa contratada é obrigatória.")
    taxa_contratada = parse_number(taxa_bruta)
    ensure_non_negative_balance(taxa_contratada, "A taxa contratada não pode ser negativa.")

    data_contratacao = parse_date(form.get("data_contratacao"))
    data_vencimento = parse_date(form.get("data_vencimento"))
    if not data_contratacao or not data_vencimento:
        raise ValueError("As datas de contratação e vencimento são obrigatórias.")
    if data_vencimento < data_contratacao:
        raise ValueError("A data de vencimento não pode ser anterior à data de contratação.")
    competencia_id = competencia_da_operacao(
        conn, form.get("competencia_id"), empresa_id, data_contratacao,
        current["competencia_id"] if current and "competencia_id" in current.keys() else None,
    )

    posicao = (form.get("posicao") or "").strip().upper()
    if posicao not in NDF_POSICOES:
        raise ValueError("Selecione uma posição válida: Compra ou Venda.")

    status = (form.get("status") or "").strip().upper()
    if status not in NDF_STATUSES:
        raise ValueError("Selecione um status válido.")

    return {
        "numero_operacao": numero_operacao,
        "cnpj": cnpj,
        "competencia_id": competencia_id,
        "cliente_id": cliente["id"] if cliente else None,
        "contraparte_id": contraparte_registro["id"] if contraparte_registro else None,
        "contraparte": contraparte,
        "tipo": tipo,
        "moeda": moeda,
        "valor_contratado": valor_contratado,
        "taxa_contratada": taxa_contratada,
        "data_contratacao": data_contratacao,
        "data_vencimento": data_vencimento,
        "posicao": posicao,
        "finalidade": (form.get("finalidade") or "").strip() or None,
        "observacao": (form.get("observacao") or "").strip() or None,
        "status": status,
    }

def parse_ptax_period(form):
    moeda = (form.get("moeda") or "").strip().upper()
    if moeda not in PTAX_MOEDAS:
        raise ValueError("Selecione uma moeda válida: USD ou EUR.")

    data_inicial = parse_date(form.get("data_inicial"))
    data_final = parse_date(form.get("data_final"))
    if not data_inicial or not data_final:
        raise ValueError("As datas inicial e final são obrigatórias.")
    if data_inicial > data_final:
        raise ValueError("A data inicial não pode ser posterior à data final.")
    return moeda, data_inicial, data_final

def ptax_api_url(moeda, data_inicial, data_final):
    parametros = urlencode({
        "@moeda": f"'{moeda}'",
        "@dataInicial": f"'{datetime.strptime(data_inicial, '%Y-%m-%d').strftime('%m-%d-%Y')}'",
        "@dataFinalCotacao": f"'{datetime.strptime(data_final, '%Y-%m-%d').strftime('%m-%d-%Y')}'",
        "$format": "json",
    })
    return (
        f"{PTAX_API_URL}(moeda=@moeda,dataInicial=@dataInicial,"
        f"dataFinalCotacao=@dataFinalCotacao)?{parametros}"
    )

def ptax_rate_value(value):
    if value is None or str(value).strip() in {"", "nan", "NaT", "None"}:
        return None
    try:
        rate_value = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    if not rate_value.is_finite() or rate_value <= 0:
        return None
    return float(rate_value)

def parse_ptax_rate(value, label):
    if not str(value or "").strip():
        raise ValueError(f"A {label} é obrigatória.")
    rate_value = ptax_rate_value(parse_number(value))
    if rate_value is None:
        raise ValueError(f"A {label} retornada é inválida.")
    return rate_value

def consultar_ptax_api(moeda, data_inicial, data_final):
    url = ptax_api_url(moeda, data_inicial, data_final)
    requisicao = Request(url, headers={
        "Accept": "application/json",
        "User-Agent": "DUE-Control PTAX",
    })
    try:
        with urlopen(requisicao, timeout=PTAX_API_TIMEOUT) as resposta:
            payload = json.loads(resposta.read().decode("utf-8"))
    except HTTPError as exc:
        raise ValueError(f"O Banco Central respondeu com erro HTTP {exc.code}.")
    except (URLError, TimeoutError, OSError):
        raise ValueError("Não foi possível acessar a API PTAX do Banco Central.")
    except (UnicodeDecodeError, json.JSONDecodeError):
        raise ValueError("A API PTAX retornou uma resposta inválida.")

    registros_api = payload.get("value") if isinstance(payload, dict) else None
    if not isinstance(registros_api, list):
        raise ValueError("A API PTAX retornou um formato de dados inválido.")

    fechamentos = {}
    for registro in registros_api:
        if not isinstance(registro, dict):
            continue
        tipo_boletim = str(registro.get("tipoBoletim") or "").strip().casefold()
        if tipo_boletim != "fechamento":
            continue
        data_hora = str(registro.get("dataHoraCotacao") or "").strip()
        if not data_hora:
            continue
        data_cotacao = data_hora[:10]
        try:
            data_cotacao = parse_date(data_cotacao)
        except ValueError:
            continue
        ptax_compra = ptax_rate_value(registro.get("cotacaoCompra"))
        ptax_venda = ptax_rate_value(registro.get("cotacaoVenda"))
        if ptax_compra is None or ptax_venda is None:
            continue
        candidato = {
            "data_cotacao": data_cotacao,
            "moeda": moeda,
            "ptax_compra": ptax_compra,
            "ptax_venda": ptax_venda,
            "_data_hora": data_hora,
        }
        anterior = fechamentos.get(data_cotacao)
        if anterior is None or candidato["_data_hora"] > anterior["_data_hora"]:
            fechamentos[data_cotacao] = candidato

    resultado = []
    for registro in sorted(fechamentos.values(), key=lambda item: item["data_cotacao"], reverse=True):
        registro.pop("_data_hora", None)
        resultado.append(registro)
    return resultado

def parse_ptax_importacao(form):
    datas = form.getlist("cotacao_data")
    moedas = form.getlist("cotacao_moeda")
    compras = form.getlist("cotacao_compra")
    vendas = form.getlist("cotacao_venda")
    tamanhos = {len(datas), len(moedas), len(compras), len(vendas)}
    if len(tamanhos) != 1 or not datas:
        raise ValueError("A prévia de PTAX está incompleta e não pode ser importada.")

    registros = []
    for data, moeda, compra, venda in zip(datas, moedas, compras, vendas):
        moeda = (moeda or "").strip().upper()
        if moeda not in PTAX_MOEDAS:
            raise ValueError("A prévia contém uma moeda inválida.")
        data = parse_date(data)
        if not data:
            raise ValueError("A prévia contém uma data inválida.")
        registros.append({
            "data_cotacao": data,
            "moeda": moeda,
            "ptax_compra": parse_ptax_rate(compra, "PTAX Compra"),
            "ptax_venda": parse_ptax_rate(venda, "PTAX Venda"),
        })
    return registros

def render_ptax_page(previsao=None, consulta=None):
    consulta = consulta or {}
    form_values = {
        "moeda": consulta.get("moeda") or "USD",
        "data_inicial": consulta.get("data_inicial") or "",
        "data_final": consulta.get("data_final") or "",
    }
    conn = db()
    historico = conn.execute("""
        SELECT data_cotacao, moeda, ptax_compra, ptax_venda
        FROM ptax_cotacoes
        ORDER BY data_cotacao DESC, moeda ASC
    """).fetchall()
    conn.close()
    return render_template("ptax.html", moedas=PTAX_MOEDAS, consulta=form_values,
                           previsao=previsao, historico=historico)

def contract_summary(conn, contrato_id):
    row = conn.execute("""
        SELECT c.id, c.numero_contrato, c.moeda, c.valor_moeda, c.status,
               COALESCE((SELECT SUM(v.valor_alocado) FROM invoice_contrato_cambio v
                         WHERE v.contrato_id=c.id), c.valor_moeda) AS valor_moeda_consolidado,
               c.saldo_zerado_manual,
               COALESCE(SUM(CASE WHEN m.tipo='VINCULACAO' THEN m.valor ELSE 0 END),0) AS vinculado
        FROM contratos c
        LEFT JOIN due_movimentacoes m ON m.contrato_id=c.id
        WHERE c.id=?
        GROUP BY c.id
    """, (contrato_id,)).fetchone()
    if not row:
        return None
    data = decorate_contract(row)
    return data

init_db()

@app.route("/")
def index():
    conn = db()
    dues = [decorate_due(row) for row in conn.execute("""
        SELECT d.*, e.apelido AS empresa_apelido,
               COALESCE(SUM(CASE WHEN m.tipo IN ('UTILIZACAO','VINCULACAO') THEN m.valor ELSE -m.valor END),0) AS utilizado
        FROM dues d
        LEFT JOIN due_movimentacoes m ON m.due_id=d.id
        LEFT JOIN empresas e ON e.cnpj=d.cnpj
        GROUP BY d.id ORDER BY d.id DESC
    """).fetchall()]
    contratos = [decorate_contract(row) for row in conn.execute("""
        SELECT c.*, e.apelido AS empresa_apelido,
               COALESCE((SELECT SUM(v.valor_alocado) FROM invoice_contrato_cambio v
                         WHERE v.contrato_id=c.id), c.valor_moeda) AS valor_moeda_consolidado,
               COALESCE(SUM(CASE WHEN m.tipo='VINCULACAO' THEN m.valor ELSE 0 END),0) AS vinculado
        FROM contratos c
        LEFT JOIN due_movimentacoes m ON m.contrato_id=c.id
        LEFT JOIN empresas e ON e.cnpj=c.cnpj
        GROUP BY c.id ORDER BY c.id DESC
    """).fetchall()]
    resumo = {
        "invoices": conn.execute("SELECT COUNT(*) FROM invoices").fetchone()[0],
        "dues": conn.execute("SELECT COUNT(*) FROM dues").fetchone()[0],
        "contratos": conn.execute("SELECT COUNT(*) FROM contratos").fetchone()[0],
        "dues_sem_vinculo": conn.execute("""
            SELECT COUNT(*) FROM dues d
            WHERE NOT EXISTS (SELECT 1 FROM due_contratos v WHERE v.due_id=d.id)
        """).fetchone()[0],
        "contratos_sem_vinculo": conn.execute("""
            SELECT COUNT(*) FROM contratos c
            WHERE NOT EXISTS (SELECT 1 FROM due_contratos v WHERE v.contrato_id=c.id)
        """).fetchone()[0],
    }
    conn.close()
    return render_template("index.html", dues=dues, contratos=contratos, resumo=resumo)

def contratos_filtros(args):
    empresa_id = form_record_id(args.get("empresa_id"))
    competencia_id = form_record_id(args.get("competencia_id"))
    numero_contrato = (args.get("numero_contrato") or "").strip()
    where, params = [], []
    if numero_contrato:
        where.append("c.numero_contrato LIKE ?"); params.append(f"%{numero_contrato}%")
    if empresa_id:
        where.append("e.id=?"); params.append(empresa_id)
    if competencia_id:
        where.append("c.competencia_id=?"); params.append(competencia_id)
    clause = " WHERE " + " AND ".join(where) if where else ""
    return empresa_id, competencia_id, clause, params

def consulta_contratos(conn, args):
    empresa_id, competencia_id, clause, params = contratos_filtros(args)
    rows = conn.execute(f"""
        SELECT c.*, e.id AS empresa_id_filtro, e.razao_social AS empresa_razao_social,
               e.apelido AS empresa_apelido, comp.descricao AS competencia_descricao,
               COALESCE((SELECT SUM(v.valor_alocado) FROM invoice_contrato_cambio v
                         WHERE v.contrato_id=c.id), c.valor_moeda) AS valor_moeda_consolidado,
               COALESCE(SUM(CASE WHEN m.tipo='VINCULACAO' THEN m.valor ELSE 0 END),0) AS vinculado
        FROM contratos c
        LEFT JOIN empresas e ON REPLACE(REPLACE(REPLACE(REPLACE(COALESCE(c.cnpj,''),'.',''),'/',''),'-',''),' ','')=e.cnpj
        LEFT JOIN competencias comp ON comp.id=c.competencia_id
        LEFT JOIN due_movimentacoes m ON m.contrato_id=c.id
        {clause}
        GROUP BY c.id ORDER BY c.id DESC
    """, params).fetchall()
    return rows, empresa_id, competencia_id

@app.route("/contratos")
def lista_contratos():
    conn = db()
    rows, empresa_id, competencia_id = consulta_contratos(conn, request.args)
    contratos = [decorate_contract(row) for row in rows]
    empresas = conn.execute("SELECT id, razao_social, apelido, cnpj FROM empresas ORDER BY razao_social").fetchall()
    competencias = conn.execute("SELECT id, descricao, data_inicial, data_final FROM competencias ORDER BY data_inicial DESC, descricao").fetchall()
    conn.close()
    return render_template("contratos.html", contratos=contratos, empresas=empresas, competencias=competencias,
                           empresa_id=empresa_id, competencia_id=competencia_id)

@app.route("/contratos/excluir-lote", methods=["POST"])
def excluir_contratos_lote():
    conn = db()
    try:
        contrato_ids = selected_record_ids(request.form)
        conn.execute("BEGIN IMMEDIATE")
        ensure_existing_record_ids(conn, "contratos", contrato_ids, "contratos")
        placeholders = ",".join("?" for _ in contrato_ids)
        invoice_links = conn.execute(f"""
            SELECT COUNT(*)
            FROM invoice_contrato_cambio
            WHERE contrato_id IN ({placeholders})
        """, contrato_ids).fetchone()[0]
        if invoice_links:
            raise ValueError("Contratos Câmbio derivados de Invoices devem ser excluídos pelos vínculos da própria Invoice.")
        due_rows = conn.execute(f"""
            SELECT DISTINCT due_id
            FROM due_movimentacoes
            WHERE contrato_id IN ({placeholders})
            UNION
            SELECT DISTINCT due_id
            FROM due_contratos
            WHERE contrato_id IN ({placeholders})
        """, contrato_ids + contrato_ids).fetchall()
        due_ids = [row[0] for row in due_rows]

        conn.execute(f"""
            DELETE FROM due_movimentacoes
            WHERE contrato_id IN ({placeholders})
               OR due_contrato_id IN (
                   SELECT id FROM due_contratos WHERE contrato_id IN ({placeholders})
               )
        """, contrato_ids + contrato_ids)
        conn.execute(f"DELETE FROM due_contratos WHERE contrato_id IN ({placeholders})", contrato_ids)
        conn.execute(f"DELETE FROM contratos WHERE id IN ({placeholders})", contrato_ids)
        recalculate_statuses(conn, due_ids=due_ids, contrato_ids=[])
        conn.commit()
        flash(f"{len(contrato_ids)} Contrato(s) Câmbio excluído(s) com sucesso.", "success")
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), "danger")
    except (sqlite3.Error, OverflowError):
        conn.rollback()
        flash("Não foi possível concluir a exclusão dos Contratos Câmbio.", "danger")
    except Exception:
        conn.rollback()
        flash("Não foi possível concluir a exclusão dos Contratos Câmbio.", "danger")
    finally:
        conn.close()
    return redirect_batch_result("lista_contratos")

@app.route("/contratos/exportar")
def exportar_contratos():
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    conn = db()
    rows, _, _ = consulta_contratos(conn, request.args)
    contrato_ids = [row["id"] for row in rows]
    vinculos = []
    if contrato_ids:
        placeholders = ",".join("?" for _ in contrato_ids)
        vinculos = conn.execute(f"""
            SELECT v.id, v.due_id, v.contrato_id, v.valor_vinculado, v.observacao,
                   v.created_at, c.numero_contrato, c.moeda AS contrato_moeda,
                   d.numero_due, d.chave_acesso, d.created_at AS data_lancamento,
                   COALESCE(SUM(m.valor), v.valor_vinculado) AS valor_calculado
            FROM due_contratos v
            JOIN contratos c ON c.id=v.contrato_id
            JOIN dues d ON d.id=v.due_id
            LEFT JOIN due_movimentacoes m ON m.due_contrato_id=v.id AND m.tipo='VINCULACAO'
            WHERE v.contrato_id IN ({placeholders})
            GROUP BY v.id
            ORDER BY c.numero_contrato, d.numero_due, v.id
        """, contrato_ids).fetchall()
    conn.close()

    def excel_value(value, field=None):
        if value is None:
            return None
        if field in {"data_contrato", "data_recebimento", "data_liquidacao", "created_at", "data_lancamento"}:
            return date_br(value)
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, bool):
            return "Sim" if value else "Não"
        return value

    labels = {
        "id": "ID", "numero_contrato": "Número do contrato", "banco": "Banco legado",
        "banco_credito": "Banco de Crédito", "banco_liquidacao": "Banco de Liquidação",
        "data_contrato": "Data do contrato", "data_recebimento": "Data de recebimento",
        "data_liquidacao": "Data de liquidação", "cnpj": "CNPJ", "cliente": "Cliente",
        "moeda": "Moeda", "valor_moeda": "Valor na moeda", "taxa_cambio": "Taxa de câmbio",
        "valor_reais": "Valor em reais", "status": "Status", "saldo_zerado_manual": "Saldo zerado manualmente",
        "observacao": "Observação", "created_at": "Criado em", "competencia_id": "Competência ID",
        "empresa_id_filtro": "Empresa ID", "empresa_razao_social": "Empresa - Razão social",
        "empresa_apelido": "Empresa - Apelido", "competencia_descricao": "Competência",
        "vinculado": "Total vinculado", "saldo": "Saldo disponível",
    }
    contratos_data = []
    for row in rows:
        item = decorate_contract(row)
        contratos_data.append({labels.get(key, key): excel_value(value, key) for key, value in item.items()})
    vinculos_data = []
    vinculo_labels = {
        "id": "Vínculo ID", "due_id": "DU-E ID", "contrato_id": "Contrato ID",
        "numero_contrato": "Número do contrato", "numero_due": "Número da DU-E",
        "chave_acesso": "Chave de acesso", "data_lancamento": "Data de lançamento",
        "contrato_moeda": "Moeda do contrato", "valor_vinculado": "Valor vinculado registrado",
        "valor_calculado": "Valor vinculado calculado", "observacao": "Observação",
        "created_at": "Vínculo criado em",
    }
    for row in vinculos:
        vinculos_data.append({vinculo_labels.get(key, key): excel_value(value, key) for key, value in dict(row).items()})

    contrato_columns = list(labels.values())
    vinculo_columns = list(vinculo_labels.values())
    contratos_df = pd.DataFrame(contratos_data, columns=contrato_columns)
    vinculos_df = pd.DataFrame(vinculos_data, columns=vinculo_columns)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        contratos_df.to_excel(writer, index=False, sheet_name="Contratos")
        vinculos_df.to_excel(writer, index=False, sheet_name="Vínculos")
        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1769AA")
                cell.alignment = Alignment(horizontal="center")
            for column in worksheet.columns:
                letter = get_column_letter(column[0].column)
                width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 45)
                worksheet.column_dimensions[letter].width = width
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="contratos_exportacao.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

REPORT_GRANULARITIES = ("diario", "semanal", "mensal")
REPORT_GRANULARITY_LABELS = {"diario": "diário", "semanal": "semanal", "mensal": "mensal"}

def report_date(value):
    try:
        return date.fromisoformat(str(value)[:10])
    except (TypeError, ValueError):
        return None

def report_bucket(value, granularity):
    current = report_date(value)
    if not current:
        return None
    if granularity == "semanal":
        current -= timedelta(days=current.weekday())
    elif granularity == "mensal":
        current = current.replace(day=1)
    return current.isoformat()

def report_period_label(value, granularity):
    current = report_date(value)
    if not current:
        return "Data não informada"
    if granularity == "semanal":
        return f"Semana de {current.strftime('%d/%m/%Y')}"
    if granularity == "mensal":
        return current.strftime("%m/%Y")
    return current.strftime("%d/%m/%Y")

def report_chart_label(value, granularity):
    current = report_date(value)
    if not current:
        return "-"
    if granularity == "semanal":
        return f"Sem. {current.strftime('%d/%m')}"
    if granularity == "mensal":
        return current.strftime("%m/%Y")
    return current.strftime("%d/%m")

def choose_report_granularity(items):
    dates = [report_date(item.get("data_contrato")) for item in items]
    dates = [item for item in dates if item]
    if not dates:
        return "diario"
    span_days = (max(dates) - min(dates)).days
    distinct_days = len(set(dates))
    record_count = len(items)
    if record_count <= 60 and distinct_days <= 45:
        return "diario"
    if span_days <= 365 and distinct_days <= 180 and record_count <= 3000:
        return "semanal"
    return "mensal"

def report_summary():
    return {
        "contratos": 0, "volume": Decimal("0"), "usd_volume": Decimal("0"),
        "brl_total": Decimal("0"), "resultado": Decimal("0"), "resultado_count": 0,
        "taxa_valor": Decimal("0"), "taxa_volume": Decimal("0"),
        "ptax_valor": Decimal("0"), "ptax_volume": Decimal("0"),
        "ptax_dia": None,
    }

def report_brl_value(item):
    if item.get("valor_reais") is not None:
        return item["valor_reais"]
    if item.get("taxa_cambio") is not None:
        return item["valor_moeda"] * item["taxa_cambio"]
    return None

def add_report_item(summary, item):
    volume = item["valor_moeda"]
    summary["contratos"] += 1
    summary["volume"] += volume
    if item.get("moeda") == "USD":
        summary["usd_volume"] += volume
    brl_value = item.get("valor_brl")
    if brl_value is not None:
        summary["brl_total"] += brl_value
    if item["resultado"] is not None:
        summary["resultado"] += item["resultado"]
        summary["resultado_count"] += 1
    if item["taxa_cambio"] is not None and volume > 0:
        summary["taxa_valor"] += item["taxa_cambio"] * volume
        summary["taxa_volume"] += volume
    if item.get("ptax_venda") is not None and volume > 0:
        summary["ptax_valor"] += item["ptax_venda"] * volume
        summary["ptax_volume"] += volume
    if item.get("ptax_venda") is not None and summary["ptax_dia"] is None:
        summary["ptax_dia"] = item["ptax_venda"]

def finish_report_summary(summary):
    summary["taxa_ponderada"] = summary["taxa_valor"] / summary["taxa_volume"] if summary["taxa_volume"] else None
    summary["ptax_ponderada"] = summary["ptax_valor"] / summary["ptax_volume"] if summary["ptax_volume"] else None
    summary["resultado"] = summary["resultado"] if summary["resultado_count"] else None
    return summary

def add_result_accumulated(summaries):
    acumulado = Decimal("0")
    tem_resultado = False
    for summary in summaries:
        if summary["resultado"] is not None:
            acumulado += summary["resultado"]
            tem_resultado = True
        summary["resultado_acumulado"] = acumulado if tem_resultado else None

def add_result_accumulated_by_period(items):
    """Aplica o mesmo acumulado a todas as moedas de cada período."""
    acumulado = Decimal("0")
    tem_resultado = False
    index = 0
    while index < len(items):
        periodo = items[index]["periodo"]
        end = index
        resultado_periodo = Decimal("0")
        tem_resultado_periodo = False
        while end < len(items) and items[end]["periodo"] == periodo:
            resumo = items[end]["resumo"]
            if resumo["resultado"] is not None:
                resultado_periodo += resumo["resultado"]
                tem_resultado_periodo = True
            end += 1
        if tem_resultado_periodo:
            acumulado += resultado_periodo
            tem_resultado = True
        for item in items[index:end]:
            item["resumo"]["resultado_acumulado"] = acumulado if tem_resultado else None
        index = end

def grouped_report_summaries(items, granularity):
    grouped = {}
    for item in items:
        bucket = report_bucket(item.get("data_contrato"), granularity)
        if not bucket:
            continue
        key = (bucket, item["moeda"])
        grouped.setdefault(key, report_summary())
        add_report_item(grouped[key], item)
    result = []
    for (bucket, moeda), summary in sorted(grouped.items(), key=lambda entry: (entry[0][0], entry[0][1])):
        result.append({"periodo": bucket, "rotulo": report_period_label(bucket, granularity),
                       "moeda": moeda, "resumo": finish_report_summary(summary)})
    add_result_accumulated_by_period(result)
    return result

def build_contract_report_chart(rows, moeda, granularity="diario"):
    """Prepara a série temporal compartilhada pela tela e pelo PDF."""
    daily = {}
    for row in rows:
        if row["moeda"] != moeda or not row["data_contrato"]:
            continue
        item = daily.setdefault(row["data_contrato"], {"date": row["data_contrato"], "ptax_soma": Decimal("0"), "ptax_count": 0, "volume": Decimal("0"), "weighted_rate": Decimal("0")})
        if row["ptax_venda"] is not None:
            item["ptax_soma"] += row["ptax_venda"]
            item["ptax_count"] += 1
        if row["taxa_cambio"] is not None and row["valor_moeda"] > 0:
            item["volume"] += row["valor_moeda"]
            item["weighted_rate"] += row["valor_moeda"] * row["taxa_cambio"]
    daily = sorted(daily.values(), key=lambda item: report_date(item["date"]) or date.max)
    for item in daily:
        item["ptax"] = item["ptax_soma"] / item["ptax_count"] if item["ptax_count"] else None
        item["fechamento"] = item["weighted_rate"] / item["volume"] if item["volume"] else None
    if granularity != "diario":
        grouped = {}
        for item in daily:
            bucket = report_bucket(item["date"], granularity)
            target = grouped.setdefault(bucket, {"date": bucket, "ptax_soma": Decimal("0"), "ptax_count": 0, "volume": Decimal("0"), "weighted_rate": Decimal("0")})
            if item["ptax"] is not None:
                target["ptax_soma"] += item["ptax"]
                target["ptax_count"] += 1
            target["volume"] += item["volume"]
            target["weighted_rate"] += item["weighted_rate"]
        daily = sorted(grouped.values(), key=lambda item: report_date(item["date"]) or date.max)
        for item in daily:
            item["ptax"] = item["ptax_soma"] / item["ptax_count"] if item["ptax_count"] else None
            item["fechamento"] = item["weighted_rate"] / item["volume"] if item["volume"] else None
    values = [value for item in daily for value in (item["ptax"], item["fechamento"]) if value is not None]
    base = {"moeda": moeda, "granularidade": granularity,
            "granularidade_label": REPORT_GRANULARITY_LABELS[granularity], "points": []}
    if not values:
        return base
    minimum, maximum = min(values), max(values)
    padding = max((maximum - minimum) * Decimal("0.12"), Decimal("0.001"))
    minimum -= padding; maximum += padding
    count = max(len(daily) - 1, 1)
    label_step = max(1, (len(daily) + 7) // 8)
    points = []
    for index, item in enumerate(daily):
        show_label = index == 0 or index == len(daily) - 1 or index % label_step == 0
        point = {"date": item["date"], "label": report_chart_label(item["date"], granularity),
                 "show_label": show_label, "x": round(40 + index * 720 / count, 2),
                 "ptax_valor": item["ptax"], "fechamento_valor": item["fechamento"]}
        for key in ("ptax", "fechamento"):
            value = item[key]
            point[key] = round(float(160 - ((value - minimum) / (maximum - minimum) * 130)), 2) if value is not None else None
        points.append(point)
    segments = {"ptax": [], "fechamento": []}
    for key in segments:
        current = []
        for point in points:
            if point[key] is None:
                if current: segments[key].append(current); current = []
            else:
                current.append(f"{point['x']},{point[key]}")
        if current: segments[key].append(current)
    base["points"] = points
    base["segments"] = segments
    return base

def report_dimension_rows(items, name_getter, total_brl):
    grouped = {}
    for item in items:
        name = name_getter(item) or "Não informado"
        summary = grouped.setdefault(name, report_summary())
        add_report_item(summary, item)
    result = []
    for name, summary in sorted(grouped.items(), key=lambda entry: entry[0].casefold()):
        participation = summary["brl_total"] / total_brl * Decimal("100") if total_brl else Decimal("0")
        result.append({"nome": name, "total_usd": summary["usd_volume"],
                       "total_brl": summary["brl_total"], "participacao": participation})
    return result

def report_period_range_label(inicio, fim, periodo):
    if inicio and fim:
        return f"{date_br(inicio)} a {date_br(fim)}"
    if inicio:
        return f"A partir de {date_br(inicio)}"
    if fim:
        return f"Até {date_br(fim)}"
    if periodo != "todos":
        return f"Últimos {periodo} dias"
    return "Todo o período"

def parse_report_filters(args, forced_granularity=None):
    data_de = (args.get("data_de", "") or "").strip()
    data_ate = (args.get("data_ate", "") or "").strip()
    numero_contrato = (args.get("numero_contrato", "") or "").strip()
    moeda = (args.get("moeda", "") or "").strip().upper()
    periodo = (args.get("periodo", "todos") or "todos").strip()
    if periodo not in {"todos", "30", "90", "180", "365"}:
        periodo = "todos"
    agrupamento_solicitado = forced_granularity or (args.get("agrupamento", "auto") or "auto").strip().lower()
    if agrupamento_solicitado not in {"auto", *REPORT_GRANULARITIES}:
        agrupamento_solicitado = "auto"
    try:
        inicio = parse_date(data_de) if data_de else None
        fim = parse_date(data_ate) if data_ate else None
    except ValueError:
        raise ValueError("Data inválida. Use o formato dd/mm/aaaa.")
    if not data_de and not data_ate and periodo != "todos":
        fim = date.today().isoformat()
        inicio = (date.today() - timedelta(days=int(periodo) - 1)).isoformat()
    if inicio and fim and inicio > fim:
        raise ValueError("A data inicial não pode ser posterior à data final.")
    return {
        "data_de": data_de, "data_ate": data_ate, "numero_contrato": numero_contrato,
        "moeda": moeda, "periodo": periodo, "inicio": inicio, "fim": fim,
        "agrupamento_solicitado": agrupamento_solicitado,
        "empresa_id": form_record_id(args.get("empresa_id")),
        "competencia_id": form_record_id(args.get("competencia_id")),
    }

def build_contract_report_context(args, forced_granularity=None):
    filters = parse_report_filters(args, forced_granularity=forced_granularity)
    where, params = [], []
    if filters["inicio"]:
        where.append("c.data_contrato >= ?"); params.append(filters["inicio"])
    if filters["fim"]:
        where.append("c.data_contrato <= ?"); params.append(filters["fim"])
    if filters["numero_contrato"]:
        where.append("c.numero_contrato LIKE ?"); params.append(f"%{filters['numero_contrato']}%")
    if filters["moeda"]:
        where.append("c.moeda = ?"); params.append(filters["moeda"])
    if filters["empresa_id"]:
        where.append("e.id = ?"); params.append(filters["empresa_id"])
    if filters["competencia_id"]:
        where.append("c.competencia_id = ?"); params.append(filters["competencia_id"])
    clause = " WHERE " + " AND ".join(where) if where else ""
    conn = db()
    try:
        rows = conn.execute(f"""SELECT c.id, c.numero_contrato, c.data_contrato, c.cnpj, c.moeda,
                c.valor_moeda, c.taxa_cambio, c.valor_reais, c.cliente, c.cliente_id,
                c.banco, c.banco_credito, c.banco_liquidacao, c.competencia_id,
                e.id AS empresa_id, e.razao_social AS empresa_razao_social,
                e.apelido AS empresa_apelido, comp.descricao AS competencia_descricao,
                cl.nome AS cliente_base_nome,
                p.ptax_venda
            FROM contratos c
            LEFT JOIN empresas e ON REPLACE(REPLACE(REPLACE(REPLACE(COALESCE(c.cnpj,''),'.',''),'/',''),'-',''),' ','')=e.cnpj
            LEFT JOIN competencias comp ON comp.id=c.competencia_id
            LEFT JOIN clientes cl ON cl.id=c.cliente_id
            LEFT JOIN ptax_cotacoes p ON p.moeda=c.moeda AND p.data_cotacao=date(c.data_contrato)
            {clause}
            ORDER BY CASE WHEN c.data_contrato IS NULL OR c.data_contrato='' THEN 1 ELSE 0 END,
                     c.data_contrato ASC, c.numero_contrato ASC""", params).fetchall()
        empresas = conn.execute("""SELECT id, razao_social, apelido, cnpj FROM empresas
                                  ORDER BY CASE WHEN TRIM(COALESCE(apelido,''))<>'' THEN 0 ELSE 1 END,
                                           apelido, razao_social""").fetchall()
        competencias = conn.execute("""SELECT comp.id, comp.empresa_id, comp.descricao,
                comp.data_inicial, comp.data_final, e.apelido, e.razao_social
            FROM competencias comp JOIN empresas e ON e.id=comp.empresa_id
            ORDER BY comp.data_inicial DESC, comp.descricao""").fetchall()
    finally:
        conn.close()

    contratos = []
    for row in rows:
        item = dict(row)
        item["valor_moeda"] = decimal_value(item.get("valor_moeda"))
        item["taxa_cambio"] = decimal_value(item["taxa_cambio"]) if item.get("taxa_cambio") is not None else None
        item["valor_reais"] = decimal_value(item["valor_reais"]) if item.get("valor_reais") is not None else None
        item["ptax_venda"] = decimal_value(item["ptax_venda"]) if item.get("ptax_venda") is not None else None
        item["resultado"] = (item["valor_moeda"] * (item["taxa_cambio"] - item["ptax_venda"])
                             if item["taxa_cambio"] is not None and item["ptax_venda"] is not None else None)
        item["cliente_nome"] = item.get("cliente_base_nome") or item.get("cliente") or "Não informado"
        item["empresa_nome"] = item.get("empresa_apelido") or item.get("empresa_razao_social") or item.get("cnpj") or "Não informado"
        item["banco_recebedor"] = item.get("banco_liquidacao") or item.get("banco_credito") or item.get("banco") or "Não informado"
        item["valor_brl"] = report_brl_value(item)
        contratos.append(item)

    granularidade = (choose_report_granularity(contratos)
                     if filters["agrupamento_solicitado"] == "auto"
                     else filters["agrupamento_solicitado"])
    monthly_groups = grouped_report_summaries(contratos, "mensal")
    daily_groups = grouped_report_summaries(contratos, "diario")
    period_groups = grouped_report_summaries(contratos, granularidade)
    mensais = [((item["periodo"], item["moeda"]), item["resumo"]) for item in monthly_groups]
    diarios = [((item["periodo"], item["moeda"]), item["resumo"]) for item in daily_groups]
    moedas = sorted({item["moeda"] for item in contratos if item["moeda"]})
    graficos = [build_contract_report_chart(contratos, item, granularidade) for item in moedas]

    total = report_summary()
    currency_totals = {}
    for item in contratos:
        add_report_item(total, item)
        currency_totals.setdefault(item["moeda"], report_summary())
        add_report_item(currency_totals[item["moeda"]], item)
    finish_report_summary(total)
    for summary in currency_totals.values():
        finish_report_summary(summary)
    operation_days = sorted({item["data_contrato"][:10] for item in contratos if item.get("data_contrato")})
    operations = len(contratos)
    day_count = len(operation_days)
    accumulated = next((item["resumo"]["resultado_acumulado"] for item in reversed(period_groups)
                        if item["resumo"].get("resultado_acumulado") is not None), None)
    kpis = {
        "total_usd": total["usd_volume"], "total_brl": total["brl_total"],
        "resultado": total["resultado"], "resultado_acumulado": accumulated,
        "operacoes": operations, "dias": day_count,
        "media_usd_operacao": total["usd_volume"] / operations if operations else Decimal("0"),
        "media_brl_operacao": total["brl_total"] / operations if operations else Decimal("0"),
        "media_usd_dia": total["usd_volume"] / day_count if day_count else Decimal("0"),
        "media_brl_dia": total["brl_total"] / day_count if day_count else Decimal("0"),
        "taxas": [{"moeda": moeda, "valor": summary["taxa_ponderada"]}
                  for moeda, summary in sorted(currency_totals.items())
                  if summary["taxa_ponderada"] is not None],
    }
    comparacoes_por_dia = {}
    for (dia, _moeda), resumo in diarios:
        if resumo.get("taxa_ponderada") is not None and resumo.get("ptax_dia") is not None:
            comparacoes_por_dia.setdefault(dia, []).append(resumo["taxa_ponderada"] - resumo["ptax_dia"])
    stats = {
        "dias_acima": sum(1 for deltas in comparacoes_por_dia.values() if any(delta > 0 for delta in deltas)),
        "dias_abaixo": sum(1 for deltas in comparacoes_por_dia.values() if any(delta < 0 for delta in deltas)),
        "dias_operacao": day_count,
        "resultado_acumulado": accumulated,
    }
    total_brl = total["brl_total"]
    por_empresa = report_dimension_rows(contratos, lambda item: item.get("empresa_nome"), total_brl)
    por_cliente = report_dimension_rows(contratos, lambda item: item.get("cliente_nome"), total_brl)
    por_banco = report_dimension_rows(contratos, lambda item: item.get("banco_recebedor"), total_brl)
    selected_empresa = next((item for item in empresas if item["id"] == filters["empresa_id"]), None)
    selected_competencia = next((item for item in competencias if item["id"] == filters["competencia_id"]), None)
    empresa_label = ((selected_empresa["apelido"] or selected_empresa["razao_social"])
                     if selected_empresa else "Todas as empresas")
    safra_label = selected_competencia["descricao"] if selected_competencia else "Todas as safras"
    return {
        "contratos": contratos, "mensais": mensais, "diarios": diarios,
        "agrupados": period_groups, "periodos": period_groups,
        "agrupamento_solicitado": filters["agrupamento_solicitado"],
        "agrupamento": granularidade, "agrupamento_label": REPORT_GRANULARITY_LABELS[granularidade],
        "periodo": filters["periodo"], "periodo_label": report_period_range_label(filters["inicio"], filters["fim"], filters["periodo"]),
        "graficos": graficos, "moedas": moedas,
        "numero_contrato": filters["numero_contrato"], "moeda": filters["moeda"],
        "data_de": filters["data_de"], "data_ate": filters["data_ate"],
        "empresa_id": filters["empresa_id"], "competencia_id": filters["competencia_id"],
        "empresas": empresas, "competencias": competencias,
        "empresa_label": empresa_label, "safra_label": safra_label,
        "kpis": kpis, "stats": stats,
        "por_empresa": por_empresa, "por_cliente": por_cliente, "por_banco": por_banco,
        "data_emissao": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }

@app.route("/contratos/relatorios")
def relatorios_contratos():
    try:
        context = build_contract_report_context(request.args)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("relatorios_contratos"))
    return render_template("contratos_relatorios.html", **context)

def report_pdf_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = (
        (Path("C:/Windows/Fonts/arial.ttf"), Path("C:/Windows/Fonts/arialbd.ttf")),
        (Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"), Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")),
    )
    for regular_path, bold_path in candidates:
        if not regular_path.exists():
            continue
        try:
            if "DueReport" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("DueReport", str(regular_path)))
            if bold_path.exists() and "DueReportBold" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("DueReportBold", str(bold_path)))
            return "DueReport", "DueReportBold" if bold_path.exists() else "DueReport"
        except Exception:
            continue
    return "Helvetica", "Helvetica-Bold"

def report_pdf_styles():
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

    regular, bold = report_pdf_fonts()
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("DuePdfTitle", parent=styles["Title"], fontName=bold,
                              fontSize=18, leading=21, textColor=colors.HexColor("#18324f"),
                              spaceAfter=3))
    styles.add(ParagraphStyle("DuePdfSubtitle", parent=styles["Normal"], fontName=regular,
                              fontSize=8.5, leading=11, textColor=colors.HexColor("#557084"),
                              spaceAfter=8))
    styles.add(ParagraphStyle("DuePdfSection", parent=styles["Heading2"], fontName=bold,
                              fontSize=10.5, leading=13, textColor=colors.HexColor("#18324f"),
                              spaceBefore=8, spaceAfter=4))
    styles.add(ParagraphStyle("DuePdfBody", parent=styles["BodyText"], fontName=regular,
                              fontSize=7.5, leading=9.5, textColor=colors.HexColor("#263b4d")))
    styles.add(ParagraphStyle("DuePdfSmall", parent=styles["BodyText"], fontName=regular,
                              fontSize=6.8, leading=8, textColor=colors.HexColor("#557084")))
    styles.add(ParagraphStyle("DuePdfTableHeader", parent=styles["BodyText"], fontName=bold,
                              fontSize=7, leading=8.5, textColor=colors.HexColor("#18324f")))
    styles.add(ParagraphStyle("DuePdfKpiLabel", parent=styles["BodyText"], fontName=regular,
                              fontSize=6.8, leading=8, textColor=colors.HexColor("#557084")))
    styles.add(ParagraphStyle("DuePdfKpiValue", parent=styles["BodyText"], fontName=bold,
                              fontSize=10.5, leading=12, textColor=colors.HexColor("#18324f")))
    styles.add(ParagraphStyle("DuePdfChartLabel", parent=styles["BodyText"], fontName=regular,
                              fontSize=6.5, leading=7.5, textColor=colors.HexColor("#557084")))
    return styles, regular, bold

def report_pdf_paragraph(value, style, bold=False):
    from html import escape
    from reportlab.platypus import Paragraph

    text = "-" if value is None or str(value) == "" else str(value)
    text = escape(text, quote=False).replace("\n", "<br/>")
    if bold:
        text = f"<b>{text}</b>"
    return Paragraph(text, style)

def report_pdf_percent(value):
    if value is None:
        return "-"
    return f"{float(value):,.1f}%".replace(",", "X").replace(".", ",").replace("X", ".")

def report_pdf_table(story, title, headers, rows, widths, styles, accent="#EAF3FA"):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    story.append(report_pdf_paragraph(title, styles["DuePdfSection"], bold=True))
    if not rows:
        story.append(report_pdf_paragraph("Nenhum registro encontrado para os filtros informados.", styles["DuePdfSmall"]))
        return
    data = [[report_pdf_paragraph(header, styles["DuePdfTableHeader"]) for header in headers]]
    for row in rows:
        data.append([report_pdf_paragraph(value, styles["DuePdfBody"]) for value in row])
    table = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(accent)),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8E1E8")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFCFD")]),
    ]))
    story.append(table)

def report_pdf_kpis(story, kpis, styles, available_width):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    rates = "<br/>".join(
        f"{item['moeda']}: {report_pdf_paragraph(item['valor'], styles['DuePdfKpiValue']).getPlainText() if item['valor'] is not None else '-'}"
        for item in kpis["taxas"]
    ) or "-"
    values = [
        ("Total USD fechado", money(kpis["total_usd"]), "#EAF3FA"),
        ("Total BRL", money(kpis["total_brl"]), "#FFF5D6"),
        ("Taxa ponderada", rates, "#E5F4E8"),
        ("Resultado financeiro vs PTAX", money(kpis["resultado"]) if kpis["resultado"] is not None else "-", "#EAF3FA"),
        ("Quantidade de operações", str(kpis["operacoes"]), "#FFF5D6"),
        ("Dias com operação", str(kpis["dias"]), "#E5F4E8"),
        ("Volume médio/operação USD", money(kpis["media_usd_operacao"]), "#EAF3FA"),
        ("Volume médio/operação BRL", money(kpis["media_brl_operacao"]), "#FFF5D6"),
        ("Volume médio/dia USD", money(kpis["media_usd_dia"]), "#E5F4E8"),
        ("Volume médio/dia BRL", money(kpis["media_brl_dia"]), "#EAF3FA"),
    ]
    cards = []
    card_width = available_width / 5
    for label, value, background in values:
        value_html = value if label == "Taxa ponderada" else report_pdf_paragraph(value, styles["DuePdfKpiValue"]).getPlainText()
        card = Table([[report_pdf_paragraph(label, styles["DuePdfKpiLabel"]),
                       report_pdf_paragraph(value_html, styles["DuePdfKpiValue"])]] ,
                     colWidths=[card_width * .65, card_width * .35])
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(background)),
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#D8E1E8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        cards.append(card)
    rows = []
    for index in range(0, len(cards), 5):
        row = cards[index:index + 5]
        while len(row) < 5:
            row.append("")
        rows.append(row)
    table = Table(rows, colWidths=[card_width] * 5, hAlign="LEFT")
    table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                               ("LEFTPADDING", (0, 0), (-1, -1), 0),
                               ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                               ("TOPPADDING", (0, 0), (-1, -1), 0),
                               ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
    story.append(report_pdf_paragraph("Indicadores do período", styles["DuePdfSection"], bold=True))
    story.append(table)

def report_pdf_chart_drawing(grafico, styles, width=790, height=175):
    from reportlab.graphics.shapes import Drawing, Line, PolyLine, String
    from reportlab.lib import colors

    points = grafico.get("points", [])
    values = [point[key] for point in points for key in ("ptax_valor", "fechamento_valor") if point.get(key) is not None]
    if not values:
        return None
    minimum, maximum = min(values), max(values)
    padding = max((maximum - minimum) * Decimal("0.12"), Decimal("0.001"))
    minimum -= padding; maximum += padding
    left, right, bottom, top = 34, width - 18, 25, height - 15
    plot_height = top - bottom
    plot_width = right - left
    drawing = Drawing(width, height)
    drawing.add(Line(left, bottom, right, bottom, strokeColor=colors.HexColor("#CBD7E2"), strokeWidth=0.6))
    drawing.add(Line(left, bottom, left, top, strokeColor=colors.HexColor("#CBD7E2"), strokeWidth=0.6))
    drawing.add(String(2, bottom - 2, rate(minimum), fontName=styles["DuePdfChartLabel"].fontName,
                      fontSize=6.5, fillColor=colors.HexColor("#557084")))
    drawing.add(String(2, top - 2, rate(maximum), fontName=styles["DuePdfChartLabel"].fontName,
                      fontSize=6.5, fillColor=colors.HexColor("#557084")))
    scale_x = plot_width / Decimal("720")
    series = (("ptax_valor", "#1769AA"), ("fechamento_valor", "#2F9E54"))
    for key, color in series:
        segment = []
        for point in points:
            value = point.get(key)
            if value is None:
                if len(segment) > 1:
                    drawing.add(PolyLine(segment, strokeColor=colors.HexColor(color), strokeWidth=1.8))
                segment = []
                continue
            x = float(left + (Decimal(str(point["x"])) - Decimal("40")) * scale_x)
            y = float(bottom + ((value - minimum) / (maximum - minimum)) * Decimal(str(plot_height)))
            segment.append((x, y))
        if len(segment) > 1:
            drawing.add(PolyLine(segment, strokeColor=colors.HexColor(color), strokeWidth=1.8))
    for point in points:
        if not point.get("show_label"):
            continue
        x = float(left + (Decimal(str(point["x"])) - Decimal("40")) * scale_x)
        drawing.add(String(x, 8, str(point.get("label") or "-"),
                          fontName=styles["DuePdfChartLabel"].fontName, fontSize=6.5,
                          textAnchor="middle", fillColor=colors.HexColor("#557084")))
    return drawing

def report_pdf_html_metadata(source_html):
    from html.parser import HTMLParser

    class MetadataParser(HTMLParser):
        def __init__(self):
            super().__init__(convert_charrefs=True)
            self.in_title = False
            self.in_h1 = False
            self.title = []
            self.h1 = []

        def handle_starttag(self, tag, attrs):
            self.in_title = self.in_title or tag == "title"
            self.in_h1 = self.in_h1 or tag == "h1"

        def handle_endtag(self, tag):
            if tag == "title":
                self.in_title = False
            if tag == "h1":
                self.in_h1 = False

        def handle_data(self, data):
            if self.in_title:
                self.title.append(data)
            if self.in_h1:
                self.h1.append(data)

    parser = MetadataParser()
    parser.feed(source_html)
    return " ".join("".join(parser.h1).split()) or " ".join("".join(parser.title).split())

def render_contract_report_pdf(rendered_html, context):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Spacer

    styles, regular, bold = report_pdf_styles()
    title = report_pdf_html_metadata(rendered_html) or "PAINEL – FECHAMENTOS DE CÂMBIO DE PRONTO"
    output = io.BytesIO()
    page_width, page_height = landscape(A4)
    document = SimpleDocTemplate(output, pagesize=(page_width, page_height),
                                 leftMargin=24, rightMargin=24, topMargin=22, bottomMargin=24,
                                 title=title, author="DUE Control")
    story = [report_pdf_paragraph(title, styles["DuePdfTitle"], bold=True),
             report_pdf_paragraph(
                 f"Safra: {context['safra_label']}  |  Empresa: {context['empresa_label']}  |  "
                 f"Período de apuração: {context['periodo_label']}  |  "
                 f"Visualização: {context['agrupamento_label']}", styles["DuePdfSubtitle"])]
    available_width = page_width - 48
    report_pdf_kpis(story, context["kpis"], styles, available_width)
    story.append(Spacer(1, 2))

    report_pdf_table(story, "Fechamentos por Empresa",
                     ["Empresa", "Total USD", "Total BRL", "Participação %"],
                     [[item["nome"], money(item["total_usd"]), money(item["total_brl"]), report_pdf_percent(item["participacao"])]
                      for item in context["por_empresa"]],
                     [available_width * .43, available_width * .18, available_width * .21, available_width * .18], styles,
                     accent="#EAF3FA")
    report_pdf_table(story, "Fechamentos por Cliente",
                     ["Cliente", "Total USD", "Total BRL", "Participação %"],
                     [[item["nome"], money(item["total_usd"]), money(item["total_brl"]), report_pdf_percent(item["participacao"])]
                      for item in context["por_cliente"]],
                     [available_width * .43, available_width * .18, available_width * .21, available_width * .18], styles,
                     accent="#FFF5D6")
    report_pdf_table(story, "Fechamentos por Banco Recebedor",
                     ["Banco", "Total BRL", "Participação %"],
                     [[item["nome"], money(item["total_brl"]), report_pdf_percent(item["participacao"])]
                      for item in context["por_banco"]],
                     [available_width * .61, available_width * .21, available_width * .18], styles,
                     accent="#E5F4E8")

    report_pdf_table(story, "Estatísticas",
                     ["Indicador", "Valor"],
                     [["Dias acima da PTAX", context["stats"]["dias_acima"]],
                      ["Dias abaixo da PTAX", context["stats"]["dias_abaixo"]],
                      ["Total de dias com operação", context["stats"]["dias_operacao"]],
                      ["Resultado acumulado", money(context["stats"]["resultado_acumulado"]) if context["stats"]["resultado_acumulado"] is not None else "-"]],
                     [available_width * .68, available_width * .32], styles, accent="#FFF5D6")
    report_pdf_table(story, f"Fechamentos por período — {context['agrupamento_label']}",
                     ["Período", "Moeda", "Operações", "Volume", "Taxa ponderada", "Resultado acumulado", "Resultado (R$)"],
                     [[item["rotulo"], item["moeda"], item["resumo"]["contratos"], money(item["resumo"]["volume"]),
                       rate(item["resumo"]["taxa_ponderada"]) if item["resumo"]["taxa_ponderada"] is not None else "-",
                       money(item["resumo"]["resultado_acumulado"]) if item["resumo"].get("resultado_acumulado") is not None else "-",
                       money(item["resumo"]["resultado"]) if item["resumo"]["resultado"] is not None else "-"]
                      for item in context["periodos"]],
                     [available_width * .19, available_width * .09, available_width * .10, available_width * .16,
                      available_width * .14, available_width * .17, available_width * .15], styles,
                     accent="#EAF3FA")

    story.append(report_pdf_paragraph("Evolução por período — Taxa ponderada × PTAX BACEN", styles["DuePdfSection"], bold=True))
    for grafico in context["graficos"]:
        story.append(report_pdf_paragraph(f"{grafico['moeda']} — {grafico['granularidade_label']}", styles["DuePdfSmall"], bold=True))
        drawing = report_pdf_chart_drawing(grafico, styles)
        if drawing:
            story.append(drawing)
            story.append(report_pdf_paragraph("PTAX BACEN   •   Taxa média ponderada pelo volume", styles["DuePdfSmall"]))
        else:
            story.append(report_pdf_paragraph("Não há taxas suficientes para desenhar o gráfico.", styles["DuePdfSmall"]))

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont(regular, 7)
        canvas.setFillColor(colors.HexColor("#557084"))
        canvas.drawString(24, 12, f"Emitido em {context['data_emissao']}")
        canvas.drawRightString(page_width - 24, 12, f"Página {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=footer, onLaterPages=footer)
    output.seek(0)
    return output

@app.route("/contratos/relatorios/pdf")
def relatorios_contratos_pdf():
    agrupamento = (request.args.get("agrupamento", "auto") or "auto").strip().lower()
    if agrupamento not in {"auto", *REPORT_GRANULARITIES}:
        return "Granularidade inválida. Use diário, semanal ou mensal.", 400
    try:
        context = build_contract_report_context(request.args, forced_granularity=agrupamento)
    except ValueError as exc:
        return str(exc), 400
    rendered_html = render_template("relatorios_fechamentos_pdf.html", **context)
    output = render_contract_report_pdf(rendered_html, context)
    filename = f"fechamentos_cambio_{context['agrupamento']}.pdf"
    response = send_file(output, as_attachment=False, download_name=filename, mimetype="application/pdf")
    response.headers["Content-Disposition"] = f'inline; filename="{filename}"'
    return response

@app.route("/derivativos")
def dashboard_derivativos():
    hoje = date.today().isoformat()
    conn = db()
    resumo_row = conn.execute("""
        SELECT COUNT(*) AS ativas,
               COALESCE(SUM(CASE WHEN moeda='USD' THEN valor_contratado ELSE 0 END), 0) AS usd_contratado,
               COALESCE(SUM(CASE WHEN moeda='USD' AND data_vencimento >= ?
                                 THEN valor_contratado ELSE 0 END), 0) AS usd_a_vencer
        FROM ndfs
        WHERE status=?
    """, (hoje, NDF_STATUS_ATIVA)).fetchone()
    proximos = [decorate_ndf(row) for row in conn.execute("""
        SELECT *
        FROM ndfs
        WHERE status=?
        ORDER BY CASE WHEN data_vencimento < ? THEN 0 ELSE 1 END,
                 data_vencimento ASC, id DESC
        LIMIT 5
    """, (NDF_STATUS_ATIVA, hoje)).fetchall()]
    conn.close()
    resumo = {
        "ativas": resumo_row["ativas"],
        "usd_contratado": decimal_value(resumo_row["usd_contratado"]),
        "usd_a_vencer": decimal_value(resumo_row["usd_a_vencer"]),
    }
    return render_template("derivativos.html", resumo=resumo, proximos=proximos)

@app.route("/derivativos/ndfs")
def lista_ndfs():
    conn = db()
    ndfs = [decorate_ndf(row) for row in conn.execute(
        "SELECT * FROM ndfs ORDER BY id DESC"
    ).fetchall()]
    conn.close()
    return render_template("ndfs.html", ndfs=ndfs)

@app.route("/ndf/novo", methods=["GET", "POST"])
def novo_ndf():
    conn = db()
    if request.method == "POST":
        try:
            data = parse_ndf_form(request.form, conn)
            conn.execute("""
                INSERT INTO ndfs
                    (numero_operacao,cnpj,cliente_id,contraparte_id,contraparte,tipo,moeda,valor_contratado,
                     taxa_contratada,data_contratacao,data_vencimento,posicao,
                     finalidade,observacao,status,competencia_id)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                data["numero_operacao"], data["cnpj"], data["cliente_id"], data["contraparte_id"],
                data["contraparte"], data["tipo"],
                data["moeda"], data["valor_contratado"], data["taxa_contratada"],
                data["data_contratacao"], data["data_vencimento"], data["posicao"],
                data["finalidade"], data["observacao"], data["status"], data["competencia_id"],
            ))
            ndf_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.commit()
            conn.close()
            flash("NDF cadastrada com sucesso.", "success")
            return redirect(url_for("detalhe_ndf", ndf_id=ndf_id))
        except sqlite3.IntegrityError:
            conn.rollback()
            flash("Já existe uma NDF com esse número/ID de operação.", "danger")
        except ValueError as exc:
            conn.rollback()
            flash(str(exc), "danger")
    empresas, empresa_id = empresas_for_form(
        conn, selected_id=request.form.get("empresa_id") if request.method == "POST" else None
    )
    clientes = clientes_for_form(conn)
    contrapartes = contrapartes_for_form(conn)
    cliente_id = form_record_id(request.form.get("cliente_id")) if request.method == "POST" else None
    contraparte_id = form_record_id(request.form.get("contraparte_id")) if request.method == "POST" else None
    competencias = competencias_for_empresa(conn, empresa_id)
    competencia_id = form_record_id(request.form.get("competencia_id")) if request.method == "POST" else None
    if not competencia_id and empresa_id:
        sugerida = sugerir_competencia(conn, empresa_id, request.form.get("data_contratacao") or date.today().isoformat())
        competencia_id = sugerida["id"] if sugerida else None
    conn.close()
    return render_template("ndf_form.html", ndf=None, empresas=empresas, empresa_id=empresa_id,
                           clientes=clientes, cliente_id=cliente_id,
                           contrapartes=contrapartes, contraparte_id=contraparte_id,
                           competencias=competencias, competencia_id=competencia_id,
                           ndf_tipos=NDF_TIPOS, ndf_posicoes=NDF_POSICOES, ndf_statuses=NDF_STATUSES)

@app.route("/ndf/<int:ndf_id>")
def detalhe_ndf(ndf_id):
    conn = db()
    ndf = conn.execute("""
        SELECT n.*, c.nome AS cliente_nome, c.pais AS cliente_pais
        FROM ndfs n
        LEFT JOIN clientes c ON c.id=n.cliente_id
        WHERE n.id=?
    """, (ndf_id,)).fetchone()
    conn.close()
    if not ndf:
        return "NDF não encontrada", 404
    return render_template("ndf_detalhe.html", ndf=decorate_ndf(ndf))

@app.route("/ndf/<int:ndf_id>/editar", methods=["GET", "POST"])
def editar_ndf(ndf_id):
    conn = db()
    ndf = conn.execute("SELECT * FROM ndfs WHERE id=?", (ndf_id,)).fetchone()
    if not ndf:
        conn.close()
        return "NDF não encontrada", 404
    if request.method == "POST":
        try:
            data = parse_ndf_form(request.form, conn, current=ndf)
            conn.execute("""
                UPDATE ndfs SET numero_operacao=?,cnpj=?,cliente_id=?,contraparte_id=?,contraparte=?,tipo=?,moeda=?,
                    valor_contratado=?,taxa_contratada=?,data_contratacao=?,data_vencimento=?,
                    posicao=?,finalidade=?,observacao=?,status=?,competencia_id=?
                WHERE id=?
            """, (
                data["numero_operacao"], data["cnpj"], data["cliente_id"], data["contraparte_id"],
                data["contraparte"], data["tipo"],
                data["moeda"], data["valor_contratado"], data["taxa_contratada"],
                data["data_contratacao"], data["data_vencimento"], data["posicao"],
                data["finalidade"], data["observacao"], data["status"], data["competencia_id"], ndf_id,
            ))
            conn.commit()
            conn.close()
            flash("NDF atualizada com sucesso.", "success")
            return redirect(url_for("detalhe_ndf", ndf_id=ndf_id))
        except sqlite3.IntegrityError:
            conn.rollback()
            flash("Já existe uma NDF com esse número/ID de operação.", "danger")
        except ValueError as exc:
            conn.rollback()
            flash(str(exc), "danger")
    empresas, empresa_id = empresas_for_form(
        conn, ndf["cnpj"], request.form.get("empresa_id") if request.method == "POST" else None
    )
    clientes = clientes_for_form(conn)
    contrapartes = contrapartes_for_form(conn)
    empresa_id = empresa_id_por_cnpj(conn, ndf["cnpj"])
    competencias = competencias_for_empresa(conn, empresa_id)
    competencia_id = form_record_id(request.form.get("competencia_id"), ndf["competencia_id"] if "competencia_id" in ndf.keys() else None)
    cliente_id = form_record_id(
        request.form.get("cliente_id") if request.method == "POST" else ndf["cliente_id"]
    )
    contraparte_id = form_record_id(
        request.form.get("contraparte_id") if request.method == "POST" else ndf["contraparte_id"]
    )
    ndf_data = decorate_ndf(ndf)
    conn.close()
    return render_template("ndf_form.html", ndf=ndf_data, empresas=empresas, empresa_id=empresa_id,
                           clientes=clientes, cliente_id=cliente_id,
                           contrapartes=contrapartes, contraparte_id=contraparte_id,
                           competencias=competencias, competencia_id=competencia_id,
                           ndf_tipos=NDF_TIPOS, ndf_posicoes=NDF_POSICOES, ndf_statuses=NDF_STATUSES)

@app.route("/derivativos/ndfs/excluir-lote", methods=["POST"])
def excluir_ndfs_lote():
    conn = db()
    try:
        ndf_ids = selected_record_ids(request.form)
        conn.execute("BEGIN IMMEDIATE")
        ensure_existing_record_ids(conn, "ndfs", ndf_ids, "NDFs")
        placeholders = ",".join("?" for _ in ndf_ids)
        conn.execute(f"DELETE FROM ndfs WHERE id IN ({placeholders})", ndf_ids)
        conn.commit()
        flash(f"{len(ndf_ids)} NDF(s) excluída(s) com sucesso.", "success")
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), "danger")
    except (sqlite3.Error, OverflowError):
        conn.rollback()
        flash("Não foi possível concluir a exclusão das NDFs.", "danger")
    except Exception:
        conn.rollback()
        flash("Não foi possível concluir a exclusão das NDFs.", "danger")
    finally:
        conn.close()
    return redirect_batch_result("lista_ndfs")

@app.route("/derivativos/ptax")
def ptax():
    return render_ptax_page()

@app.route("/derivativos/ptax/consultar", methods=["POST"])
def consultar_ptax():
    consulta = {
        "moeda": request.form.get("moeda"),
        "data_inicial": request.form.get("data_inicial"),
        "data_final": request.form.get("data_final"),
    }
    try:
        moeda, data_inicial, data_final = parse_ptax_period(request.form)
        previsao = consultar_ptax_api(moeda, data_inicial, data_final)
        return render_ptax_page(previsao=previsao, consulta=consulta)
    except ValueError as exc:
        flash(str(exc), "danger")
        return render_ptax_page(consulta=consulta)

@app.route("/derivativos/ptax/importar", methods=["POST"])
def importar_ptax():
    conn = db()
    try:
        registros = parse_ptax_importacao(request.form)
        conn.execute("BEGIN IMMEDIATE")
        conn.executemany("""
            INSERT INTO ptax_cotacoes (data_cotacao,moeda,ptax_compra,ptax_venda)
            VALUES (?,?,?,?)
            ON CONFLICT(moeda,data_cotacao) DO UPDATE SET
                ptax_compra=excluded.ptax_compra,
                ptax_venda=excluded.ptax_venda
        """, [(
            registro["data_cotacao"], registro["moeda"],
            registro["ptax_compra"], registro["ptax_venda"],
        ) for registro in registros])
        conn.commit()
        flash(f"{len(registros)} cotação(ões) PTAX importada(s)/atualizada(s) com sucesso.", "success")
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), "danger")
    except sqlite3.Error:
        conn.rollback()
        flash("Não foi possível gravar as cotações PTAX.", "danger")
    finally:
        conn.close()
    return redirect(url_for("ptax"))

@app.route("/configuracoes")
def configuracoes():
    return render_template("configuracoes.html")

@app.route("/configuracoes/empresas", methods=["GET", "POST"])
def cadastro_empresas():
    conn = db()
    if request.method == "POST":
        try:
            razao_social = (request.form.get("razao_social") or "").strip()
            if not razao_social:
                raise ValueError("A Razão Social é obrigatória.")
            cnpj = normalize_cnpj(request.form.get("cnpj"))
            apelido = (request.form.get("apelido") or "").strip() or None
            conn.execute(
                "INSERT INTO empresas (razao_social, cnpj, apelido) VALUES (?,?,?)",
                (razao_social, cnpj, apelido),
            )
            conn.commit()
            conn.close()
            flash("Empresa cadastrada com sucesso.", "success")
            return redirect(url_for("cadastro_empresas"))
        except sqlite3.IntegrityError:
            conn.rollback()
            flash("Já existe uma empresa cadastrada com este CNPJ.", "danger")
        except ValueError as exc:
            conn.rollback()
            flash(str(exc), "danger")
    empresas = conn.execute("""
        SELECT id, razao_social, cnpj, apelido
        FROM empresas
        ORDER BY CASE WHEN TRIM(COALESCE(apelido, '')) <> '' THEN 0 ELSE 1 END,
                 apelido, razao_social
    """).fetchall()
    conn.close()
    return render_template("empresas.html", empresas=empresas)

@app.route("/configuracoes/clientes", methods=["GET", "POST"])
def cadastro_clientes():
    conn = db()
    nome = normalize_client_name_display(request.form.get("nome")) if request.method == "POST" else ""
    pais_selecionado = request.form.get("pais") if request.method == "POST" else ""
    if request.method == "POST":
        try:
            if not nome:
                raise ValueError("O nome do cliente é obrigatório.")
            pais = normalize_pais(pais_selecionado)
            existing_clients = conn.execute(
                "SELECT id, nome FROM clientes WHERE pais=?", (pais,)
            ).fetchall()
            if any(normalize_client_name_key(client["nome"]) == normalize_client_name_key(nome)
                   for client in existing_clients):
                raise ValueError("Já existe um cliente equivalente cadastrado para este país.")
            conn.execute(
                "INSERT INTO clientes (nome, pais) VALUES (?, ?)",
                (nome, pais),
            )
            conn.commit()
            conn.close()
            flash("Cliente cadastrado com sucesso.", "success")
            return redirect(url_for("cadastro_clientes"))
        except sqlite3.IntegrityError:
            conn.rollback()
            flash("Já existe um cliente cadastrado com este nome e país.", "danger")
        except ValueError as exc:
            conn.rollback()
            flash(str(exc), "danger")
    clientes = conn.execute("""
        SELECT id, nome, pais
        FROM clientes
        ORDER BY nome, pais
    """).fetchall()
    conn.close()
    return render_template(
        "clientes.html", clientes=clientes, paises=CLIENTE_PAISES_ORDENADOS,
        nome=nome, pais_selecionado=pais_selecionado,
    )

@app.route("/configuracoes/contrapartes", methods=["GET", "POST"])
def cadastro_contrapartes():
    conn = db()
    nome = (request.form.get("nome") or "").strip() if request.method == "POST" else ""
    if request.method == "POST":
        try:
            if not nome:
                raise ValueError("O nome do Banco / Contraparte é obrigatório.")
            conn.execute(
                "INSERT INTO contrapartes (nome) VALUES (?)",
                (nome,),
            )
            conn.commit()
            conn.close()
            flash("Banco / Contraparte cadastrado com sucesso.", "success")
            return redirect(url_for("cadastro_contrapartes"))
        except sqlite3.IntegrityError:
            conn.rollback()
            flash("Já existe um Banco / Contraparte cadastrado com este nome.", "danger")
        except ValueError as exc:
            conn.rollback()
            flash(str(exc), "danger")
    contrapartes = conn.execute("""
        SELECT id, nome
        FROM contrapartes
        ORDER BY nome
    """).fetchall()
    conn.close()
    return render_template("contrapartes.html", contrapartes=contrapartes, nome=nome)

@app.route("/configuracoes/competencias", methods=["GET", "POST"])
def cadastro_competencias():
    conn = db()
    competencia = None
    if request.method == "POST":
        try:
            data = competencia_data(request.form, conn)
            salvar_competencia(conn, data)
            flash("Competência cadastrada com sucesso.", "success")
            conn.close()
            return redirect(url_for("cadastro_competencias"))
        except sqlite3.IntegrityError:
            conn.rollback(); flash("Já existe uma competência com essa descrição para a empresa selecionada.", "danger")
        except ValueError as exc:
            conn.rollback(); flash(str(exc), "danger")
        except sqlite3.Error as exc:
            conn.rollback(); flash(f"Não foi possível salvar a competência: {exc}", "danger")
        competencia = dict(request.form)
    competencias = conn.execute("""SELECT c.id, c.empresa_id, c.descricao, c.data_inicial, c.data_final, c.status,
        e.razao_social, e.apelido FROM competencias c JOIN empresas e ON e.id=c.empresa_id
        ORDER BY c.data_inicial DESC, c.descricao""").fetchall()
    empresas = conn.execute("SELECT id, razao_social, apelido, cnpj FROM empresas ORDER BY razao_social").fetchall()
    conn.close()
    return render_template("competencias.html", competencia=competencia, competencias=competencias, empresas=empresas, statuses=COMPETENCIA_STATUSES)

@app.route("/configuracoes/competencias/<int:competencia_id>/editar", methods=["GET", "POST"])
def editar_competencia(competencia_id):
    conn = db()
    competencia = conn.execute("SELECT * FROM competencias WHERE id=?", (competencia_id,)).fetchone()
    if not competencia:
        conn.close(); return "Competência não encontrada", 404
    if request.method == "POST":
        try:
            data = competencia_data(request.form, conn, current=competencia)
            salvar_competencia(conn, data, competencia_id=competencia_id)
            flash("Competência atualizada com sucesso.", "success")
            conn.close()
            return redirect(url_for("cadastro_competencias"))
        except sqlite3.IntegrityError:
            conn.rollback(); flash("Já existe uma competência com essa descrição para a empresa selecionada.", "danger")
        except ValueError as exc:
            conn.rollback(); flash(str(exc), "danger")
        except sqlite3.Error as exc:
            conn.rollback(); flash(f"Não foi possível salvar a competência: {exc}", "danger")
        competencia = dict(request.form); competencia["id"] = competencia_id
    empresas = conn.execute("SELECT id, razao_social, apelido, cnpj FROM empresas ORDER BY razao_social").fetchall()
    competencias = conn.execute("""SELECT c.id, c.empresa_id, c.descricao, c.data_inicial, c.data_final, c.status,
        e.razao_social, e.apelido FROM competencias c JOIN empresas e ON e.id=c.empresa_id
        ORDER BY c.data_inicial DESC, c.descricao""").fetchall()
    conn.close()
    return render_template("competencias.html", competencia=competencia, competencias=competencias, empresas=empresas, statuses=COMPETENCIA_STATUSES)

@app.route("/configuracoes/competencias/<int:competencia_id>/encerrar", methods=["POST"])
def encerrar_competencia(competencia_id):
    conn = db()
    if not conn.execute("SELECT id FROM competencias WHERE id=?", (competencia_id,)).fetchone():
        conn.close(); return "Competência não encontrada", 404
    conn.execute("UPDATE competencias SET status=? WHERE id=?", (COMPETENCIA_STATUS_ENCERRADA, competencia_id))
    conn.commit(); conn.close()
    flash("Competência encerrada com sucesso.", "success")
    return redirect(url_for("cadastro_competencias"))

@app.route("/configuracoes/competencias/<int:competencia_id>/excluir", methods=["POST"])
def excluir_competencia(competencia_id):
    conn = db()
    if not conn.execute("SELECT id FROM competencias WHERE id=?", (competencia_id,)).fetchone():
        conn.close(); return "Competência não encontrada", 404
    conn.execute("DELETE FROM competencias WHERE id=?", (competencia_id,))
    conn.commit(); conn.close()
    flash("Competência excluída com sucesso.", "success")
    return redirect(url_for("cadastro_competencias"))

def optional_number(value):
    return parse_number(value) if value and str(value).strip() else None

@app.route("/contrato/novo", methods=["GET", "POST"])
def novo_contrato():
    return redirect(url_for("lista_invoices"))
    if request.method == "POST":
        f = request.form
        conn = db()
        try:
            valor_moeda = parse_number(f.get("valor_moeda"))
            ensure_non_negative_balance(valor_moeda, "O valor do contrato não pode ser negativo.")
            banco_id, banco, banco_liquidacao_id, banco_liquidacao = bancos_do_contrato(conn, f)
            cliente_id, cliente = cliente_do_contrato(
                conn, f.get("cliente_id"), legacy_name=f.get("cliente")
            )
            cnpj = cnpj_da_empresa(conn, f.get("empresa_id"))
            empresa_id = form_record_id(f.get("empresa_id"), empresa_id_por_cnpj(conn, cnpj))
            data_contrato = parse_date(f.get("data_contrato"))
            data_recebimento = parse_date(f.get("data_recebimento"))
            data_liquidacao = parse_date(f.get("data_liquidacao"))
            competencia_id = competencia_da_operacao(conn, f.get("competencia_id"), empresa_id, data_contrato or date.today().isoformat())
            conn.execute("""INSERT INTO contratos
                (numero_contrato,banco_id,banco,banco_credito,banco_liquidacao,data_contrato,data_recebimento,data_liquidacao,cnpj,cliente_id,cliente,moeda,valor_moeda,taxa_cambio,valor_reais,status,observacao,competencia_id)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (f["numero_contrato"].strip(), banco_id, banco, banco, banco_liquidacao,
                 data_contrato, data_recebimento, data_liquidacao, cnpj, cliente_id, cliente,
                 f.get("moeda") or "USD", valor_moeda,
                 optional_number(f.get("taxa_cambio")), optional_number(f.get("valor_reais")),
                 status_from_balance(valor_moeda), f.get("observacao"), competencia_id))
            conn.commit(); conn.close()
            flash("Contrato Câmbio cadastrado com sucesso.", "success")
            return redirect(url_for("lista_contratos"))
        except sqlite3.IntegrityError:
            conn.close()
            flash("Já existe um Contrato Câmbio com esse número.", "danger")
        except ValueError as exc:
            conn.close()
            flash(str(exc), "danger")
    conn = db()
    empresas, empresa_id = empresas_for_form(conn, selected_id=request.form.get("empresa_id"))
    contrapartes = contrapartes_for_form(conn)
    banco_credito_id, banco_liquidacao_id = banco_ids_for_contrato_form(
        conn, credito_id=request.form.get("banco_credito_id") or request.form.get("banco_id"),
        liquidacao_id=request.form.get("banco_liquidacao_id"),
    )
    clientes = clientes_for_form(conn)
    cliente_id = cliente_id_for_form(selected_id=request.form.get("cliente_id"))
    competencias = competencias_for_empresa(conn, empresa_id)
    competencia_id = form_record_id(request.form.get("competencia_id"))
    if not competencia_id and empresa_id:
        sugerida = sugerir_competencia(conn, empresa_id, request.form.get("data_contrato") or date.today().isoformat())
        competencia_id = sugerida["id"] if sugerida else None
    conn.close()
    return render_template("contrato_form.html", contrato=None, empresas=empresas, empresa_id=empresa_id,
                           contrapartes=contrapartes, banco_id=banco_credito_id,
                           banco_credito_id=banco_credito_id, banco_liquidacao_id=banco_liquidacao_id,
                           clientes=clientes, cliente_id=cliente_id,
                           competencias=competencias, competencia_id=competencia_id)

def carregar_detalhe_contrato(conn, contrato_id):
    contrato = conn.execute("""
        SELECT c.*, cl.pais AS cliente_pais,
               e.razao_social AS empresa_razao_social,
               e.apelido AS empresa_apelido
        FROM contratos c
        LEFT JOIN clientes cl ON cl.id=c.cliente_id
        LEFT JOIN empresas e ON e.cnpj=c.cnpj
        WHERE c.id=?
    """, (contrato_id,)).fetchone()
    if not contrato:
        return None
    vinculos = conn.execute("""
        SELECT v.*, m.id AS movimentacao_id, m.valor AS valor_movimentacao,
               d.chave_acesso, d.numero_due, d.created_at AS data_lancamento,
               d.cliente AS cliente_due
        FROM due_contratos v JOIN dues d ON d.id=v.due_id
        LEFT JOIN due_movimentacoes m ON m.due_contrato_id=v.id AND m.tipo='VINCULACAO'
        WHERE v.contrato_id=? ORDER BY v.id DESC
    """, (contrato_id,)).fetchall()
    invoice_links = conn.execute("""
        SELECT v.*, i.numero_invoice, i.tipo_documento, i.data_emissao, i.moeda AS invoice_moeda,
               e.apelido AS empresa_apelido, e.razao_social AS empresa_razao_social,
               cl.nome AS cliente_nome
        FROM invoice_contrato_cambio v
        JOIN invoices i ON i.id=v.invoice_id
        JOIN empresas e ON e.id=i.empresa_id
        LEFT JOIN clientes cl ON cl.id=i.cliente_id
        WHERE v.contrato_id=? ORDER BY v.id DESC
    """, (contrato_id,)).fetchall()
    summary = contract_summary(conn, contrato_id)
    contrato = dict(contrato)
    contrato["valor_moeda"] = summary["valor_moeda"]
    contrato.update({"vinculado": summary["vinculado"], "saldo": summary["saldo"], "status": summary["status"],
                     "invoice_links": invoice_links})
    return contrato, vinculos, summary

@app.route("/contrato/<int:contrato_id>")
def detalhe_contrato(contrato_id):
    conn = db()
    dados = carregar_detalhe_contrato(conn, contrato_id)
    conn.close()
    if not dados:
        return "Contrato Câmbio não encontrado", 404
    contrato, vinculos, summary = dados
    return render_template("contrato_detalhe.html", contrato=contrato, vinculos=vinculos,
                           vinculado=summary["vinculado"], saldo=summary["saldo"])

@app.route("/contrato/<int:contrato_id>/relatorio")
def relatorio_contrato(contrato_id):
    conn = db()
    dados = carregar_detalhe_contrato(conn, contrato_id)
    conn.close()
    if not dados:
        return "Contrato Câmbio não encontrado", 404
    contrato, vinculos, summary = dados
    return render_template("contrato_relatorio.html", contrato=contrato, vinculos=vinculos,
                           vinculado=summary["vinculado"], saldo=summary["saldo"])

@app.route("/contrato/<int:contrato_id>/editar", methods=["GET", "POST"])
def editar_contrato(contrato_id):
    conn = db()
    contrato = conn.execute("SELECT * FROM contratos WHERE id=?", (contrato_id,)).fetchone()
    if not contrato:
        conn.close(); return "Contrato Câmbio não encontrado", 404
    resumo = contract_summary(conn, contrato_id)
    if request.method == "GET":
        contrapartes = contrapartes_for_form(conn)
        conn.close()
        return render_template("contrato_form_derived.html", contrato=contrato, resumo=resumo,
                               contrapartes=contrapartes)
    if request.form.get("derived_contract_form") == "1":
        try:
            metadata = contract_metadata_from_form(request.form, conn)
            numero = metadata["numero_contrato"]
            if not numero:
                raise ValueError("O número do Contrato Câmbio é obrigatório.")
            duplicate = conn.execute("SELECT id FROM contratos WHERE numero_contrato=? AND id<>?", (numero, contrato_id)).fetchone()
            if duplicate:
                raise ValueError("Já existe um Contrato Câmbio com esse número.")
            if metadata["data_fechamento"] and metadata["data_liquidacao"] and metadata["data_liquidacao"] < metadata["data_fechamento"]:
                raise ValueError("A data de liquidação não pode ser anterior ao fechamento.")
            conn.execute("""
                UPDATE contratos SET numero_contrato=?, banco_liquidacao_id=?, banco_liquidacao=?,
                    data_fechamento=?, data_liquidacao=?, data_contrato=?, taxa_cambio=?, observacao=?
                WHERE id=?
            """, (numero, metadata["banco_liquidacao_id"], metadata["banco_liquidacao"],
                  metadata["data_fechamento"], metadata["data_liquidacao"], metadata["data_fechamento"],
                  metadata["taxa_cambio"], metadata["observacao"], contrato_id))
            sync_contract_cache(conn, contrato_id)
            conn.commit(); conn.close()
            flash("Contrato Câmbio atualizado com sucesso.", "success")
            return redirect(url_for("detalhe_contrato", contrato_id=contrato_id))
        except (ValueError, sqlite3.Error) as exc:
            conn.rollback(); flash(str(exc), "danger")
            contrapartes = contrapartes_for_form(conn)
            conn.close()
            return render_template("contrato_form_derived.html", contrato=dict(contrato), resumo=resumo,
                                   contrapartes=contrapartes), 400
    conn.close()
    flash("Contratos Câmbio são mantidos pelas Invoices vinculadas; edite-os a partir do fluxo de Invoice.", "danger")
    return redirect(url_for("detalhe_contrato", contrato_id=contrato_id))
    if request.method == "POST":
        f = request.form
        try:
            valor_moeda = parse_number(f.get("valor_moeda"))
            linked = resumo["vinculado"]
            ensure_non_negative_balance(
                decimal_value(valor_moeda) - linked,
                "O valor do contrato não pode ficar abaixo do total já vinculado."
            )
            banco_id, banco, banco_liquidacao_id, banco_liquidacao = bancos_do_contrato(conn, f, current=contrato)
            cliente_id, cliente = cliente_do_contrato(
                conn, f.get("cliente_id"), current=contrato, legacy_name=f.get("cliente")
            )
            cnpj = cnpj_da_empresa(conn, f.get("empresa_id"), contrato["cnpj"])
            empresa_id = form_record_id(f.get("empresa_id"), empresa_id_por_cnpj(conn, cnpj))
            data_contrato = parse_date(f.get("data_contrato"))
            data_recebimento = parse_date(f.get("data_recebimento"))
            data_liquidacao = parse_date(f.get("data_liquidacao"))
            competencia_id = competencia_da_operacao(conn, f.get("competencia_id"), empresa_id, data_contrato or date.today().isoformat(), contrato["competencia_id"] if "competencia_id" in contrato.keys() else None)
            conn.execute("""UPDATE contratos SET numero_contrato=?,banco_id=?,banco=?,banco_credito=?,banco_liquidacao=?,data_contrato=?,data_recebimento=?,data_liquidacao=?,cnpj=?,cliente_id=?,cliente=?,moeda=?,
                            valor_moeda=?,taxa_cambio=?,valor_reais=?,status=?,observacao=?,competencia_id=? WHERE id=?""",
                (f["numero_contrato"].strip(), banco_id, banco, banco, banco_liquidacao, data_contrato, data_recebimento, data_liquidacao, cnpj,
                 cliente_id, cliente, f.get("moeda") or "USD", valor_moeda,
                 optional_number(f.get("taxa_cambio")), optional_number(f.get("valor_reais")),
                 STATUS_CONCLUIDO if resumo["saldo_zerado_manual"] else status_from_balance(contract_balance(valor_moeda, linked), linked),
                 f.get("observacao"), competencia_id, contrato_id))
            conn.commit(); conn.close()
            flash("Contrato Câmbio atualizado com sucesso.", "success")
            return redirect(url_for("detalhe_contrato", contrato_id=contrato_id))
        except sqlite3.IntegrityError:
            flash("Já existe um Contrato Câmbio com esse número.", "danger")
        except ValueError as exc:
            flash(str(exc), "danger")
    empresas, empresa_id = empresas_for_form(
        conn, contrato["cnpj"], request.form.get("empresa_id") if request.method == "POST" else None
    )
    contrapartes = contrapartes_for_form(conn)
    banco_credito_id, banco_liquidacao_id = banco_ids_for_contrato_form(
        conn, contrato,
        request.form.get("banco_credito_id") if request.method == "POST" else None,
        request.form.get("banco_liquidacao_id") if request.method == "POST" else None,
    )
    clientes = clientes_for_form(conn)
    cliente_id = cliente_id_for_form(
        contrato,
        request.form.get("cliente_id") if request.method == "POST" else None,
    )
    competencias = competencias_for_empresa(conn, empresa_id)
    competencia_id = form_record_id(request.form.get("competencia_id"), contrato["competencia_id"] if "competencia_id" in contrato.keys() else None)
    conn.close()
    return render_template("contrato_form.html", contrato=contrato, resumo=resumo, empresas=empresas, empresa_id=empresa_id,
                           contrapartes=contrapartes, banco_id=banco_credito_id,
                           banco_credito_id=banco_credito_id, banco_liquidacao_id=banco_liquidacao_id,
                           clientes=clientes, cliente_id=cliente_id,
                           competencias=competencias, competencia_id=competencia_id)

@app.route("/contrato/<int:contrato_id>/zerar-saldo", methods=["POST"])
def zerar_saldo(contrato_id):
    conn = db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        contrato = contract_summary(conn, contrato_id)
        if not contrato:
            conn.rollback()
            return "Contrato Câmbio não encontrado", 404
        if contrato["saldo_zerado_manual"]:
            raise ValueError("O saldo deste Contrato Câmbio já foi zerado manualmente.")
        if contrato["saldo"] <= 0:
            raise ValueError("Só é possível zerar manualmente Contratos Câmbio com saldo positivo.")
        conn.execute("UPDATE contratos SET saldo_zerado_manual=1,status=? WHERE id=?",
                     (STATUS_CONCLUIDO, contrato_id))
        conn.commit()
        flash("Saldo do Contrato Câmbio zerado manualmente e Contrato Câmbio marcado como concluído.", "success")
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), "danger")
    finally:
        conn.close()
    return redirect(url_for("editar_contrato", contrato_id=contrato_id))

@app.route("/contrato/<int:contrato_id>/reverter-saldo", methods=["POST"])
def reverter_saldo(contrato_id):
    conn = db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        contrato = conn.execute("SELECT saldo_zerado_manual FROM contratos WHERE id=?", (contrato_id,)).fetchone()
        if not contrato:
            conn.rollback()
            return "Contrato Câmbio não encontrado", 404
        if not contrato["saldo_zerado_manual"]:
            raise ValueError("Este Contrato Câmbio não possui zeramento manual ativo.")
        conn.execute("UPDATE contratos SET saldo_zerado_manual=0 WHERE id=?", (contrato_id,))
        update_contract_status(conn, contrato_id)
        conn.commit()
        flash("Zeramento manual revertido; saldo recalculado com os vínculos atuais.", "success")
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), "danger")
    finally:
        conn.close()
    return redirect(url_for("editar_contrato", contrato_id=contrato_id))

@app.route("/contratos/<int:contrato_id>/saldo")
def saldo_contrato(contrato_id):
    conn = db()
    contrato = contract_summary(conn, contrato_id)
    conn.close()
    if not contrato:
        return jsonify({"error": "Contrato Câmbio não encontrado."}), 404
    total = float(contrato["valor_moeda"] or 0)
    vinculado = float(contrato["vinculado"] or 0)
    return jsonify({"id": contrato["id"], "numero_contrato": contrato["numero_contrato"],
                    "moeda": contrato["moeda"], "valor_total": total,
                    "total_vinculado": vinculado, "saldo_disponivel": float(contrato["saldo"]),
                    "status": contrato["status"]})

@app.route("/dues")
def consulta_dues():
    filters = {key: value for key, value in request.args.items()
               if key not in {"sort", "direction", "page"} and value}
    where, params = [], []
    if request.args.get("numero_due"):
        where.append("d.numero_due LIKE ?"); params.append(f"%{request.args['numero_due'].strip()}%")
    if request.args.get("chave_acesso"):
        where.append("d.chave_acesso LIKE ?"); params.append(f"%{request.args['chave_acesso'].strip().upper()}%")
    if request.args.get("cliente"):
        where.append("d.cliente LIKE ?"); params.append(f"%{request.args['cliente'].strip()}%")
    if request.args.get("cnpj"):
        where.append("d.cnpj LIKE ?"); params.append(f"%{request.args['cnpj'].strip()}%")
    if request.args.get("moeda"):
        where.append("d.moeda = ?"); params.append(request.args["moeda"].strip().upper())
    if request.args.get("status"):
        where.append("d.status = ?"); params.append(request.args["status"].strip().upper())
    for key, operator in (("data_de", ">="), ("data_ate", "<=")):
        if request.args.get(key):
            try:
                where.append(f"date(d.created_at) {operator} ?"); params.append(parse_date(request.args[key]))
            except ValueError as exc:
                flash(str(exc), "danger")
    for key, operator in (("valor_min", ">="), ("valor_max", "<=")):
        if request.args.get(key):
            try:
                where.append(f"d.valor_original {operator} ?"); params.append(parse_number(request.args[key]))
            except ValueError as exc:
                flash(str(exc), "danger")

    sort_fields = {"chave_acesso": "d.chave_acesso", "numero_due": "d.numero_due", "data_due": "d.created_at",
                   "cliente": "d.cliente", "moeda": "d.moeda",
                   "valor_original": "d.valor_original", "status": "d.status"}
    sort = request.args.get("sort", "data_due")
    sort = sort if sort in sort_fields else "data_due"
    direction = "ASC" if request.args.get("direction", "desc").lower() == "asc" else "DESC"
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    per_page = 20
    clause = (" WHERE " + " AND ".join(where)) if where else ""
    conn = db()
    total = conn.execute(f"SELECT COUNT(*) FROM dues d{clause}", params).fetchone()[0]
    pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, pages)
    dues = [decorate_due(row) for row in conn.execute(f"""
        SELECT d.*, COALESCE(SUM(CASE WHEN m.tipo IN ('UTILIZACAO','VINCULACAO') THEN m.valor ELSE -m.valor END),0) AS utilizado
        FROM dues d LEFT JOIN due_movimentacoes m ON m.due_id=d.id
        {clause} GROUP BY d.id ORDER BY {sort_fields[sort]} {direction}, d.id DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, (page - 1) * per_page]).fetchall()]
    moedas = [row[0] for row in conn.execute("SELECT DISTINCT moeda FROM dues WHERE moeda IS NOT NULL ORDER BY moeda")]
    conn.close()
    sort_links = {}
    for key in sort_fields:
        next_direction = "asc" if sort == key and direction == "DESC" else "desc"
        sort_links[key] = {**filters, "sort": key, "direction": next_direction}
    previous_args = {**filters, "sort": sort, "direction": direction, "page": page - 1}
    next_args = {**filters, "sort": sort, "direction": direction, "page": page + 1}
    return render_template("due_consulta.html", dues=dues, total=total, page=page, pages=pages,
                           filters=filters, moedas=moedas, sort=sort, direction=direction,
                           sort_links=sort_links, previous_args=previous_args, next_args=next_args,
                           status_concluido=STATUS_CONCLUIDO, status_parcial=STATUS_PARCIAL)

@app.route("/dues/excluir-lote", methods=["POST"])
def excluir_dues_lote():
    conn = db()
    try:
        due_ids = selected_record_ids(request.form)
        conn.execute("BEGIN IMMEDIATE")
        ensure_existing_record_ids(conn, "dues", due_ids, "DU-Es")
        placeholders = ",".join("?" for _ in due_ids)
        contrato_rows = conn.execute(f"""
            SELECT DISTINCT contrato_id
            FROM due_movimentacoes
            WHERE due_id IN ({placeholders}) AND contrato_id IS NOT NULL
            UNION
            SELECT DISTINCT contrato_id
            FROM due_contratos
            WHERE due_id IN ({placeholders})
        """, due_ids + due_ids).fetchall()
        contrato_ids = [row[0] for row in contrato_rows]

        conn.execute(f"DELETE FROM due_movimentacoes WHERE due_id IN ({placeholders})", due_ids)
        conn.execute(f"DELETE FROM due_contratos WHERE due_id IN ({placeholders})", due_ids)
        conn.execute(f"DELETE FROM dues WHERE id IN ({placeholders})", due_ids)
        recalculate_statuses(conn, due_ids=[], contrato_ids=contrato_ids)
        conn.commit()
        flash(f"{len(due_ids)} DU-E(s) excluída(s) com sucesso.", "success")
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), "danger")
    except (sqlite3.Error, OverflowError):
        conn.rollback()
        flash("Não foi possível concluir a exclusão das DU-Es.", "danger")
    except Exception:
        conn.rollback()
        flash("Não foi possível concluir a exclusão das DU-Es.", "danger")
    finally:
        conn.close()
    return redirect_batch_result("consulta_dues")

def write_excel_model_orientations(writer, conn, pandas):
    """Adiciona ao modelo uma lista de referência para preenchimento do Excel."""
    banks = [row["nome"] for row in conn.execute(
        "SELECT nome FROM contrapartes ORDER BY nome"
    ).fetchall()]
    clients = [row["nome"] for row in conn.execute(
        "SELECT nome FROM clientes ORDER BY nome, pais"
    ).fetchall()]
    companies = conn.execute("""
        SELECT cnpj
        FROM empresas
        ORDER BY cnpj
    """).fetchall()
    company_cnpjs = [format_cnpj(row["cnpj"]) for row in companies]
    competencies = [row["descricao"] for row in conn.execute("""
        SELECT descricao
        FROM competencias
        ORDER BY data_inicial DESC, descricao
    """).fetchall()]

    orientations = pandas.DataFrame({
        "BANCOS": pandas.Series(banks),
        "CLIENTES": pandas.Series(clients),
        "STATUS": pandas.Series(INVOICE_STATUS_OPTIONS),
        "CNPJ": pandas.Series(company_cnpjs),
        "COMPETENCIAS": pandas.Series(competencies),
    })
    orientations.to_excel(writer, index=False, sheet_name="Orientações")

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    worksheet = writer.book["Orientações"]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1769AA")
        cell.alignment = Alignment(horizontal="center")
    for column in worksheet.columns:
        letter = get_column_letter(column[0].column)
        width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 14), 45)
        worksheet.column_dimensions[letter].width = width

@app.route("/dues/modelo")
def modelo_dues():
    import pandas as pd
    df = pd.DataFrame(columns=["numero_due", "chave_acesso", "cnpj", "cliente", "moeda", "valor_original"])
    out = io.BytesIO()
    conn = db()
    try:
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="DU-Es")
            write_excel_model_orientations(writer, conn, pd)
    finally:
        conn.close()
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="modelo_dues.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/dues/importar", methods=["POST"])
def importar_dues():
    import pandas as pd
    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        flash("Selecione um arquivo Excel de DU-Es.", "danger")
        return redirect(url_for("consulta_dues"))

    try:
        df = pd.read_excel(arquivo)
        df.columns = [str(c).strip().lower() for c in df.columns]
        obrigatorias = {"chave_acesso", "numero_due", "valor_original"}
        faltantes = sorted(obrigatorias - set(df.columns))
        if faltantes:
            flash("O Excel precisa conter as colunas obrigatórias: " + ", ".join(sorted(obrigatorias)) + ".", "danger")
            return redirect(url_for("consulta_dues"))
    except Exception as exc:
        flash(f"Não foi possível ler o arquivo Excel: {exc}", "danger")
        return redirect(url_for("consulta_dues"))

    conn = db()
    existentes = {str(row[0]).strip() for row in conn.execute("SELECT numero_due FROM dues")}
    chaves_existentes = {str(row[0]).strip().upper() for row in conn.execute("SELECT chave_acesso FROM dues WHERE chave_acesso IS NOT NULL AND chave_acesso <> ''")}
    conn.close()
    registros, rejeitados, vistos_numeros, vistos_chaves = [], [], set(), set()

    def valor_linha(row, coluna, default=None):
        value = row[coluna] if coluna in df.columns else default
        return None if pd.isna(value) else value

    for indice, row in df.iterrows():
        linha = indice + 2
        numero = str(valor_linha(row, "numero_due", "") or "").strip()
        chave = str(valor_linha(row, "chave_acesso", "") or "").strip()
        try:
            chave = parse_chave_acesso(chave)
            if not numero:
                raise ValueError("numero_due é obrigatório")
            if numero in existentes:
                raise ValueError("a DU-E já está cadastrada no banco")
            if numero in vistos_numeros:
                raise ValueError("número de DU-E repetido no arquivo")
            if chave in chaves_existentes:
                raise ValueError("a Chave de Acesso já está cadastrada no banco")
            if chave in vistos_chaves:
                raise ValueError("Chave de Acesso repetida no arquivo")
            valor_bruto = valor_linha(row, "valor_original")
            if valor_bruto is None or str(valor_bruto).strip() == "":
                raise ValueError("valor_original é obrigatório")
            valor_original = parse_number(valor_bruto)
            ensure_non_negative_balance(valor_original, "valor_original não pode ser negativo")
            moeda = str(valor_linha(row, "moeda", "USD") or "USD").strip().upper()
            if not moeda:
                moeda = "USD"
            registros.append((chave, numero, valor_linha(row, "cnpj"), valor_linha(row, "cliente"),
                              moeda, valor_original, status_from_balance(valor_original), launch_timestamp()))
            vistos_numeros.add(numero)
            vistos_chaves.add(chave)
        except (TypeError, ValueError, InvalidOperation) as exc:
            rejeitados.append({"linha": linha, "chave": chave or "-", "numero": numero or "-", "motivo": str(exc)})

    conn = None
    try:
        if registros:
            conn = db()
            for tentativa in range(6):
                try:
                    conn.execute("BEGIN IMMEDIATE")
                    conn.executemany("""INSERT INTO dues
                        (chave_acesso,numero_due,cnpj,cliente,moeda,valor_original,status,created_at)
                        VALUES (?,?,?,?,?,?,?,?)""", registros)
                    conn.commit()
                    break
                except sqlite3.OperationalError as exc:
                    conn.rollback()
                    if "locked" not in str(exc).lower() or tentativa == 5:
                        raise
                    time.sleep(0.5)
    except Exception as exc:
        if conn is not None:
            conn.rollback()
        rejeitados.append({"linha": "-", "chave": "-", "numero": "-", "motivo": f"Falha ao gravar os registros válidos: {exc}"})
        registros = []
    finally:
        if conn is not None:
            conn.close()

    resumo = {"encontrados": len(df), "importados": len(registros), "rejeitados": len(rejeitados),
              "erros": rejeitados}
    return render_template("due_import_resultado.html", resumo=resumo)

@app.route("/due/nova", methods=["GET","POST"])
def nova_due():
    if request.method == "POST":
        f=request.form
        conn = None
        try:
            chave = parse_chave_acesso(f.get("chave_acesso"))
            valor_original = parse_number(f.get("valor_original"))
            ensure_non_negative_balance(valor_original, "O valor original da DU-E não pode ser negativo.")
            conn = db()
            cnpj = cnpj_da_empresa(conn, f.get("empresa_id"))
            empresa_id = form_record_id(f.get("empresa_id"), empresa_id_por_cnpj(conn, cnpj))
            data_due = parse_date(f.get("data_due")) or date.today().isoformat()
            competencia_id = competencia_da_operacao(conn, f.get("competencia_id"), empresa_id, data_due)
            if conn.execute("SELECT 1 FROM dues WHERE chave_acesso=?", (chave,)).fetchone():
                raise ValueError("A Chave de Acesso já está cadastrada.")
            conn.execute("""INSERT INTO dues
                (chave_acesso,numero_due,cnpj,cliente,moeda,valor_original,status,created_at,observacao,data_due,competencia_id)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (chave,f["numero_due"].strip(),cnpj,f.get("cliente"),
                 f.get("moeda") or "USD",valor_original, status_from_balance(valor_original),
                 launch_timestamp(),
                 f.get("observacao"), data_due, competencia_id))
            conn.commit()
            flash("DU-E cadastrada com sucesso.", "success")
            return redirect(url_for("index"))
        except sqlite3.IntegrityError:
            flash("Já existe uma DU-E com esse número ou Chave de Acesso.", "danger")
        except ValueError as exc:
            flash(str(exc), "danger")
        finally:
            if conn is not None:
                conn.close()
    conn = db()
    empresas, empresa_id = empresas_for_form(conn, selected_id=request.form.get("empresa_id"))
    competencias = competencias_for_empresa(conn, empresa_id)
    competencia_id = form_record_id(request.form.get("competencia_id"))
    if not competencia_id and empresa_id:
        sugerida = sugerir_competencia(conn, empresa_id, request.form.get("data_due") or date.today().isoformat())
        competencia_id = sugerida["id"] if sugerida else None
    conn.close()
    return render_template("due_form.html", due=None, empresas=empresas, empresa_id=empresa_id, competencias=competencias, competencia_id=competencia_id)

@app.route("/due/<int:due_id>/editar", methods=["GET", "POST"])
def editar_due(due_id):
    conn = db()
    due = conn.execute("SELECT * FROM dues WHERE id=?", (due_id,)).fetchone()
    if not due:
        conn.close(); return "DU-E não encontrada", 404
    if request.method == "POST":
        f = request.form
        try:
            chave = parse_chave_acesso(f.get("chave_acesso"))
            valor_original = parse_number(f.get("valor_original"))
            ensure_non_negative_balance(valor_original, "O valor original da DU-E não pode ser negativo.")
            utilizado_atual = due_effect(conn, due_id)
            ensure_non_negative_balance(
                decimal_value(valor_original) - decimal_value(utilizado_atual),
                "O valor original da DU-E não pode ficar abaixo do total utilizado."
            )
            cnpj = cnpj_da_empresa(conn, f.get("empresa_id"), due["cnpj"])
            empresa_id = form_record_id(f.get("empresa_id"), empresa_id_por_cnpj(conn, cnpj))
            data_due = parse_date(f.get("data_due")) or due["data_due"] or date.today().isoformat()
            competencia_id = competencia_da_operacao(conn, f.get("competencia_id"), empresa_id, data_due, due["competencia_id"] if "competencia_id" in due.keys() else None)
            if conn.execute("SELECT 1 FROM dues WHERE chave_acesso=? AND id<>?", (chave, due_id)).fetchone():
                raise ValueError("A Chave de Acesso já está cadastrada.")
            conn.execute("""UPDATE dues SET chave_acesso=?,numero_due=?,cnpj=?,cliente=?,moeda=?,valor_original=?,observacao=?,data_due=?,competencia_id=?
                            WHERE id=?""",
                         (chave, f["numero_due"].strip(), cnpj,
                          f.get("cliente"), f.get("moeda") or "USD", valor_original,
                          f.get("observacao"), data_due, competencia_id, due_id))
            update_due_status(conn, due_id)
            conn.commit(); conn.close()
            flash("DU-E atualizada com sucesso.", "success")
            return redirect(url_for("due_detalhe", due_id=due_id))
        except sqlite3.IntegrityError:
            flash("Já existe uma DU-E com esse número ou Chave de Acesso.", "danger")
        except ValueError as exc:
            flash(str(exc), "danger")
    empresas, empresa_id = empresas_for_form(
        conn, due["cnpj"], request.form.get("empresa_id") if request.method == "POST" else None
    )
    competencias = competencias_for_empresa(conn, empresa_id)
    competencia_id = form_record_id(request.form.get("competencia_id"), due["competencia_id"] if "competencia_id" in due.keys() else None)
    conn.close()
    return render_template("due_form.html", due=due, empresas=empresas, empresa_id=empresa_id, competencias=competencias, competencia_id=competencia_id)

def carregar_detalhe_due(conn, due_id):
    due_row=conn.execute("SELECT * FROM dues WHERE id=?", (due_id,)).fetchone()
    if not due_row:
        return None
    mov=conn.execute("""SELECT m.*,c.numero_contrato,c.moeda AS contrato_moeda
                       FROM due_movimentacoes m
                       LEFT JOIN contratos c ON c.id=m.contrato_id
                       WHERE m.due_id=? ORDER BY m.data_movimentacao DESC,m.id DESC""",(due_id,)).fetchall()
    vinc=conn.execute("""SELECT v.*,c.numero_contrato,c.moeda,c.valor_moeda,
                         m.id AS movimentacao_id,COALESCE(m.valor,v.valor_vinculado) AS valor_calculado
                         FROM due_contratos v JOIN contratos c ON c.id=v.contrato_id
                         LEFT JOIN due_movimentacoes m ON m.due_contrato_id=v.id AND m.tipo='VINCULACAO'
                         WHERE v.due_id=? ORDER BY v.id DESC""",(due_id,)).fetchall()
    contratos=[decorate_contract(row) for row in conn.execute("""SELECT c.*,COALESCE(SUM(CASE WHEN m.tipo='VINCULACAO' THEN m.valor ELSE 0 END),0) AS vinculado
                               FROM contratos c LEFT JOIN due_movimentacoes m ON m.contrato_id=c.id
                               GROUP BY c.id
                               HAVING c.saldo_zerado_manual=0
                                  AND c.valor_moeda-COALESCE(SUM(CASE WHEN m.tipo='VINCULACAO' THEN m.valor ELSE 0 END),0)>?
                               ORDER BY c.numero_contrato""", (float(SALDO_TOLERANCE),)).fetchall()]
    utilizado=due_effect(conn, due_id)
    due = decorate_due({**dict(due_row), "utilizado": utilizado})
    due["invoice_links"] = conn.execute("""
        SELECT di.*, i.numero_invoice, i.tipo_documento, i.moeda AS invoice_moeda
        FROM due_invoice di JOIN invoices i ON i.id=di.invoice_id
        WHERE di.due_id=? ORDER BY di.id DESC
    """, (due_id,)).fetchall()
    saldo=due["saldo"]
    return due, mov, vinc, contratos, utilizado, saldo

@app.route("/due/<int:due_id>")
def due_detalhe(due_id):
    conn=db()
    dados=carregar_detalhe_due(conn, due_id)
    conn.close()
    if not dados:
        return "DU-E não encontrada",404
    due, mov, vinc, contratos, utilizado, saldo = dados
    return render_template("due_detalhe.html", due=due, mov=mov, vinc=vinc, contratos=contratos,
                           utilizado=utilizado, saldo=saldo)

@app.route("/due/<int:due_id>/relatorio")
def relatorio_due(due_id):
    conn=db()
    dados=carregar_detalhe_due(conn, due_id)
    conn.close()
    if not dados:
        return "DU-E não encontrada",404
    due, mov, vinc, contratos, utilizado, saldo = dados
    historico=[]
    saldo_parcial=decimal_value(due.get("valor_original"))
    for item in reversed(mov):
        valor=decimal_value(item["valor"])
        if item["tipo"] in ("UTILIZACAO", "VINCULACAO"):
            saldo_parcial-=valor
        elif item["tipo"] == "DEVOLUCAO":
            saldo_parcial+=valor
        historico.append({"mov": item, "saldo": saldo_parcial})
    return render_template("due_relatorio.html", due=due, vinc=vinc, historico=historico,
                           utilizado=utilizado, saldo=saldo)

@app.route("/due/<int:due_id>/movimentacao", methods=["POST"])
def movimentacao(due_id):
    f=request.form
    conn = None
    try:
        conn=db()
        due = conn.execute("SELECT id,valor_original FROM dues WHERE id=?", (due_id,)).fetchone()
        if not due:
            raise ValueError("DU-E não encontrada.")
        tipo = f.get("tipo", "UTILIZACAO").upper()
        if tipo not in {"UTILIZACAO", "DEVOLUCAO"}:
            raise ValueError("Tipo de movimentação inválido.")
        valor = parse_number(f.get("valor"))
        if valor <= 0:
            raise ValueError("O valor da movimentação deve ser maior que zero.")
        utilizado_atual = decimal_value(due_effect(conn, due_id))
        if tipo == "UTILIZACAO":
            ensure_non_negative_balance(
                decimal_value(due["valor_original"]) - utilizado_atual - decimal_value(valor),
                "A utilização não pode ultrapassar o saldo disponível da DU-E."
            )
        else:
            ensure_non_negative_balance(
                utilizado_atual - decimal_value(valor),
                "A devolução não pode ultrapassar o total utilizado da DU-E."
            )
        conn.execute("""INSERT INTO due_movimentacoes
            (due_id,data_movimentacao,tipo,documento,valor,observacao)
            VALUES (?,?,?,?,?,?)""",
            (due_id,parse_date(f["data_movimentacao"]),tipo,f.get("documento"),valor,f.get("observacao")))
        update_due_status(conn, due_id)
        conn.commit(); conn.close()
        flash("Movimentação registrada com sucesso.", "success")
    except ValueError as exc:
        if conn is not None:
            conn.close()
        flash(str(exc), "danger")
    return redirect(url_for("due_detalhe", due_id=due_id))

@app.route("/due/<int:due_id>/vincular", methods=["POST"])
def vincular(due_id):
    f=request.form
    conn=None
    conn=db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        due=conn.execute("SELECT valor_original FROM dues WHERE id=?",(due_id,)).fetchone()
        if not due:
            raise ValueError("DU-E não encontrada.")
        contrato_id=int(f["contrato_id"])
        contrato=contract_summary(conn, contrato_id)
        if not contrato:
            raise ValueError("Contrato Câmbio não encontrado.")
        if contrato["saldo_zerado_manual"]:
            raise ValueError("O Contrato Câmbio foi zerado manualmente e não pode receber vínculos.")
        if contrato["status"] not in {STATUS_PENDENTE, STATUS_PARCIAL}:
            raise ValueError("O Contrato Câmbio selecionado não possui saldo disponível.")
        valor=Decimal(str(parse_number(f.get("valor_vinculado"))))
        if valor<=0:
            raise ValueError("O valor do vínculo deve ser maior que zero.")
        saldo_due=due_balance(due["valor_original"], due_effect(conn, due_id))
        saldo_contrato=contract_balance(contrato["valor_moeda"], contrato["vinculado"])
        if saldo_due <= 0:
            raise ValueError("A DU-E não possui saldo disponível.")
        if saldo_contrato <= 0:
            raise ValueError("O Contrato Câmbio não possui saldo disponível.")
        if valor>saldo_due:
            raise ValueError(f"Valor maior que o saldo disponível da DU-E ({money(saldo_due)}).")
        if valor>saldo_contrato:
            raise ValueError(f"Valor maior que o saldo disponível do Contrato Câmbio ({money(saldo_contrato)}).")
        link=conn.execute("SELECT id,valor_vinculado FROM due_contratos WHERE due_id=? AND contrato_id=?",
                          (due_id,contrato_id)).fetchone()
        if link:
            due_contrato_id=link["id"]
            conn.execute("UPDATE due_contratos SET valor_vinculado=valor_vinculado+?,observacao=? WHERE id=?",
                         (float(valor),f.get("observacao"),due_contrato_id))
        else:
            conn.execute("""INSERT INTO due_contratos(due_id,contrato_id,valor_vinculado,observacao)
                            VALUES (?,?,?,?)""",(due_id,contrato_id,float(valor),f.get("observacao")))
            due_contrato_id=conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.execute("""INSERT INTO due_movimentacoes
            (due_id,contrato_id,due_contrato_id,data_movimentacao,tipo,documento,valor,observacao)
            VALUES (?,?,?,?,?,?,?,?)""",(due_id,contrato_id,due_contrato_id,date.today().isoformat(),
                                          "VINCULACAO",f"CONTRATO:{contrato['numero_contrato']}",float(valor),f.get("observacao")))
        recalculate_statuses(conn, due_ids=[due_id], contrato_ids=[contrato_id])
        conn.commit(); flash("Contrato Câmbio vinculado e movimentação registrada com sucesso.", "success")
    except sqlite3.IntegrityError:
        if conn: conn.rollback()
        flash("Esse Contrato Câmbio já está vinculado a esta DU-E.", "danger")
    except ValueError as exc:
        if conn: conn.rollback()
        flash(str(exc), "danger")
    finally:
        if conn: conn.close()
    return redirect(url_for("due_detalhe", due_id=due_id))

@app.route("/due/<int:due_id>/movimentacao/<int:mov_id>/excluir", methods=["POST"])
def excluir_movimentacao(due_id, mov_id):
    conn=db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        mov=conn.execute("SELECT * FROM due_movimentacoes WHERE id=? AND due_id=?",(mov_id,due_id)).fetchone()
        if not mov:
            raise ValueError("Movimentação não encontrada.")
        contrato_id = None
        if mov["tipo"]=="VINCULACAO":
            contrato_id = mov["contrato_id"]
            if mov["due_contrato_id"]:
                link=conn.execute("SELECT contrato_id,valor_vinculado FROM due_contratos WHERE id=?",(mov["due_contrato_id"],)).fetchone()
                if link:
                    contrato_id = contrato_id or link["contrato_id"]
                    saldo_link=decimal_value(link["valor_vinculado"])-decimal_value(mov["valor"])
                    if saldo_link <= SALDO_TOLERANCE:
                        conn.execute("DELETE FROM due_contratos WHERE id=?",(mov["due_contrato_id"],))
                    else:
                        conn.execute("UPDATE due_contratos SET valor_vinculado=? WHERE id=?",
                                     (float(saldo_link),mov["due_contrato_id"]))
            elif mov["contrato_id"]:
                link=conn.execute("SELECT id,valor_vinculado FROM due_contratos WHERE due_id=? AND contrato_id=?",
                                  (due_id,mov["contrato_id"])).fetchone()
                if link:
                    saldo_link=decimal_value(link["valor_vinculado"])-decimal_value(mov["valor"])
                    if saldo_link <= SALDO_TOLERANCE:
                        conn.execute("DELETE FROM due_contratos WHERE id=?",(link["id"],))
                    else:
                        conn.execute("UPDATE due_contratos SET valor_vinculado=? WHERE id=?",(float(saldo_link),link["id"]))
        conn.execute("DELETE FROM due_movimentacoes WHERE id=?",(mov_id,))
        recalculate_statuses(conn, due_ids=[due_id], contrato_ids=[contrato_id] if mov["tipo"] == "VINCULACAO" and contrato_id else [])
        conn.commit(); flash("Movimentação excluída e saldo revertido com sucesso.", "success")
    except ValueError as exc:
        conn.rollback(); flash(str(exc), "danger")
    finally:
        conn.close()
    return redirect(url_for("due_detalhe", due_id=due_id))

def contract_import_stage_path(token):
    if not isinstance(token, str) or not re.fullmatch(r"[A-Za-z0-9_-]{20,128}", token):
        raise ValueError("A prévia da importação é inválida ou expirou.")
    return Path(tempfile.gettempdir()) / f"{CONTRACT_IMPORT_STAGE_PREFIX}{token}.json"

def cleanup_contract_import_stages():
    cutoff = time.time() - CONTRACT_IMPORT_STAGE_TTL
    directory = Path(tempfile.gettempdir())
    for path in directory.glob(f"{CONTRACT_IMPORT_STAGE_PREFIX}*.json"):
        try:
            if path.stat().st_mtime < cutoff:
                path.unlink()
        except OSError:
            pass

def remove_contract_import_stage(token=None):
    stage_token = token if token is not None else session.get("contract_import_stage")
    if stage_token:
        try:
            contract_import_stage_path(stage_token).unlink(missing_ok=True)
        except (OSError, ValueError):
            pass
    if session.get("contract_import_stage") == stage_token:
        session.pop("contract_import_stage", None)

def save_contract_import_stage(payload):
    cleanup_contract_import_stages()
    old_token = session.get("contract_import_stage")
    if old_token:
        remove_contract_import_stage(old_token)
    token = secrets.token_urlsafe(24)
    payload = dict(payload)
    payload["created_at"] = time.time()
    path = contract_import_stage_path(token)
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    session["contract_import_stage"] = token
    return token

def load_contract_import_stage(token=None):
    token = (token or session.get("contract_import_stage") or "").strip()
    if not token or token != session.get("contract_import_stage"):
        raise ValueError("A prévia da importação não pertence a esta sessão.")
    path = contract_import_stage_path(token)
    try:
        if time.time() - path.stat().st_mtime > CONTRACT_IMPORT_STAGE_TTL:
            raise FileNotFoundError
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, OSError, ValueError, json.JSONDecodeError):
        remove_contract_import_stage(token)
        raise ValueError("A prévia da importação é inválida ou expirou.")
    allowed_date_columns = {"data_recebimento", "data_liquidacao"}
    if (
        not isinstance(payload, dict)
        or payload.get("version") != 1
        or not isinstance(payload.get("rows"), list)
        or not isinstance(payload.get("date_columns", []), list)
        or not set(payload.get("date_columns", [])).issubset(allowed_date_columns)
    ):
        remove_contract_import_stage(token)
        raise ValueError("A prévia da importação é inválida ou expirou.")
    return payload

def contract_import_cell(record, columns, column, pandas, default=None):
    if column not in columns:
        return default
    value = record[column]
    return None if pandas.isna(value) else value

def contract_import_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None

def prepare_contract_import_rows(df, pandas):
    date_columns = {
        column for column in ("data_recebimento", "data_liquidacao")
        if column in df.columns
    }
    rows = []
    for line_number, (_, record) in enumerate(df.iterrows(), start=2):
        raw_number = contract_import_cell(record, df.columns, "numero_contrato", pandas, "")
        numero = contract_import_text(raw_number)
        if not numero:
            values = [contract_import_cell(record, df.columns, column, pandas) for column in df.columns]
            if any(contract_import_text(value) for value in values):
                raise ValueError(f"Linha {line_number}: o numero_contrato do Contrato Câmbio é obrigatório.")
            continue
        try:
            valor_moeda = parse_number(contract_import_cell(record, df.columns, "valor_moeda", pandas, 0))
            ensure_non_negative_balance(valor_moeda, "valor_moeda não pode ser negativo")
            data_contrato = parse_date(contract_import_cell(record, df.columns, "data_contrato", pandas))
            data_recebimento = parse_date(contract_import_cell(record, df.columns, "data_recebimento", pandas))
            data_liquidacao = parse_date(contract_import_cell(record, df.columns, "data_liquidacao", pandas))
            taxa_cambio = optional_number(contract_import_cell(record, df.columns, "taxa_cambio", pandas))
            valor_reais = optional_number(contract_import_cell(record, df.columns, "valor_reais", pandas))
        except ValueError as exc:
            raise ValueError(f"Linha {line_number}: {exc}") from exc
        banco_credito = contract_import_cell(record, df.columns, "banco_credito", pandas)
        banco_credito = banco_credito or contract_import_cell(record, df.columns, "banco", pandas)
        banco_credito = contract_import_text(banco_credito)
        banco_liquidacao = contract_import_cell(record, df.columns, "banco_liquidacao", pandas)
        banco_liquidacao = contract_import_text(banco_liquidacao) or banco_credito
        rows.append({
            "row_id": f"r{line_number}",
            "source_row": line_number,
            "numero_contrato": numero,
            "banco_credito": banco_credito,
            "banco_liquidacao": banco_liquidacao,
            "data_contrato": data_contrato,
            "data_recebimento": data_recebimento,
            "data_liquidacao": data_liquidacao,
            "cnpj": contract_import_text(contract_import_cell(record, df.columns, "cnpj", pandas)),
            "cliente": contract_import_text(contract_import_cell(record, df.columns, "cliente", pandas)),
            "moeda": (contract_import_text(contract_import_cell(record, df.columns, "moeda", pandas, "USD")) or "USD").upper(),
            "valor_moeda": float(valor_moeda),
            "taxa_cambio": taxa_cambio,
            "valor_reais": valor_reais,
        })
    if not rows:
        raise ValueError("A planilha não contém Contratos Câmbio válidos para importar.")
    return rows, date_columns

def contracts_for_import(conn, numbers):
    if not numbers:
        return {}
    numbers = [str(number).strip() for number in numbers]
    placeholders = ",".join("?" for _ in numbers)
    rows = conn.execute(f"""
        SELECT c.*, COALESCE(SUM(CASE WHEN m.tipo='VINCULACAO' THEN m.valor ELSE 0 END), 0) AS vinculado
        FROM contratos c
        LEFT JOIN due_movimentacoes m ON m.contrato_id=c.id
        WHERE TRIM(c.numero_contrato) IN ({placeholders})
        GROUP BY c.id
    """, numbers).fetchall()
    result = {}
    for row in rows:
        numero = str(row["numero_contrato"] or "").strip()
        if numero in result and result[numero]["id"] != row["id"]:
                raise ValueError(f"O banco contém mais de um Contrato Câmbio com o número {numero} após a normalização.")
        result[numero] = dict(row)
    return result

def contract_import_snapshot(row):
    if not row:
        return None
    data = dict(row)
    snapshot = {
        key: data.get(key) for key in (
            "id", "numero_contrato", "banco", "banco_credito", "banco_liquidacao",
            "data_contrato", "data_recebimento", "data_liquidacao", "cnpj", "cliente",
            "moeda", "valor_moeda", "taxa_cambio", "valor_reais", "status",
            "saldo_zerado_manual", "observacao", "competencia_id", "vinculado",
        )
    }
    if snapshot["vinculado"] is not None:
        snapshot["vinculado"] = float(snapshot["vinculado"])
    return snapshot

def build_contract_import_payload(conn, rows, date_columns):
    grouped = {}
    for row in rows:
        grouped.setdefault(row["numero_contrato"], []).append(row)
    existing = contracts_for_import(conn, list(grouped))
    existing_snapshots = {}
    for numero, group_rows in grouped.items():
        current = contract_import_snapshot(existing.get(numero))
        existing_snapshots[numero] = current
        if current:
            for row in group_rows:
                try:
                    ensure_non_negative_balance(
                        decimal_value(row["valor_moeda"]) - decimal_value(current.get("vinculado")),
                        "valor_moeda não pode ficar abaixo do total já vinculado"
                    )
                except ValueError as exc:
                    row["update_error"] = str(exc)
    return {
        "version": 1,
        "date_columns": sorted(date_columns),
        "rows": rows,
        "existing_by_number": existing_snapshots,
    }

def contract_import_preview_context(payload):
    grouped = {}
    for row in payload["rows"]:
        grouped.setdefault(row["numero_contrato"], []).append(row)
    conflicts = []
    new_rows = []
    for group_index, (numero, rows) in enumerate(grouped.items()):
        existing = payload.get("existing_by_number", {}).get(numero)
        if existing or len(rows) > 1:
            conflicts.append({
                "group_id": f"g{group_index}",
                "numero_contrato": numero,
                "rows": rows,
                "existing": existing,
                "kind": "existing" if existing else "planilha",
                "requires_choice": len(rows) > 1,
                "can_update": bool(existing) and any(not row.get("update_error") for row in rows),
            })
        else:
            new_rows.append(rows[0])
    return {
        "new_rows": new_rows,
        "conflicts": conflicts,
        "total_rows": len(payload["rows"]),
        "duplicate_rows": sum(len(group["rows"]) for group in conflicts),
        "date_columns": payload.get("date_columns", []),
    }

def render_contract_import_preview(payload, error=None):
    context = contract_import_preview_context(payload)
    context.update(stage_token=session.get("contract_import_stage"), error=error)
    return render_template("contrato_import_preview.html", **context)

def contract_import_decisions(payload, form):
    context = contract_import_preview_context(payload)
    rows_to_apply = list(context["new_rows"])
    discarded = 0
    for conflict in context["conflicts"]:
        rows = conflict["rows"]
        selected = rows[0]
        if conflict["requires_choice"]:
            selected_id = (form.get(f"chosen_row_{conflict['group_id']}") or "").strip()
            selected = next((row for row in rows if row["row_id"] == selected_id), None)
            if selected is None:
                raise ValueError(f"Escolha uma linha para o Contrato Câmbio {conflict['numero_contrato']}.")
        if conflict["existing"]:
            action = form.get(f"duplicate_action_{conflict['group_id']}")
            if action not in {"update", "discard"}:
                raise ValueError(f"Escolha uma ação para o Contrato Câmbio {conflict['numero_contrato']}.")
            if action == "update":
                if selected.get("update_error"):
                    raise ValueError(
                        f"Contrato Câmbio {conflict['numero_contrato']}: {selected['update_error']}."
                    )
                rows_to_apply.append(selected)
            discarded += len(rows) if action == "discard" else len(rows) - 1
        else:
            action = form.get(f"file_duplicate_action_{conflict['group_id']}")
            if action not in {"insert", "discard"}:
                raise ValueError(f"Escolha uma ação para o grupo do Contrato Câmbio {conflict['numero_contrato']}.")
            if action == "insert":
                rows_to_apply.append(selected)
                discarded += len(rows) - 1
            else:
                discarded += len(rows)
    return rows_to_apply, discarded

def apply_contract_import_rows(conn, rows, date_columns):
    allowed_date_columns = {"data_recebimento", "data_liquidacao"}
    date_columns = set(date_columns)
    if not date_columns.issubset(allowed_date_columns):
        raise ValueError("A prévia contém colunas de data inválidas.")
    update_columns = [
        "numero_contrato=?", "banco=?", "banco_credito=?", "banco_liquidacao=?",
        "data_contrato=?", "cnpj=?", "cliente=?", "moeda=?", "valor_moeda=?",
        "taxa_cambio=?", "valor_reais=?", "status=?",
    ]
    for column in sorted(date_columns):
        update_columns.insert(5, f"{column}=?")
    update_query = "UPDATE contratos SET " + ",".join(update_columns) + " WHERE id=?"
    insert_query = """INSERT INTO contratos
        (numero_contrato,banco,banco_credito,banco_liquidacao,data_contrato,data_recebimento,data_liquidacao,cnpj,cliente,moeda,valor_moeda,taxa_cambio,valor_reais,status)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """
    inserted = 0
    updated = 0
    contrato_ids = []
    for row in rows:
        existing = conn.execute(
            "SELECT id FROM contratos WHERE TRIM(numero_contrato)=?", (row["numero_contrato"],)
        ).fetchone()
        if existing:
            linked = contract_summary(conn, existing["id"])["vinculado"]
            ensure_non_negative_balance(
                decimal_value(row["valor_moeda"]) - decimal_value(linked),
                "valor_moeda não pode ficar abaixo do total já vinculado"
            )
            updated += 1
        else:
            inserted += 1
        values = [
            row["numero_contrato"], row["banco_credito"], row["banco_credito"],
            row["banco_liquidacao"], row["data_contrato"], row["cnpj"], row["cliente"],
            row["moeda"], row["valor_moeda"], row["taxa_cambio"], row["valor_reais"],
            status_from_balance(row["valor_moeda"]),
        ]
        for column in sorted(date_columns):
            values.insert(5, row[column])
        if existing:
            conn.execute(update_query, values + [existing["id"]])
        else:
            conn.execute(insert_query, (
                row["numero_contrato"], row["banco_credito"], row["banco_credito"],
                row["banco_liquidacao"], row["data_contrato"], row["data_recebimento"],
                row["data_liquidacao"], row["cnpj"], row["cliente"], row["moeda"],
                row["valor_moeda"], row["taxa_cambio"], row["valor_reais"],
                status_from_balance(row["valor_moeda"]),
            ))
        contrato_ids.append(conn.execute(
            "SELECT id FROM contratos WHERE TRIM(numero_contrato)=?", (row["numero_contrato"],)
        ).fetchone()[0])
    recalculate_statuses(conn, due_ids=[], contrato_ids=contrato_ids)
    return {"inserted": inserted, "updated": updated, "discarded": 0}

def revalidate_contract_import_stage(conn, payload):
    numbers = list(payload.get("existing_by_number", {}))
    current = contracts_for_import(conn, numbers)
    expected = payload.get("existing_by_number", {})
    for numero in numbers:
        if contract_import_snapshot(current.get(numero)) != expected.get(numero):
            raise ValueError(
                f"O contrato {numero} foi alterado desde a prévia. Analise a planilha novamente."
            )

@app.route("/contratos/importar", methods=["POST"])
def importar_contratos():
    return importar_invoices()
    conn = None
    try:
        import pandas as pd
        arquivo = request.files.get("arquivo")
        if not arquivo or not arquivo.filename:
            flash("Selecione um arquivo Excel.", "danger")
            return redirect(url_for("index"))
        df = pd.read_excel(arquivo)
        df.columns = normalize_contract_import_columns(df.columns)
        if "numero_contrato" not in df.columns:
            flash("O Excel precisa conter a coluna numero_contrato do Contrato Câmbio.", "danger")
            return redirect(url_for("index"))
        conn = db()
        rows, date_columns = prepare_contract_import_rows(df, pd)
        payload = build_contract_import_payload(conn, rows, date_columns)
        save_contract_import_stage(payload)
        return render_contract_import_preview(payload)
    except ValueError as exc:
        flash(str(exc), "danger")
    except Exception as exc:
        flash(f"Erro na análise da importação: {exc}", "danger")
    finally:
        if conn is not None:
            conn.close()
    return redirect(url_for("index"))

@app.route("/contratos/importar/confirmar", methods=["POST"])
def confirmar_importacao_contratos():
    return confirmar_importacao_invoices()
    token = request.form.get("stage_token")
    payload = None
    conn = None
    preview_error = None
    try:
        payload = load_contract_import_stage(token)
        conn = db()
        conn.execute("BEGIN IMMEDIATE")
        revalidate_contract_import_stage(conn, payload)
        rows, discarded = contract_import_decisions(payload, request.form)
        result = apply_contract_import_rows(conn, rows, set(payload.get("date_columns", [])))
        result["discarded"] = discarded
        conn.commit()
        remove_contract_import_stage(token)
        flash(
            f"Importação concluída: {result['inserted']} inserido(s), "
            f"{result['updated']} atualizado(s) e {result['discarded']} descartado(s).",
            "success",
        )
        return redirect(url_for("index"))
    except ValueError as exc:
        if conn is not None:
            conn.rollback()
        if payload is not None and session.get("contract_import_stage") == token:
            preview_error = str(exc)
        else:
            flash(str(exc), "danger")
    except Exception as exc:
        if conn is not None:
            conn.rollback()
        if payload is not None and session.get("contract_import_stage") == token:
            preview_error = f"Não foi possível concluir a importação: {exc}"
        else:
            flash(f"Erro na confirmação da importação: {exc}", "danger")
    finally:
        if conn is not None:
            conn.close()
    if preview_error:
        return render_contract_import_preview(payload, error=preview_error), 400
    return redirect(url_for("index"))

@app.route("/contratos/importar/cancelar", methods=["POST"])
def cancelar_importacao_contratos():
    return cancelar_importacao_invoices()
    token = request.form.get("stage_token")
    if token and token == session.get("contract_import_stage"):
        remove_contract_import_stage(token)
        flash("Importação cancelada.", "success")
    else:
        flash("A prévia da importação já expirou ou não é válida.", "danger")
    return redirect(url_for("index"))

@app.route("/contratos/modelo")
def modelo_contratos():
    return modelo_invoices()
    import pandas as pd
    df=pd.DataFrame(columns=["numero_contrato","banco_credito","banco_liquidacao","data_contrato","data_recebimento","data_liquidacao","cnpj","cliente","moeda","valor_moeda","taxa_cambio","valor_reais"])
    out=io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer,index=False,sheet_name="Contratos")
    out.seek(0)
    return send_file(out,as_attachment=True,download_name="modelo_contratos.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def invoice_type_label(value):
    return {
        "PROFORMA": "Proforma",
        "COMMERCIAL_INVOICE": "Commercial Invoice",
        "SERVICE_INVOICE": "Service Invoice",
        "DEBIT_NOTE": "Debit Note",
    }.get(str(value or "").upper(), value or "-")

app.jinja_env.filters["invoice_type_label"] = invoice_type_label

def invoice_status_class(status):
    return {
        INVOICE_STATUS_AGUARDANDO_RECEBIMENTO: "status-invoice-awaiting",
        INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO: "status-invoice-exchange",
        INVOICE_STATUS_LIQUIDADA: "status-invoice-settled",
    }.get(status, "status-pendente")

app.jinja_env.filters["invoice_status_class"] = invoice_status_class
app.jinja_env.filters["invoice_status_label"] = lambda value: INVOICE_STATUS_LABELS.get(value, value or "-")

def invoice_status_from_totals(valor_invoice, total_recebido, total_cambio):
    valor_invoice = decimal_value(valor_invoice)
    total_recebido = decimal_value(total_recebido)
    total_cambio = decimal_value(total_cambio)
    saldo_recebimento = normalize_balance(valor_invoice - total_recebido)
    saldo_cambio = normalize_balance(total_recebido - total_cambio)
    if saldo_recebimento > SALDO_TOLERANCE:
        return INVOICE_STATUS_AGUARDANDO_RECEBIMENTO
    if saldo_cambio > SALDO_TOLERANCE:
        return INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO
    return INVOICE_STATUS_LIQUIDADA

def invoice_summary(conn, invoice_id):
    invoice = conn.execute("""
        SELECT i.*, e.razao_social AS empresa_razao_social, e.apelido AS empresa_apelido,
               e.cnpj AS empresa_cnpj, cl.nome AS cliente_nome, cl.pais AS cliente_pais,
               comp.descricao AS competencia_descricao, comp.data_inicial AS competencia_data_inicial,
               comp.data_final AS competencia_data_final
        FROM invoices i
        JOIN empresas e ON e.id=i.empresa_id
        LEFT JOIN clientes cl ON cl.id=i.cliente_id
        LEFT JOIN competencias comp ON comp.id=i.competencia_id
        WHERE i.id=?
    """, (invoice_id,)).fetchone()
    if not invoice:
        return None
    recebimentos = conn.execute("""
        SELECT r.*, cp.nome AS banco_credito_nome
        FROM recebimentos_invoice r
        LEFT JOIN contrapartes cp ON cp.id=r.banco_credito_id
        WHERE r.invoice_id=? ORDER BY r.data_credito DESC, r.id DESC
    """, (invoice_id,)).fetchall()
    cambios = conn.execute("""
        SELECT v.*, c.numero_contrato, c.moeda AS contrato_moeda,
               c.banco_liquidacao, c.data_fechamento, c.data_liquidacao,
               c.taxa_cambio, c.valor_reais AS contrato_valor_reais
        FROM invoice_contrato_cambio v
        JOIN contratos c ON c.id=v.contrato_id
        WHERE v.invoice_id=? ORDER BY v.id DESC
    """, (invoice_id,)).fetchall()
    due_links = conn.execute("""
        SELECT di.*, d.numero_due, d.chave_acesso, d.moeda AS due_moeda
        FROM due_invoice di JOIN dues d ON d.id=di.due_id
        WHERE di.invoice_id=? ORDER BY di.id DESC
    """, (invoice_id,)).fetchall()
    total_recebido = sum((decimal_value(row["valor_moeda"]) for row in recebimentos), Decimal("0"))
    total_cambio = sum((decimal_value(row["valor_alocado"]) for row in cambios), Decimal("0"))
    valor_invoice = decimal_value(invoice["valor_moeda"])
    saldo_recebimento = normalize_balance(valor_invoice - total_recebido)
    saldo_cambio = normalize_balance(total_recebido - total_cambio)
    taxa_volume = sum((decimal_value(row["valor_alocado"]) for row in cambios
                       if row["taxa_cambio"] is not None), Decimal("0"))
    taxa_valor = sum((decimal_value(row["valor_alocado"]) * decimal_value(row["taxa_cambio"])
                      for row in cambios if row["taxa_cambio"] is not None), Decimal("0"))
    valor_brl_calculado = sum((decimal_value(row["valor_alocado"]) * decimal_value(row["taxa_cambio"])
                               for row in cambios if row["taxa_cambio"] is not None), Decimal("0"))
    try:
        status = normalize_invoice_status(invoice["status"], default=INVOICE_STATUS_AGUARDANDO_RECEBIMENTO)
    except ValueError:
        status = INVOICE_STATUS_AGUARDANDO_RECEBIMENTO
    if not invoice["status_manual"]:
        status = invoice_status_from_totals(valor_invoice, total_recebido, total_cambio)
    datas_credito = {row["data_credito"] for row in recebimentos if row["data_credito"]}
    if invoice["data_credito"]:
        datas_credito.add(invoice["data_credito"])
    data = dict(invoice)
    data.update({
        "valor_moeda": valor_invoice,
        "total_recebido": total_recebido,
        "total_cambio": total_cambio,
        "saldo_recebimento": saldo_recebimento,
        "saldo_cambio": saldo_cambio,
        "taxa_cambio_media": taxa_valor / taxa_volume if taxa_volume else None,
        "valor_brl": valor_brl_calculado if any(row["taxa_cambio"] is not None for row in cambios) else None,
        "status": status,
        "recebimentos": recebimentos,
        "cambios": cambios,
        "due_links": due_links,
        "bancos_credito": sorted({row["banco_credito_nome"] for row in recebimentos if row["banco_credito_nome"]}),
        "bancos_liquidacao": sorted({row["banco_liquidacao"] for row in cambios if row["banco_liquidacao"]}),
        "contratos_numeros": sorted({row["numero_contrato"] for row in cambios}),
        "datas_credito": sorted(datas_credito),
        "datas_fechamento": sorted({row["data_fechamento"] for row in cambios if row["data_fechamento"]}),
        "datas_liquidacao": sorted({row["data_liquidacao"] for row in cambios if row["data_liquidacao"]}),
    })
    return data

def refresh_invoice_status(conn, invoice_id):
    summary = invoice_summary(conn, invoice_id)
    if not summary:
        return None
    current = conn.execute("SELECT status, status_manual, data_credito FROM invoices WHERE id=?", (invoice_id,)).fetchone()
    if current and not current["status_manual"]:
        data_credito = None
        if summary["status"] in {INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO, INVOICE_STATUS_LIQUIDADA}:
            data_credito = conn.execute(
                "SELECT MIN(data_credito) FROM recebimentos_invoice WHERE invoice_id=?", (invoice_id,)
            ).fetchone()[0]
        if current["status"] != summary["status"] or current["data_credito"] != data_credito:
            conn.execute("UPDATE invoices SET status=?, data_credito=? WHERE id=?",
                         (summary["status"], data_credito, invoice_id))
    return summary["status"]


def invoice_report_copy_text(title, rows, total):
    lines = [title, "", "Cliente / Trading        USD"]
    for row in rows:
        lines.append(f"{row['cliente']:<24} US$ {money(row['valor'])}")
    lines.extend([
        "--------------------------------",
        f"{'TOTAL':<24} US$ {money(total)}",
    ])
    return "\n".join(lines)


def build_invoice_report_context():
    conn = db()
    try:
        invoice_ids = [row["id"] for row in conn.execute("SELECT id FROM invoices").fetchall()]
        for invoice_id in invoice_ids:
            refresh_invoice_status(conn, invoice_id)
        conn.commit()
        summaries = [invoice_summary(conn, invoice_id) for invoice_id in invoice_ids]
    finally:
        conn.close()

    summaries = [
        summary for summary in summaries
        if summary and str(summary.get("moeda") or "").upper() == "USD"
    ]

    def grouped_rows(status, value_key):
        grouped = {}
        for summary in summaries:
            if summary["status"] != status:
                continue
            cliente = summary.get("cliente_nome") or "Não informado"
            grouped[cliente] = grouped.get(cliente, Decimal("0")) + decimal_value(summary[value_key])
        rows = [{"cliente": cliente, "valor": valor} for cliente, valor in grouped.items()]
        rows.sort(key=lambda row: (-row["valor"], row["cliente"].casefold()))
        total = sum((row["valor"] for row in rows), Decimal("0"))
        return rows, total

    recebido_aguardando_cambio, total_recebido_aguardando_cambio = grouped_rows(
        INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO, "saldo_cambio"
    )
    aguardando_recebimento, total_aguardando_recebimento = grouped_rows(
        INVOICE_STATUS_AGUARDANDO_RECEBIMENTO, "saldo_recebimento"
    )
    total_recebido = sum((decimal_value(summary["total_recebido"]) for summary in summaries), Decimal("0"))
    total_cambio = sum((decimal_value(summary["total_cambio"]) for summary in summaries), Decimal("0"))

    tables = [
        {"title": "RECEBIDO AGUARDANDO CÂMBIO", "rows": recebido_aguardando_cambio,
         "total": total_recebido_aguardando_cambio},
        {"title": "AGUARDANDO RECEBIMENTO", "rows": aguardando_recebimento,
         "total": total_aguardando_recebimento},
    ]
    for table in tables:
        table["copy_text"] = invoice_report_copy_text(table["title"], table["rows"], table["total"])

    return {
        "kpis": {
            "total_recebido": total_recebido,
            "total_cambio": total_cambio,
            "total_aguardando_recebimento": total_aguardando_recebimento,
        },
        "tables": tables,
    }

def validate_invoice_balances(summary, extra_recebido=0, extra_cambio=0,
                              replacement_recebido=None, replacement_cambio=None):
    valor = decimal_value(summary["valor_moeda"])
    recebido = decimal_value(summary["total_recebido"]) + decimal_value(extra_recebido)
    cambio = decimal_value(summary["total_cambio"]) + decimal_value(extra_cambio)
    if replacement_recebido is not None:
        recebido = decimal_value(replacement_recebido)
    if replacement_cambio is not None:
        cambio = decimal_value(replacement_cambio)
    ensure_non_negative_balance(valor - recebido,
                                "O recebimento não pode ultrapassar o valor da Invoice.")
    ensure_non_negative_balance(recebido - cambio,
                                "O câmbio não pode ultrapassar o total recebido da Invoice.")

def normalize_contract_commercial(value):
    text_value = str(value or "").strip()
    if any(unicodedata.category(character) == "Cc" for character in text_value):
        raise ValueError("O Contrato comercial não pode conter caracteres de controle.")
    if len(text_value) > 120:
        raise ValueError("O Contrato comercial deve ter no máximo 120 caracteres.")
    return text_value or None

def invoice_form_data(form, conn, current=None):
    try:
        empresa_id = int(form.get("empresa_id"))
    except (TypeError, ValueError):
        empresa_id = current["empresa_id"] if current else None
    if not empresa_id or not conn.execute("SELECT id FROM empresas WHERE id=?", (empresa_id,)).fetchone():
        raise ValueError("Selecione uma empresa cadastrada.")
    numero = (form.get("numero_invoice") or "").strip()
    if not numero:
        raise ValueError("O número da Invoice é obrigatório.")
    tipo = (form.get("tipo_documento") or "").strip().upper()
    if tipo not in INVOICE_TYPE_CODES:
        raise ValueError("Selecione um tipo de documento válido.")
    data_emissao = parse_date(form.get("data_emissao"))
    moeda = (form.get("moeda") or "USD").strip().upper()
    if not re.fullmatch(r"[A-Z]{3}", moeda):
        raise ValueError("Informe uma moeda válida com três letras.")
    valor = parse_number(form.get("valor_moeda"))
    if decimal_value(valor) <= 0:
        raise ValueError("O valor da Invoice deve ser maior que zero.")
    competencia_id = competencia_da_operacao(
        conn, form.get("competencia_id"), empresa_id,
        data_emissao or date.today().isoformat(),
        current["competencia_id"] if current and "competencia_id" in current.keys() else None,
    )
    cliente_id = form_record_id(form.get("cliente_id"), current["cliente_id"] if current else None)
    if cliente_id and not conn.execute("SELECT id FROM clientes WHERE id=?", (cliente_id,)).fetchone():
        raise ValueError("O cliente selecionado não foi encontrado.")
    contrato_comercial = normalize_contract_commercial(form.get("contrato_comercial"))
    if "status" in form:
        status = normalize_invoice_status(form.get("status"))
        status_manual = 1
    elif current:
        status = normalize_invoice_status(
            current["status"], default=INVOICE_STATUS_AGUARDANDO_RECEBIMENTO
        )
        status_manual = int(current["status_manual"]) if "status_manual" in current.keys() else 0
    else:
        status = INVOICE_STATUS_AGUARDANDO_RECEBIMENTO
        status_manual = 0
    if "data_credito" in form:
        data_credito = parse_date(form.get("data_credito"))
    elif current:
        data_credito = current["data_credito"] if "data_credito" in current.keys() else None
    else:
        data_credito = None
    if status == INVOICE_STATUS_AGUARDANDO_RECEBIMENTO:
        data_credito = None
    elif status == INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO and not data_credito:
        raise ValueError("Informe a data do crédito para alterar o status para recebido.")
    return {
        "empresa_id": empresa_id, "numero_invoice": numero, "tipo_documento": tipo,
        "competencia_id": competencia_id, "cliente_id": cliente_id, "data_emissao": data_emissao, "moeda": moeda,
        "valor_moeda": valor, "contrato_comercial": contrato_comercial,
        "status": status, "status_manual": status_manual, "data_credito": data_credito,
        "observacao": (form.get("observacao") or "").strip() or None,
    }

INVOICE_TYPE_CODES = set(INVOICE_STATUS_TYPES)

def distinct_display(values):
    values = [str(value) for value in values if value]
    if not values:
        return "-"
    if len(values) == 1:
        return values[0]
    return "Vários"

app.jinja_env.globals["distinct_display"] = distinct_display

def resolve_counterparty(conn, raw_id, required=False):
    if raw_id in (None, ""):
        if required:
            raise ValueError("Selecione um Banco / Contraparte cadastrado.")
        return None
    try:
        record_id = int(raw_id)
    except (TypeError, ValueError):
        raise ValueError("Selecione um Banco / Contraparte válido.")
    row = conn.execute("SELECT id, nome FROM contrapartes WHERE id=?", (record_id,)).fetchone()
    if not row:
        raise ValueError("O Banco / Contraparte selecionado não foi encontrado.")
    return row

def contract_metadata_from_form(form, conn):
    bank = resolve_counterparty(conn, form.get("banco_liquidacao_id"))
    numero_contrato = (form.get("numero_contrato_cambio") or form.get("numero_contrato") or "").strip()
    return {
        "numero_contrato": numero_contrato,
        "banco_liquidacao_id": bank["id"] if bank else None,
        "banco_liquidacao": bank["nome"] if bank else None,
        "data_fechamento": parse_date(form.get("data_fechamento")),
        "data_liquidacao": parse_date(form.get("data_liquidacao")),
        "taxa_cambio": optional_number(form.get("taxa_cambio")),
        "observacao": (form.get("contrato_observacao") or "").strip() or None,
    }

def contract_for_invoice(conn, invoice, metadata):
    contrato_id = form_record_id(metadata.get("contrato_id"))
    if contrato_id:
        contrato = conn.execute("SELECT * FROM contratos WHERE id=?", (contrato_id,)).fetchone()
        if not contrato:
            raise ValueError("O Contrato Câmbio selecionado não foi encontrado.")
    else:
        numero = metadata.get("numero_contrato")
        if not numero:
            raise ValueError("Informe um Contrato Câmbio existente ou o número de um novo Contrato Câmbio.")
        contrato = conn.execute("SELECT * FROM contratos WHERE numero_contrato=?", (numero,)).fetchone()
        if not contrato:
            conn.execute("""
                INSERT INTO contratos
                    (numero_contrato,banco_liquidacao_id,banco_liquidacao,data_fechamento,
                     data_liquidacao,moeda,taxa_cambio,observacao,status,data_contrato)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (numero, metadata.get("banco_liquidacao_id"), metadata.get("banco_liquidacao"),
                  metadata.get("data_fechamento"), metadata.get("data_liquidacao"), invoice["moeda"],
                  metadata.get("taxa_cambio"), metadata.get("observacao"), STATUS_PENDENTE,
                  metadata.get("data_fechamento")))
            contrato = conn.execute("SELECT * FROM contratos WHERE id=last_insert_rowid()").fetchone()
    has_allocations = conn.execute(
        "SELECT 1 FROM invoice_contrato_cambio WHERE contrato_id=? LIMIT 1", (contrato["id"],)
    ).fetchone()
    if contrato["moeda"] and contrato["moeda"] != invoice["moeda"] and has_allocations:
        raise ValueError("A moeda do Contrato Câmbio deve ser igual à moeda da Invoice.")
    if not has_allocations and contrato["moeda"] != invoice["moeda"]:
        conn.execute("UPDATE contratos SET moeda=? WHERE id=?", (invoice["moeda"], contrato["id"]))
    updates = {}
    for key in ("banco_liquidacao_id", "banco_liquidacao", "data_fechamento", "data_liquidacao", "taxa_cambio", "observacao"):
        if metadata.get(key) not in (None, "") and not contrato[key]:
            updates[key] = metadata[key]
    if updates:
        updates["data_contrato"] = updates.get("data_fechamento", contrato["data_contrato"])
        assignments = ",".join(f"{key}=?" for key in updates)
        conn.execute(f"UPDATE contratos SET {assignments} WHERE id=?",
                     tuple(updates.values()) + (contrato["id"],))
    return contrato["id"]

def sync_contract_cache(conn, contrato_id):
    contrato = conn.execute("SELECT * FROM contratos WHERE id=?", (contrato_id,)).fetchone()
    if not contrato:
        return None
    total = conn.execute("""
        SELECT COALESCE(SUM(v.valor_alocado),0) AS total,
               MIN(i.moeda) AS moeda, MIN(i.empresa_id) AS empresa_id,
               MIN(i.cliente_id) AS cliente_id, MIN(e.cnpj) AS cnpj,
               MIN(cl.nome) AS cliente, MIN(i.data_emissao) AS data_emissao
        FROM invoice_contrato_cambio v
        JOIN invoices i ON i.id=v.invoice_id
        LEFT JOIN empresas e ON e.id=i.empresa_id
        LEFT JOIN clientes cl ON cl.id=i.cliente_id
        WHERE v.contrato_id=?
    """, (contrato_id,)).fetchone()
    banks = conn.execute("""
        SELECT DISTINCT cp.nome
        FROM invoice_contrato_cambio v
        JOIN recebimentos_invoice r ON r.invoice_id=v.invoice_id
        JOIN contrapartes cp ON cp.id=r.banco_credito_id
        WHERE v.contrato_id=? AND cp.nome IS NOT NULL ORDER BY cp.nome
    """, (contrato_id,)).fetchall()
    total_value = decimal_value(total["total"])
    valor_reais = (total_value * decimal_value(contrato["taxa_cambio"])
                   if contrato["taxa_cambio"] is not None else None)
    linked = decimal_value(conn.execute("""
        SELECT COALESCE(SUM(valor),0) FROM due_movimentacoes
        WHERE contrato_id=? AND tipo='VINCULACAO'
    """, (contrato_id,)).fetchone()[0])
    saldo = contract_balance(total_value, linked)
    status = STATUS_CONCLUIDO if contrato["saldo_zerado_manual"] else status_from_balance(saldo, linked)
    conn.execute("""
        UPDATE contratos SET valor_moeda=?, valor_reais=?, moeda=COALESCE(?,moeda),
            cnpj=?, cliente=?, cliente_id=?, banco=?, banco_credito=?,
            data_contrato=COALESCE(data_fechamento,data_contrato), status=?
        WHERE id=?
    """, (float(total_value), float(valor_reais) if valor_reais is not None else None,
          total["moeda"] or contrato["moeda"], total["cnpj"], total["cliente"], total["cliente_id"],
          contrato["banco_liquidacao"], ", ".join(row[0] for row in banks) or None, status, contrato_id))
    return conn.execute("SELECT * FROM contratos WHERE id=?", (contrato_id,)).fetchone()

def sync_contracts_for_invoice(conn, invoice_id):
    contract_ids = [row[0] for row in conn.execute(
        "SELECT contrato_id FROM invoice_contrato_cambio WHERE invoice_id=?", (invoice_id,)
    ).fetchall()]
    for contract_id in contract_ids:
        sync_contract_cache(conn, contract_id)

def invoice_contracts_for_currency(conn, invoice):
    return conn.execute("""
        SELECT c.*, COALESCE(SUM(v.valor_alocado),0) AS valor_moeda_calculado
        FROM contratos c
        LEFT JOIN invoice_contrato_cambio v ON v.contrato_id=c.id
        WHERE c.moeda=? OR NOT EXISTS (SELECT 1 FROM invoice_contrato_cambio x WHERE x.contrato_id=c.id)
        GROUP BY c.id ORDER BY c.numero_contrato
    """, (invoice["moeda"],)).fetchall()

def invoice_detail_data(conn, invoice_id):
    summary = invoice_summary(conn, invoice_id)
    if not summary:
        return None
    empresas = conn.execute("""
        SELECT id, razao_social, apelido, cnpj FROM empresas
        ORDER BY CASE WHEN TRIM(COALESCE(apelido,''))<>'' THEN 0 ELSE 1 END, apelido, razao_social
    """).fetchall()
    clientes = clientes_for_form(conn)
    contrapartes = contrapartes_for_form(conn)
    dues = conn.execute("""
        SELECT d.id, d.numero_due, d.chave_acesso, d.moeda, d.valor_original
        FROM dues d ORDER BY d.numero_due
    """).fetchall()
    contratos = invoice_contracts_for_currency(conn, summary)
    return summary, empresas, clientes, contrapartes, dues, contratos

def invoice_stage_path(token):
    if not isinstance(token, str) or not re.fullmatch(r"[A-Za-z0-9_-]{20,128}", token):
        raise ValueError("A prévia da importação é inválida ou expirou.")
    return Path(tempfile.gettempdir()) / f"{INVOICE_IMPORT_STAGE_PREFIX}{token}.json"

def remove_invoice_stage(token=None):
    stage_token = token if token is not None else session.get("invoice_import_stage")
    if stage_token:
        try:
            invoice_stage_path(stage_token).unlink(missing_ok=True)
        except (OSError, ValueError):
            pass
    if session.get("invoice_import_stage") == stage_token:
        session.pop("invoice_import_stage", None)

def save_invoice_stage(payload):
    remove_invoice_stage()
    token = secrets.token_urlsafe(24)
    data = dict(payload)
    data["created_at"] = time.time()
    invoice_stage_path(token).write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    session["invoice_import_stage"] = token
    return token

def load_invoice_stage(token=None):
    token = (token or session.get("invoice_import_stage") or "").strip()
    if not token or token != session.get("invoice_import_stage"):
        raise ValueError("A prévia da importação não pertence a esta sessão.")
    path = invoice_stage_path(token)
    try:
        if time.time() - path.stat().st_mtime > INVOICE_IMPORT_STAGE_TTL:
            raise FileNotFoundError
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, OSError, ValueError, json.JSONDecodeError):
        remove_invoice_stage(token)
        raise ValueError("A prévia da importação é inválida ou expirou.")
    if not isinstance(payload, dict) or payload.get("version") != 1 or not isinstance(payload.get("rows"), list):
        remove_invoice_stage(token)
        raise ValueError("A prévia da importação é inválida ou expirou.")
    return payload

def invoice_import_cell(record, columns, column, pandas, default=None):
    if column not in columns:
        return default
    value = record[column]
    return None if pandas.isna(value) else value

def normalize_invoice_import_columns(columns):
    aliases = {
        "cnpj": "cnpj", "empresa": "empresa", "numero_invoice": "numero_invoice",
        "invoice": "numero_invoice", "numero_da_invoice": "numero_invoice",
        "tipo_documento": "tipo_documento", "tipo": "tipo_documento",
        "status": "status", "status_invoice": "status", "situacao": "status",
        "data_credito": "data_credito", "data_de_credito": "data_credito", "credito": "data_credito",
        "banco_credito": "banco_credito", "banco_de_credito": "banco_credito",
        "cliente": "cliente", "nome_cliente": "cliente", "data_emissao": "data_emissao", "emissao": "data_emissao",
        "competencia": "competencia", "safra": "competencia", "periodo": "competencia",
        "descricao_competencia": "competencia", "competencia_descricao": "competencia",
        # Aliases legados permanecem aceitos para não quebrar reprocessamentos antigos.
        "cliente_pais": "cliente_pais", "pais_cliente": "cliente_pais", "pais": "cliente_pais",
        "country": "cliente_pais",
        "moeda": "moeda", "valor_invoice": "valor_invoice", "valor_moeda": "valor_invoice",
        "contrato_comercial": "contrato_comercial", "numero_contrato_comercial": "contrato_comercial",
        "numero_contrato_cambio": "numero_contrato_cambio", "contrato_cambio": "numero_contrato_cambio",
        "numero_contrato": "numero_contrato_cambio", "contrato": "numero_contrato_cambio",
        "valor_alocado": "valor_alocado", "valor_cambio": "valor_alocado",
        "banco_liquidacao": "banco_liquidacao", "banco_de_liquidacao": "banco_liquidacao",
        "data_fechamento": "data_fechamento", "data_do_fechamento": "data_fechamento",
        "data_liquidacao": "data_liquidacao", "data_de_liquidacao": "data_liquidacao",
        "taxa_cambio": "taxa_cambio", "taxa_de_cambio": "taxa_cambio",
        "valor_brl": "valor_brl", "valor_em_brl": "valor_brl", "valor_reais": "valor_brl",
        "observacao": "observacao", "observacao_invoice": "observacao",
    }
    normalized = []
    for column in columns:
        text_value = unicodedata.normalize("NFKD", str(column)).encode("ascii", "ignore").decode("ascii")
        key = re.sub(r"[^a-zA-Z0-9]+", "_", text_value.strip().lower()).strip("_")
        normalized.append(aliases.get(key, key))
    duplicates = sorted({column for column in normalized if normalized.count(column) > 1})
    if duplicates:
        raise ValueError("O Excel contém colunas duplicadas após a normalização: " + ", ".join(duplicates) + ".")
    return normalized

def normalize_invoice_type(value):
    text_value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    text_value = re.sub(r"[^A-Z0-9]+", "_", text_value.upper()).strip("_")
    return {
        "PROFORMA": "PROFORMA", "PROFORMA_INVOICE": "PROFORMA",
        "COMMERCIAL_INVOICE": "COMMERCIAL_INVOICE",
        "COMMERCIAL": "COMMERCIAL_INVOICE", "SERVICE_INVOICE": "SERVICE_INVOICE",
        "SERVICE": "SERVICE_INVOICE", "DEBIT_NOTE": "DEBIT_NOTE", "DEBIT": "DEBIT_NOTE",
    }.get(text_value, text_value)

def prepare_invoice_import_rows(df, pandas):
    required = {"numero_invoice", "tipo_documento", "competencia", "valor_invoice"}
    missing = sorted(required - set(df.columns))
    if "cnpj" not in df.columns and "empresa" not in df.columns:
        missing.append("empresa/cnpj")
    if missing:
        raise ValueError("O Excel precisa conter as colunas: " + ", ".join(missing) + ".")
    rows = []
    invoice_groups = {}
    for line_number, (_, record) in enumerate(df.iterrows(), start=2):
        raw_values = {column: invoice_import_cell(record, df.columns, column, pandas) for column in df.columns}
        if not any(str(value or "").strip() for value in raw_values.values()):
            continue
        raw_empresa = raw_values.get("empresa") if "empresa" in df.columns else raw_values.get("cnpj")
        empresa = str(raw_empresa or "").strip()
        if not empresa:
            raise ValueError(f"Linha {line_number}: a empresa/CNPJ é obrigatória.")
        cnpj = normalize_import_cnpj(empresa)
        numero = str(raw_values.get("numero_invoice") or "").strip()
        if not numero:
            raise ValueError(f"Linha {line_number}: o numero_invoice é obrigatório.")
        tipo = normalize_invoice_type(raw_values.get("tipo_documento"))
        if tipo not in INVOICE_TYPE_CODES:
            raise ValueError(f"Linha {line_number}: tipo_documento inválido.")
        valor_invoice = parse_number(raw_values.get("valor_invoice"))
        if decimal_value(valor_invoice) <= 0:
            raise ValueError(f"Linha {line_number}: valor_invoice deve ser maior que zero.")
        data_emissao = parse_date(raw_values.get("data_emissao"))
        moeda = str(raw_values.get("moeda") or "USD").strip().upper()
        if not re.fullmatch(r"[A-Z]{3}", moeda):
            raise ValueError(f"Linha {line_number}: moeda inválida.")
        raw_status = raw_values.get("status")
        status_provided = raw_status is not None and str(raw_status).strip() != ""
        status = normalize_invoice_status(raw_status) if status_provided else None
        raw_data_credito = raw_values.get("data_credito")
        data_credito_provided = raw_data_credito is not None and str(raw_data_credito).strip() != ""
        data_credito = parse_date(raw_data_credito)
        if status == INVOICE_STATUS_AGUARDANDO_RECEBIMENTO:
            data_credito = None
        competencia = normalize_competencia_import(raw_values.get("competencia"))
        if not competencia:
            raise ValueError(f"Linha {line_number}: competencia e obrigatoria.")
        contrato_comercial = normalize_contract_commercial(raw_values.get("contrato_comercial"))
        banco_credito = str(raw_values.get("banco_credito") or "").strip() or None
        banco_liquidacao = str(raw_values.get("banco_liquidacao") or "").strip() or None
        contrato = str(raw_values.get("numero_contrato_cambio") or "").strip() or None
        valor_brl = optional_number(raw_values.get("valor_brl"))
        importa_cambio = bool(
            contrato or banco_liquidacao or raw_values.get("data_fechamento") is not None
            or raw_values.get("data_liquidacao") is not None
            or raw_values.get("taxa_cambio") is not None or valor_brl is not None
            or "valor_alocado" in df.columns
        )
        valor_alocado = optional_number(raw_values.get("valor_alocado")) if "valor_alocado" in df.columns else None
        if contrato and valor_alocado is None:
            valor_alocado = valor_invoice
        if importa_cambio and not contrato and any((banco_liquidacao, valor_brl, valor_alocado,
                                                    raw_values.get("data_fechamento"),
                                                    raw_values.get("data_liquidacao"),
                                                    raw_values.get("taxa_cambio"))):
            raise ValueError(f"Linha {line_number}: os dados de câmbio exigem numero_contrato_cambio.")
        if contrato and (valor_alocado is None or decimal_value(valor_alocado) <= 0):
            raise ValueError(f"Linha {line_number}: valor_alocado é obrigatório quando há Contrato Câmbio.")
        data_fechamento = parse_date(raw_values.get("data_fechamento"))
        data_liquidacao = parse_date(raw_values.get("data_liquidacao"))
        taxa_cambio = optional_number(raw_values.get("taxa_cambio"))
        if data_fechamento and data_liquidacao and data_liquidacao < data_fechamento:
            raise ValueError(f"Linha {line_number}: data_liquidacao não pode ser anterior à data_fechamento.")
        if taxa_cambio is not None and decimal_value(taxa_cambio) <= 0:
            raise ValueError(f"Linha {line_number}: taxa_cambio deve ser maior que zero.")
        if valor_brl is not None and decimal_value(valor_brl) <= 0:
            raise ValueError(f"Linha {line_number}: valor_brl deve ser maior que zero.")
        if valor_brl is not None and taxa_cambio is None:
            raise ValueError(f"Linha {line_number}: taxa_cambio é obrigatória quando valor_brl é informado.")
        if valor_brl is not None and valor_alocado is not None and taxa_cambio is not None:
            calculado_brl = decimal_value(valor_alocado) * decimal_value(taxa_cambio)
            if abs(calculado_brl - decimal_value(valor_brl)) > Decimal("0.05"):
                raise ValueError(f"Linha {line_number}: valor_brl não corresponde ao valor_moeda x taxa_cambio.")
        row = {
            "row_id": f"r{line_number}", "source_row": line_number, "empresa": empresa, "cnpj": cnpj,
            "numero_invoice": numero, "tipo_documento": tipo, "competencia": competencia,
            "cliente": normalize_client_name_display(raw_values.get("cliente")),
            "data_emissao": data_emissao, "moeda": moeda, "valor_invoice": float(valor_invoice),
            "contrato_comercial": contrato_comercial,
            "status": status, "status_provided": status_provided,
            "data_credito": data_credito, "data_credito_provided": data_credito_provided,
            "banco_credito": banco_credito, "banco_liquidacao": banco_liquidacao,
            "numero_contrato_cambio": contrato,
            "valor_brl": float(valor_brl) if valor_brl is not None else None,
            "valor_alocado": float(valor_alocado) if valor_alocado is not None else None,
            "legacy_valor_alocado": float(valor_alocado) if valor_alocado is not None else None,
            "importa_cambio": importa_cambio,
            "data_fechamento": data_fechamento,
            "data_liquidacao": data_liquidacao,
            "taxa_cambio": taxa_cambio,
            "observacao": str(raw_values.get("observacao") or "").strip() or None,
        }
        if not importa_cambio:
            row.pop("valor_alocado", None)
        company_key = cnpj or f"empresa:{normalize_client_name_key(empresa)}"
        key = (company_key, numero, tipo)
        invoice_groups.setdefault(key, []).append(row)
        rows.append(row)
    if not rows:
        raise ValueError("A planilha não contém Invoices válidas para importar.")
    for key, group in invoice_groups.items():
        first = group[0]
        for row in group[1:]:
            for field in ("valor_invoice", "cliente", "data_emissao", "moeda", "contrato_comercial", "competencia"):
                if str(row.get(field) or "") != str(first.get(field) or ""):
                    raise ValueError(
                        f"Invoice {key[1]} possui dados comerciais divergentes entre as linhas."
                    )
        statuses = {row["status"] for row in group if row.get("status_provided")}
        if len(statuses) > 1:
            raise ValueError(f"Invoice {key[1]} possui status divergentes entre as linhas.")
        if statuses:
            imported_status = statuses.pop()
            for row in group:
                row["status"] = imported_status
                row["status_provided"] = True
        credit_dates = {row["data_credito"] for row in group if row.get("data_credito_provided")}
        if len(credit_dates) > 1:
            raise ValueError(f"Invoice {key[1]} possui datas de crédito divergentes entre as linhas.")
        group_data_credito = credit_dates.pop() if credit_dates else None
        if group_data_credito:
            for row in group:
                row["data_credito"] = group_data_credito
                row["data_credito_provided"] = True
        credit_banks = {row["banco_credito"] for row in group if row.get("banco_credito")}
        if len(credit_banks) > 1:
            raise ValueError(f"Invoice {key[1]} possui bancos de crédito divergentes entre as linhas.")
        if credit_banks:
            group_banco_credito = credit_banks.pop()
            for row in group:
                row["banco_credito"] = group_banco_credito
        group_status = next((row["status"] for row in group if row.get("status_provided")), None)
        if group_status == INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO and not group_data_credito:
            raise ValueError(
                f"Invoice {key[1]}: data_credito é obrigatória quando o status é RECEBIDO AGUARDANDO CAMBIO."
            )
    contract_groups = {}
    for row in rows:
        if row["numero_contrato_cambio"]:
            contract_groups.setdefault(row["numero_contrato_cambio"], []).append(row)
    for numero, group in contract_groups.items():
        fields = ("banco_liquidacao", "data_fechamento", "data_liquidacao", "taxa_cambio", "moeda")
        first = group[0]
        for row in group[1:]:
            if any(str(row.get(field) or "") != str(first.get(field) or "") for field in fields):
                raise ValueError(f"Contrato Câmbio {numero} possui metadados divergentes entre as linhas.")
    return rows

def invoice_identity_rows(conn, rows):
    result = {}
    for row in rows:
        empresa = conn.execute("SELECT id FROM empresas WHERE cnpj=?", (row["cnpj"],)).fetchone()
        if not empresa:
            raise ValueError(f"A empresa com CNPJ {row['cnpj']} não está cadastrada.")
        db_key = (empresa["id"], row["numero_invoice"], row["tipo_documento"])
        display_key = (row["cnpj"], row["numero_invoice"], row["tipo_documento"])
        result[display_key] = conn.execute("""
            SELECT i.*, COALESCE((SELECT SUM(valor_moeda) FROM recebimentos_invoice WHERE invoice_id=i.id),0) AS total_recebido,
                   COALESCE((SELECT SUM(valor_alocado) FROM invoice_contrato_cambio WHERE invoice_id=i.id),0) AS total_cambio
            FROM invoices i WHERE i.empresa_id=? AND i.numero_invoice=? AND i.tipo_documento=?
        """, db_key).fetchone()
    return result

def invoice_import_snapshot(row):
    if not row:
        return None
    data = dict(row)
    return {key: data.get(key) for key in (
        "id", "empresa_id", "numero_invoice", "tipo_documento", "competencia_id", "cliente_id", "data_emissao",
        "data_credito", "moeda", "valor_moeda", "contrato_comercial", "status", "status_manual",
        "total_recebido", "total_cambio"
    )}

def invoice_import_counterparty(conn, name, field_label):
    name = str(name or "").strip()
    if not name:
        return None
    counterparty = conn.execute(
        "SELECT id, nome FROM contrapartes WHERE nome=? COLLATE NOCASE", (name,)
    ).fetchone()
    if not counterparty:
        raise ValueError(f"O {field_label} {name} não está cadastrado em Configurações.")
    return counterparty

def invoice_import_bank_key(field, name):
    return f"{field}|{normalize_client_name_key(name)}"

def resolve_invoice_import_banks(conn, rows, bank_overrides=None):
    """Valida bancos na previa e aplica o vinculo escolhido pelo usuario."""
    preview = bank_overrides is None
    bank_overrides = bank_overrides or {}
    counterparties = conn.execute("SELECT id, nome FROM contrapartes ORDER BY nome").fetchall()
    fields = (("banco_credito", "Banco de Cr\u00e9dito"), ("banco_liquidacao", "Banco de Liquida\u00e7\u00e3o"))
    suggestions = []
    suggestions_by_key = {}
    for row in rows:
        for field, field_label in fields:
            name = str(row.get(field) or "").strip() or None
            if not name:
                continue
            counterparty = next((item for item in counterparties
                                 if item["nome"].casefold() == name.casefold()), None)
            if not counterparty:
                key = invoice_import_bank_key(field, name)
                selected_id = bank_overrides.get(key)
                if selected_id in (None, ""):
                    if preview:
                        suggestion = suggestions_by_key.get(key)
                        if not suggestion:
                            suggestion = {
                                "suggestion_id": f"b{len(suggestions) + 1}", "key": key,
                                "nome": name, "tipo": field_label, "linhas": [],
                            }
                            suggestions_by_key[key] = suggestion
                            suggestions.append(suggestion)
                        suggestion["linhas"].append(row["source_row"])
                        continue
                    raise ValueError(
                        f"O {field_label} n\u00e3o est\u00e1 cadastrado. Selecione um banco na pr\u00e9via."
                    )
                try:
                    selected_id = int(selected_id)
                except (TypeError, ValueError):
                    raise ValueError(f"O banco selecionado para {name} \u00e9 inv\u00e1lido.")
                counterparty = next((item for item in counterparties if item["id"] == selected_id), None)
                if not counterparty:
                    raise ValueError(f"O banco selecionado para {name} n\u00e3o foi encontrado.")
            row[field] = counterparty["nome"]
    return suggestions

def apply_invoice_import_receipt(conn, invoice_id, row, status=None):
    data_credito = row.get("data_credito")
    banco_nome = row.get("banco_credito")
    if banco_nome and not data_credito:
        raise ValueError("banco_credito exige data_credito para registrar o recebimento.")
    if not data_credito or status == INVOICE_STATUS_AGUARDANDO_RECEBIMENTO:
        return
    bank = invoice_import_counterparty(conn, banco_nome, "Banco de Crédito")
    amount = decimal_value(row["valor_invoice"])
    receipts = conn.execute(
        "SELECT id, valor_moeda FROM recebimentos_invoice WHERE invoice_id=? ORDER BY id",
        (invoice_id,)
    ).fetchall()
    total_received = sum((decimal_value(receipt["valor_moeda"]) for receipt in receipts), Decimal("0"))
    if total_received >= amount - SALDO_TOLERANCE:
        return
    remaining = amount - total_received
    conn.execute("""
        INSERT INTO recebimentos_invoice
            (invoice_id,banco_credito_id,data_credito,moeda,valor_moeda,documento,observacao)
        VALUES (?,?,?,?,?,?,?)
    """, (invoice_id, bank["id"] if bank else None, data_credito, row["moeda"], float(remaining), None, None))

def apply_invoice_import_rows(conn, rows, replace_existing=True, country_overrides=None,
                              competency_overrides=None, client_overrides=None, bank_overrides=None):
    resolve_invoice_import_companies(conn, rows)
    resolve_invoice_import_banks(conn, rows, bank_overrides=bank_overrides)
    ensure_invoice_import_clients(conn, rows, country_overrides=country_overrides,
                                  client_overrides=client_overrides)
    ensure_invoice_import_competencies(conn, rows, new_competency_overrides=competency_overrides)
    groups = {}
    for row in rows:
        groups.setdefault((row["cnpj"], row["numero_invoice"], row["tipo_documento"]), []).append(row)
    changed_invoices, changed_contracts = set(), set()
    inserted = updated = 0
    for key, group in groups.items():
        company = conn.execute("SELECT id FROM empresas WHERE cnpj=?", (key[0],)).fetchone()
        if not company:
            raise ValueError(f"A empresa com CNPJ {key[0]} não está cadastrada.")
        first = group[0]
        status_provided = first.get("status_provided", first.get("status") is not None)
        imported_status = first.get("status") if status_provided else None
        data_credito_provided = first.get(
            "data_credito_provided", first.get("data_credito") is not None
        )
        imported_data_credito = first.get("data_credito")
        if imported_status == INVOICE_STATUS_AGUARDANDO_RECEBIMENTO:
            imported_data_credito = None
        elif imported_status == INVOICE_STATUS_RECEBIDA_AGUARDANDO_CAMBIO and not imported_data_credito:
            raise ValueError(
                f"Invoice {first['numero_invoice']}: data_credito é obrigatória quando o status é RECEBIDO AGUARDANDO CAMBIO."
            )
        cliente_id = first.get("cliente_id")
        current = conn.execute("""
            SELECT * FROM invoices WHERE empresa_id=? AND numero_invoice=? AND tipo_documento=?
        """, (company["id"], first["numero_invoice"], first["tipo_documento"])).fetchone()
        if current:
            summary = invoice_summary(conn, current["id"])
            if decimal_value(first["valor_invoice"]) < summary["total_recebido"]:
                raise ValueError(f"Invoice {first['numero_invoice']} não pode ficar abaixo do total recebido.")
            if first["moeda"] != current["moeda"] and (
                summary["total_recebido"] > SALDO_TOLERANCE or
                summary["total_cambio"] > SALDO_TOLERANCE
            ):
                raise ValueError(f"Invoice {first['numero_invoice']} não pode mudar de moeda com recebimento ou câmbio vinculado.")
            updated += 1
            invoice_id = current["id"]
            if not data_credito_provided and current and imported_status != INVOICE_STATUS_AGUARDANDO_RECEBIMENTO:
                imported_data_credito = current["data_credito"]
            if status_provided:
                conn.execute("""
                    UPDATE invoices SET cliente_id=?, competencia_id=?, data_emissao=?, moeda=?, valor_moeda=?,
                        contrato_comercial=?, status=?, status_manual=1, data_credito=?, observacao=? WHERE id=?
                """, (cliente_id, first["competencia_id"], first["data_emissao"], first["moeda"], first["valor_invoice"],
                      first["contrato_comercial"], imported_status, imported_data_credito, first["observacao"], invoice_id))
            else:
                conn.execute("""
                    UPDATE invoices SET cliente_id=?, competencia_id=?, data_emissao=?, moeda=?, valor_moeda=?,
                        contrato_comercial=?, data_credito=?, observacao=? WHERE id=?
                """, (cliente_id, first["competencia_id"], first["data_emissao"], first["moeda"], first["valor_invoice"],
                      first["contrato_comercial"], imported_data_credito, first["observacao"], invoice_id))
            old_contracts = [row[0] for row in conn.execute(
                "SELECT contrato_id FROM invoice_contrato_cambio WHERE invoice_id=?", (invoice_id,)
            ).fetchall()]
            if replace_existing:
                conn.execute("DELETE FROM invoice_contrato_cambio WHERE invoice_id=?", (invoice_id,))
                changed_contracts.update(old_contracts)
        else:
            inserted += 1
            cursor = conn.execute("""
                INSERT INTO invoices
                    (empresa_id,numero_invoice,tipo_documento,competencia_id,cliente_id,data_emissao,moeda,valor_moeda,
                     contrato_comercial,status,status_manual,data_credito,observacao)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (company["id"], first["numero_invoice"], first["tipo_documento"], first["competencia_id"], cliente_id,
                  first["data_emissao"], first["moeda"], first["valor_invoice"],
                  first["contrato_comercial"], imported_status or INVOICE_STATUS_AGUARDANDO_RECEBIMENTO,
                  1 if status_provided else 0, imported_data_credito, first["observacao"]))
            invoice_id = cursor.lastrowid
        apply_invoice_import_receipt(conn, invoice_id, first, status=imported_status)
        allocation_total = sum((decimal_value(
            row.get("valor_alocado") if row.get("valor_alocado") is not None
            else row.get("legacy_valor_alocado")
        ) for row in group), Decimal("0"))
        summary_after_header = invoice_summary(conn, invoice_id)
        validate_invoice_balances(summary_after_header, replacement_cambio=allocation_total)
        allocation_by_contract = {}
        for row in group:
            if row["numero_contrato_cambio"]:
                allocation_by_contract.setdefault(row["numero_contrato_cambio"], []).append(row)
        for numero, contract_rows in allocation_by_contract.items():
            first_contract = contract_rows[0]
            bank = invoice_import_counterparty(conn, first_contract["banco_liquidacao"], "Banco de Liquidação")
            metadata = {
                "numero_contrato": numero, "banco_liquidacao_id": bank["id"] if bank else None,
                "banco_liquidacao": bank["nome"] if bank else None,
                "data_fechamento": first_contract["data_fechamento"],
                "data_liquidacao": first_contract["data_liquidacao"],
                "taxa_cambio": first_contract["taxa_cambio"],
                "observacao": first_contract["observacao"],
            }
            contract_id = contract_for_invoice(conn, {"moeda": first["moeda"]}, metadata)
            changed_contracts.add(contract_id)
            amount = sum((decimal_value(
                row.get("valor_alocado") if row.get("valor_alocado") is not None
                else row.get("legacy_valor_alocado")
            ) for row in contract_rows), Decimal("0"))
            conn.execute("""
                INSERT INTO invoice_contrato_cambio(invoice_id,contrato_id,valor_alocado,observacao)
                VALUES (?,?,?,?)
                ON CONFLICT(invoice_id,contrato_id) DO UPDATE SET valor_alocado=excluded.valor_alocado,
                    observacao=excluded.observacao
            """, (invoice_id, contract_id, float(amount), first_contract["observacao"]))
        changed_invoices.add(invoice_id)
    for contract_id in changed_contracts:
        if conn.execute("SELECT COUNT(*) FROM invoice_contrato_cambio WHERE contrato_id=?", (contract_id,)).fetchone()[0]:
            sync_contract_cache(conn, contract_id)
        elif not conn.execute("SELECT 1 FROM due_contratos WHERE contrato_id=?", (contract_id,)).fetchone():
            conn.execute("DELETE FROM contratos WHERE id=?", (contract_id,))
    for invoice_id in changed_invoices:
        refresh_invoice_status(conn, invoice_id)
    return {"inserted": inserted, "updated": updated, "invoices": len(changed_invoices)}

def invoice_filter_query(args):
    where, params = [], []
    if args.get("numero_invoice"):
        where.append("i.numero_invoice LIKE ?"); params.append(f"%{args['numero_invoice'].strip()}%")
    if args.get("contrato_comercial"):
        where.append("i.contrato_comercial LIKE ?"); params.append(f"%{args['contrato_comercial'].strip()}%")
    if args.get("tipo_documento"):
        where.append("i.tipo_documento=?"); params.append(args["tipo_documento"].strip().upper())
    if args.get("empresa_id"):
        where.append("i.empresa_id=?"); params.append(form_record_id(args["empresa_id"]))
    if args.get("cliente_id"):
        where.append("i.cliente_id=?"); params.append(form_record_id(args["cliente_id"]))
    if args.get("competencia_id"):
        where.append("i.competencia_id=?"); params.append(form_record_id(args["competencia_id"]))
    if args.get("moeda"):
        where.append("i.moeda=?"); params.append(args["moeda"].strip().upper())
    if args.get("status"):
        try:
            status_filter = normalize_invoice_status(args["status"])
            where.append("i.status=?"); params.append(status_filter)
        except ValueError:
            flash("Status de Invoice invalido.", "danger")
    try:
        if args.get("data_de"):
            where.append("i.data_emissao>=?"); params.append(parse_date(args["data_de"]))
        if args.get("data_ate"):
            where.append("i.data_emissao<=?"); params.append(parse_date(args["data_ate"]))
    except ValueError as exc:
        flash(str(exc), "danger")
    clause = " WHERE " + " AND ".join(where) if where else ""
    return clause, params


def invoice_sorting(args):
    sort_fields = {
        "numero_invoice": "i.numero_invoice", "contrato_comercial": "i.contrato_comercial",
        "tipo_documento": "i.tipo_documento", "competencia_id": "i.competencia_id",
        "data_emissao": "i.data_emissao", "moeda": "i.moeda", "valor_moeda": "i.valor_moeda",
        "status": "i.status",
    }
    sort = args.get("sort", "data_emissao")
    sort = sort if sort in sort_fields else "data_emissao"
    direction = "ASC" if args.get("direction", "desc").lower() == "asc" else "DESC"
    return sort_fields, sort, direction


@app.route("/invoices")
def lista_invoices():
    filters = {key: value for key, value in request.args.items() if key not in {"sort", "direction", "page"} and value}
    clause, params = invoice_filter_query(request.args)
    sort_fields, sort, direction = invoice_sorting(request.args)
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    per_page = 20
    conn = db()
    for row in conn.execute("SELECT id FROM invoices").fetchall():
        refresh_invoice_status(conn, row["id"])
    conn.commit()
    total = conn.execute(f"SELECT COUNT(*) FROM invoices i{clause}", params).fetchone()[0]
    pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, pages)
    rows = conn.execute(f"""
        SELECT i.id FROM invoices i {clause}
        ORDER BY {sort_fields[sort]} {direction}, i.id DESC LIMIT ? OFFSET ?
    """, params + [per_page, (page - 1) * per_page]).fetchall()
    invoices = []
    for row in rows:
        item = invoice_summary(conn, row["id"])
        invoices.append(item)
    empresas = conn.execute("SELECT id,razao_social,apelido,cnpj FROM empresas ORDER BY razao_social").fetchall()
    clientes = clientes_for_form(conn)
    competencias = competencias_for_empresa(conn, None)
    moedas = [row[0] for row in conn.execute("SELECT DISTINCT moeda FROM invoices ORDER BY moeda").fetchall()]
    conn.commit(); conn.close()
    sort_links = {}
    for key in sort_fields:
        sort_links[key] = {**filters, "sort": key, "direction": "asc" if sort == key and direction == "DESC" else "desc"}
    previous_args = {**filters, "sort": sort, "direction": direction, "page": page - 1}
    next_args = {**filters, "sort": sort, "direction": direction, "page": page + 1}
    return render_template("invoices.html", invoices=invoices, total=total, page=page, pages=pages,
                           filters=filters, empresas=empresas, clientes=clientes, competencias=competencias, moedas=moedas,
                           invoice_statuses=INVOICE_STATUS_LABELS, sort=sort, direction=direction,
                           sort_links=sort_links, previous_args=previous_args, next_args=next_args)


@app.route("/invoices/relatorios")
def relatorios_invoices():
    return render_template("invoice_relatorios.html", **build_invoice_report_context())


@app.route("/invoices/exportar")
def exportar_invoices():
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    conn = db()
    for row in conn.execute("SELECT id FROM invoices").fetchall():
        refresh_invoice_status(conn, row["id"])
    conn.commit()

    clause, params = invoice_filter_query(request.args)
    sort_fields, sort, direction = invoice_sorting(request.args)
    rows = conn.execute(f"""
        SELECT i.id
        FROM invoices i {clause}
        ORDER BY {sort_fields[sort]} {direction}, i.id DESC
    """, params).fetchall()
    invoice_ids = [row["id"] for row in rows]
    invoices = [invoice_summary(conn, invoice_id) for invoice_id in invoice_ids]
    invoices = [invoice for invoice in invoices if invoice]
    invoice_by_id = {invoice["id"]: invoice for invoice in invoices}

    receipts = []
    changes = []
    due_links = []
    if invoice_ids:
        placeholders = ",".join("?" for _ in invoice_ids)
        receipts = conn.execute(f"""
            SELECT r.id AS recebimento_id, r.invoice_id, r.banco_credito_id,
                   r.data_credito, r.moeda, r.valor_moeda, r.documento,
                   r.observacao, r.created_at AS recebimento_criado_em,
                   i.numero_invoice, i.tipo_documento, i.data_emissao,
                   i.moeda AS invoice_moeda, i.valor_moeda AS invoice_valor_moeda,
                   e.id AS empresa_id, e.cnpj AS empresa_cnpj,
                   e.razao_social AS empresa_razao_social, e.apelido AS empresa_apelido,
                   cl.id AS cliente_id, cl.nome AS cliente_nome, cl.pais AS cliente_pais,
                   cp.nome AS banco_credito_nome
            FROM recebimentos_invoice r
            JOIN invoices i ON i.id=r.invoice_id
            JOIN empresas e ON e.id=i.empresa_id
            LEFT JOIN clientes cl ON cl.id=i.cliente_id
            LEFT JOIN contrapartes cp ON cp.id=r.banco_credito_id
            WHERE r.invoice_id IN ({placeholders})
            ORDER BY i.numero_invoice, r.data_credito, r.id
        """, invoice_ids).fetchall()
        changes = conn.execute(f"""
            SELECT v.id AS vinculo_id, v.invoice_id, v.contrato_id,
                   v.valor_alocado, v.observacao AS vinculo_observacao,
                   v.created_at AS vinculo_criado_em,
                   i.numero_invoice, i.tipo_documento, i.data_emissao,
                   i.moeda AS invoice_moeda, i.valor_moeda AS invoice_valor_moeda,
                   e.id AS empresa_id, e.cnpj AS empresa_cnpj,
                   e.razao_social AS empresa_razao_social, e.apelido AS empresa_apelido,
                   cl.id AS cliente_id, cl.nome AS cliente_nome, cl.pais AS cliente_pais,
                   c.id AS contrato_registro_id, c.numero_contrato, c.banco_liquidacao_id,
                   c.banco, c.banco_id, c.banco_credito, c.banco_liquidacao,
                   c.data_contrato, c.data_recebimento, c.data_fechamento,
                   c.data_liquidacao, c.cnpj AS contrato_cnpj,
                   c.cliente AS contrato_cliente, c.cliente_id AS contrato_cliente_id,
                   c.moeda AS contrato_moeda, c.valor_moeda AS contrato_valor_moeda,
                   c.taxa_cambio AS contrato_taxa_cambio, c.valor_reais AS contrato_valor_reais,
                   c.status AS contrato_status, c.saldo_zerado_manual,
                   c.observacao AS contrato_observacao, c.competencia_id AS contrato_competencia_id,
                   c.created_at AS contrato_criado_em,
                   COALESCE((SELECT SUM(x.valor_alocado)
                             FROM invoice_contrato_cambio x
                             WHERE x.contrato_id=c.id), 0) AS contrato_total_alocado
            FROM invoice_contrato_cambio v
            JOIN invoices i ON i.id=v.invoice_id
            JOIN empresas e ON e.id=i.empresa_id
            LEFT JOIN clientes cl ON cl.id=i.cliente_id
            JOIN contratos c ON c.id=v.contrato_id
            WHERE v.invoice_id IN ({placeholders})
            ORDER BY i.numero_invoice, c.numero_contrato, v.id
        """, invoice_ids).fetchall()
        due_links = conn.execute(f"""
            SELECT di.id AS vinculo_id, di.invoice_id, di.due_id,
                   di.valor_vinculado, di.observacao AS vinculo_observacao,
                   di.created_at AS vinculo_criado_em,
                   i.numero_invoice, i.tipo_documento, i.data_emissao,
                   i.moeda AS invoice_moeda, i.valor_moeda AS invoice_valor_moeda,
                   e.id AS empresa_id, e.cnpj AS empresa_cnpj,
                   e.razao_social AS empresa_razao_social, e.apelido AS empresa_apelido,
                   cl.id AS cliente_id, cl.nome AS cliente_nome, cl.pais AS cliente_pais,
                   d.numero_due, d.chave_acesso, d.data_due,
                   d.cnpj AS due_cnpj, d.cliente AS due_cliente,
                   d.moeda AS due_moeda, d.valor_original AS due_valor_original,
                   d.status AS due_status, d.observacao AS due_observacao,
                   d.created_at AS due_criado_em, d.competencia_id AS due_competencia_id,
                   comp.descricao AS due_competencia_descricao
            FROM due_invoice di
            JOIN invoices i ON i.id=di.invoice_id
            JOIN empresas e ON e.id=i.empresa_id
            LEFT JOIN clientes cl ON cl.id=i.cliente_id
            JOIN dues d ON d.id=di.due_id
            LEFT JOIN competencias comp ON comp.id=d.competencia_id
            WHERE di.invoice_id IN ({placeholders})
                   ORDER BY i.numero_invoice, d.numero_due, di.id
        """, invoice_ids).fetchall()
        due_usage = {
            row["due_id"]: decimal_value(due_effect(conn, row["due_id"]))
            for row in due_links
        }
    else:
        due_usage = {}
    conn.close()

    date_fields = {
        "data_emissao", "data_credito", "competencia_data_inicial", "competencia_data_final",
        "created_at", "recebimento_criado_em", "vinculo_criado_em", "contrato_criado_em",
        "data_contrato", "data_recebimento", "data_fechamento", "data_liquidacao",
        "due_data_due", "due_criado_em",
    }
    date_list_fields = {"datas_credito", "datas_fechamento", "datas_liquidacao"}

    def excel_value(value, field=None):
        if value is None:
            return None
        if field in date_fields:
            return date_br(value)
        if field in date_list_fields:
            return "; ".join(date_br(item) for item in value if date_br(item))
        if isinstance(value, (list, tuple, set)):
            return "; ".join(str(item) for item in value if item)
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, bool):
            return "Sim" if value else "Nao"
        return value

    def mapped_row(item, fields):
        data = dict(item)
        return {label: excel_value(data.get(key), key) for key, label in fields}

    invoice_fields = [
        ("id", "ID"), ("empresa_id", "Empresa ID"), ("empresa_cnpj", "Empresa - CNPJ"),
        ("empresa_razao_social", "Empresa - Razao social"), ("empresa_apelido", "Empresa - Apelido"),
        ("numero_invoice", "Numero da Invoice"), ("tipo_documento", "Tipo"),
        ("contrato_comercial", "Contrato comercial"), ("competencia_id", "Competencia ID"),
        ("competencia_descricao", "Competencia"), ("competencia_data_inicial", "Competencia - inicio"),
        ("competencia_data_final", "Competencia - fim"), ("cliente_id", "Cliente ID"),
        ("cliente_nome", "Cliente"), ("cliente_pais", "Pais do cliente"),
        ("data_emissao", "Data de emissao"), ("data_credito", "Data de credito"),
        ("moeda", "Moeda"), ("valor_moeda", "Valor da Invoice"), ("total_recebido", "Total recebido"),
        ("saldo_recebimento", "Saldo de recebimento"), ("total_cambio", "Total de cambio"),
        ("saldo_cambio", "Saldo sem cambio"), ("taxa_cambio_media", "Taxa media de cambio"),
        ("valor_brl", "Valor BRL"), ("status", "Status"), ("status_manual", "Status manual"),
        ("bancos_credito", "Bancos de credito"), ("bancos_liquidacao", "Bancos de liquidacao"),
        ("contratos_numeros", "Contratos de cambio"), ("datas_credito", "Datas de credito"),
        ("datas_fechamento", "Datas de fechamento"), ("datas_liquidacao", "Datas de liquidacao"),
        ("observacao", "Observacao"), ("created_at", "Criado em"),
    ]
    invoice_data = []
    for invoice in invoices:
        item = dict(invoice)
        item["tipo_documento"] = invoice_type_label(item["tipo_documento"])
        item["status"] = INVOICE_STATUS_LABELS.get(item["status"], item["status"])
        item["status_manual"] = bool(item["status_manual"])
        invoice_data.append(mapped_row(item, invoice_fields))

    receipt_fields = [
        ("recebimento_id", "Recebimento ID"), ("invoice_id", "Invoice ID"),
        ("numero_invoice", "Numero da Invoice"), ("tipo_documento", "Tipo"),
        ("empresa_id", "Empresa ID"), ("empresa_cnpj", "Empresa - CNPJ"),
        ("empresa_razao_social", "Empresa - Razao social"), ("empresa_apelido", "Empresa - Apelido"),
        ("cliente_id", "Cliente ID"), ("cliente_nome", "Cliente"), ("cliente_pais", "Pais do cliente"),
        ("data_emissao", "Data de emissao"), ("invoice_moeda", "Moeda da Invoice"),
        ("invoice_valor_moeda", "Valor da Invoice"), ("banco_credito_id", "Banco credito ID"),
        ("banco_credito_nome", "Banco de credito"), ("data_credito", "Data de credito"),
        ("moeda", "Moeda do recebimento"), ("valor_moeda", "Valor recebido"),
        ("documento", "Documento"), ("observacao", "Observacao"),
        ("recebimento_criado_em", "Recebimento criado em"), ("invoice_status", "Status da Invoice"),
    ]
    receipt_data = []
    for row in receipts:
        item = dict(row)
        invoice = invoice_by_id.get(item["invoice_id"])
        item["tipo_documento"] = invoice_type_label(item["tipo_documento"])
        item["invoice_status"] = INVOICE_STATUS_LABELS.get(invoice["status"]) if invoice else None
        receipt_data.append(mapped_row(item, receipt_fields))

    change_fields = [
        ("vinculo_id", "Vinculo ID"), ("invoice_id", "Invoice ID"),
        ("numero_invoice", "Numero da Invoice"), ("tipo_documento", "Tipo"),
        ("empresa_id", "Empresa ID"), ("empresa_cnpj", "Empresa - CNPJ"),
        ("empresa_razao_social", "Empresa - Razao social"), ("empresa_apelido", "Empresa - Apelido"),
        ("cliente_id", "Cliente ID"), ("cliente_nome", "Cliente"), ("cliente_pais", "Pais do cliente"),
        ("data_emissao", "Data de emissao"), ("invoice_moeda", "Moeda da Invoice"),
        ("invoice_valor_moeda", "Valor da Invoice"), ("contrato_id", "Contrato ID"),
        ("numero_contrato", "Numero do Contrato Cambio"), ("banco_liquidacao_id", "Banco liquidacao ID"),
        ("banco", "Banco legado"), ("banco_id", "Banco credito ID"),
        ("banco_credito", "Banco de credito"), ("banco_liquidacao", "Banco de liquidacao"),
        ("data_contrato", "Data do contrato"), ("data_recebimento", "Data de recebimento"),
        ("data_fechamento", "Data de fechamento"), ("data_liquidacao", "Data de liquidacao"),
        ("contrato_cnpj", "Contrato - CNPJ"), ("contrato_cliente", "Contrato - Cliente"),
        ("contrato_cliente_id", "Contrato - Cliente ID"), ("contrato_moeda", "Moeda do contrato"),
        ("contrato_valor_moeda", "Valor do contrato"), ("contrato_taxa_cambio", "Taxa de cambio"),
        ("contrato_valor_reais", "Valor do contrato em reais"), ("valor_alocado", "Valor alocado"),
        ("contrato_total_alocado", "Total alocado no contrato"), ("contrato_saldo", "Saldo do contrato"),
        ("contrato_status", "Status do contrato"), ("saldo_zerado_manual", "Saldo zerado manualmente"),
        ("contrato_competencia_id", "Contrato - Competencia ID"),
        ("vinculo_observacao", "Observacao do vinculo"), ("contrato_observacao", "Observacao do contrato"),
        ("vinculo_criado_em", "Vinculo criado em"), ("contrato_criado_em", "Contrato criado em"),
        ("invoice_status", "Status da Invoice"),
    ]
    change_data = []
    for row in changes:
        item = dict(row)
        invoice = invoice_by_id.get(item["invoice_id"])
        item["tipo_documento"] = invoice_type_label(item["tipo_documento"])
        item["invoice_status"] = INVOICE_STATUS_LABELS.get(invoice["status"]) if invoice else None
        item["contrato_saldo"] = Decimal("0") if item["saldo_zerado_manual"] else contract_balance(
            item["contrato_valor_moeda"], item["contrato_total_alocado"]
        )
        item["contrato_status"] = item["contrato_status"] or ""
        item["saldo_zerado_manual"] = bool(item["saldo_zerado_manual"])
        change_data.append(mapped_row(item, change_fields))

    due_fields = [
        ("vinculo_id", "Vinculo ID"), ("invoice_id", "Invoice ID"), ("due_id", "DU-E ID"),
        ("numero_invoice", "Numero da Invoice"), ("tipo_documento", "Tipo"),
        ("empresa_id", "Empresa ID"), ("empresa_cnpj", "Empresa - CNPJ"),
        ("empresa_razao_social", "Empresa - Razao social"), ("empresa_apelido", "Empresa - Apelido"),
        ("cliente_id", "Cliente ID"), ("cliente_nome", "Cliente"), ("cliente_pais", "Pais do cliente"),
        ("data_emissao", "Data de emissao"), ("invoice_moeda", "Moeda da Invoice"),
        ("invoice_valor_moeda", "Valor da Invoice"), ("numero_due", "Numero da DU-E"),
        ("chave_acesso", "Chave de acesso"), ("due_data_due", "Data da DU-E"),
        ("due_cnpj", "DU-E - CNPJ"), ("due_cliente", "DU-E - Cliente"),
        ("due_moeda", "Moeda da DU-E"), ("due_valor_original", "Valor original da DU-E"),
        ("due_utilizado", "Total utilizado na DU-E"), ("due_saldo", "Saldo da DU-E"),
        ("due_status", "Status da DU-E"), ("due_competencia_id", "DU-E - Competencia ID"),
        ("due_competencia_descricao", "DU-E - Competencia"), ("valor_vinculado", "Valor vinculado"),
        ("vinculo_observacao", "Observacao do vinculo"), ("due_observacao", "Observacao da DU-E"),
        ("vinculo_criado_em", "Vinculo criado em"), ("due_criado_em", "DU-E criada em"),
        ("invoice_status", "Status da Invoice"),
    ]
    due_data = []
    for row in due_links:
        item = dict(row)
        invoice = invoice_by_id.get(item["invoice_id"])
        item["tipo_documento"] = invoice_type_label(item["tipo_documento"])
        item["invoice_status"] = INVOICE_STATUS_LABELS.get(invoice["status"]) if invoice else None
        item["due_utilizado"] = due_usage.get(item["due_id"], Decimal("0"))
        item["due_saldo"] = due_balance(item["due_valor_original"], item["due_utilizado"])
        due_data.append(mapped_row(item, due_fields))

    invoice_columns = [label for _, label in invoice_fields]
    receipt_columns = [label for _, label in receipt_fields]
    change_columns = [label for _, label in change_fields]
    due_columns = [label for _, label in due_fields]
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheets = [
            ("Invoices", invoice_data, invoice_columns),
            ("Recebimentos", receipt_data, receipt_columns),
            ("Cambios", change_data, change_columns),
            ("DU-Es", due_data, due_columns),
        ]
        for sheet_name, data, columns in sheets:
            pd.DataFrame(data, columns=columns).to_excel(writer, index=False, sheet_name=sheet_name)
        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1769AA")
                cell.alignment = Alignment(horizontal="center")
            for column in worksheet.columns:
                letter = get_column_letter(column[0].column)
                width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 45)
                worksheet.column_dimensions[letter].width = width
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="invoices_exportacao.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/invoice/nova", methods=["GET", "POST"])
def nova_invoice():
    conn = db()
    if request.method == "POST":
        try:
            data = invoice_form_data(request.form, conn)
            conn.execute("""
                INSERT INTO invoices
                    (empresa_id,numero_invoice,tipo_documento,competencia_id,cliente_id,data_emissao,moeda,valor_moeda,
                     contrato_comercial,status,status_manual,data_credito,observacao)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (data["empresa_id"], data["numero_invoice"], data["tipo_documento"], data["competencia_id"], data["cliente_id"],
                  data["data_emissao"], data["moeda"], data["valor_moeda"],
                  data["contrato_comercial"], data["status"], data["status_manual"], data["data_credito"], data["observacao"]))
            invoice_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.commit(); conn.close()
            flash("Invoice cadastrada com sucesso.", "success")
            return redirect(url_for("detalhe_invoice", invoice_id=invoice_id))
        except sqlite3.IntegrityError:
            conn.rollback(); flash("Já existe uma Invoice com essa empresa, número e tipo documental.", "danger")
        except ValueError as exc:
            conn.rollback(); flash(str(exc), "danger")
    empresas = conn.execute("SELECT id,razao_social,apelido,cnpj FROM empresas ORDER BY razao_social").fetchall()
    clientes = clientes_for_form(conn)
    competencias = competencias_for_empresa(conn, None)
    conn.close()
    return render_template("invoice_form.html", invoice=None, empresas=empresas, clientes=clientes,
                           competencias=competencias,
                           invoice_types=INVOICE_STATUS_TYPES,
                           invoice_statuses=INVOICE_STATUS_LABELS)

@app.route("/invoice/<int:invoice_id>")
def detalhe_invoice(invoice_id):
    conn = db(); data = invoice_detail_data(conn, invoice_id); conn.close()
    if not data:
        return "Invoice não encontrada", 404
    invoice, empresas, clientes, contrapartes, dues, contratos = data
    return render_template("invoice_detail.html", invoice=invoice, empresas=empresas, clientes=clientes,
                           contrapartes=contrapartes, dues=dues, contratos=contratos,
                           invoice_types=INVOICE_STATUS_TYPES)

@app.route("/invoice/<int:invoice_id>/editar", methods=["GET", "POST"])
def editar_invoice(invoice_id):
    conn = db(); current = conn.execute("SELECT * FROM invoices WHERE id=?", (invoice_id,)).fetchone()
    if not current:
        conn.close(); return "Invoice não encontrada", 404
    if request.method == "POST":
        try:
            data = invoice_form_data(request.form, conn, current=current)
            summary = invoice_summary(conn, invoice_id)
            if decimal_value(data["valor_moeda"]) < summary["total_recebido"]:
                raise ValueError("O valor da Invoice não pode ficar abaixo do total recebido.")
            if decimal_value(data["valor_moeda"]) < summary["total_cambio"]:
                raise ValueError("O valor da Invoice não pode ficar abaixo do total de câmbio.")
            if data["moeda"] != current["moeda"] and (
                summary["total_recebido"] > SALDO_TOLERANCE or
                summary["total_cambio"] > SALDO_TOLERANCE
            ):
                raise ValueError("Não é possível alterar a moeda de uma Invoice com recebimento ou câmbio vinculado.")
            conn.execute("""
                UPDATE invoices SET empresa_id=?,numero_invoice=?,tipo_documento=?,competencia_id=?,cliente_id=?,data_emissao=?,
                    moeda=?,valor_moeda=?,contrato_comercial=?,status=?,status_manual=?,data_credito=?,observacao=? WHERE id=?
            """, (data["empresa_id"], data["numero_invoice"], data["tipo_documento"], data["competencia_id"], data["cliente_id"],
                  data["data_emissao"], data["moeda"], data["valor_moeda"], data["contrato_comercial"],
                  data["status"], data["status_manual"], data["data_credito"], data["observacao"], invoice_id))
            sync_contracts_for_invoice(conn, invoice_id)
            refresh_invoice_status(conn, invoice_id)
            conn.commit(); conn.close()
            flash("Invoice atualizada com sucesso.", "success")
            return redirect(url_for("detalhe_invoice", invoice_id=invoice_id))
        except sqlite3.IntegrityError:
            conn.rollback(); flash("Já existe uma Invoice com essa empresa, número e tipo documental.", "danger")
        except ValueError as exc:
            conn.rollback(); flash(str(exc), "danger")
    empresas = conn.execute("SELECT id,razao_social,apelido,cnpj FROM empresas ORDER BY razao_social").fetchall()
    clientes = clientes_for_form(conn)
    competencias = competencias_for_empresa(conn, None)
    conn.close()
    return render_template("invoice_form.html", invoice=current, empresas=empresas, clientes=clientes,
                           competencias=competencias,
                           invoice_types=INVOICE_STATUS_TYPES,
                           invoice_statuses=INVOICE_STATUS_LABELS)

@app.route("/invoice/<int:invoice_id>/recebimentos", methods=["POST"])
def adicionar_recebimento_invoice(invoice_id):
    conn = db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        summary = invoice_summary(conn, invoice_id)
        if not summary:
            raise ValueError("Invoice não encontrada.")
        amount = parse_number(request.form.get("valor_moeda"))
        if decimal_value(amount) <= 0:
            raise ValueError("O valor do recebimento deve ser maior que zero.")
        validate_invoice_balances(summary, extra_recebido=amount)
        bank = resolve_counterparty(conn, request.form.get("banco_credito_id"))
        data_credito = parse_date(request.form.get("data_credito"))
        if not data_credito:
            raise ValueError("A data de crédito é obrigatória.")
        conn.execute("""
            INSERT INTO recebimentos_invoice
                (invoice_id,banco_credito_id,data_credito,moeda,valor_moeda,documento,observacao)
            VALUES (?,?,?,?,?,?,?)
        """, (invoice_id, bank["id"] if bank else None, data_credito, summary["moeda"], amount,
              (request.form.get("documento") or "").strip() or None,
              (request.form.get("observacao") or "").strip() or None))
        refresh_invoice_status(conn, invoice_id); sync_contracts_for_invoice(conn, invoice_id); conn.commit()
        flash("Recebimento registrado com sucesso.", "success")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc) if isinstance(exc, ValueError) else "Não foi possível registrar o recebimento.", "danger")
    finally:
        conn.close()
    return redirect(url_for("detalhe_invoice", invoice_id=invoice_id))

@app.route("/invoice/<int:invoice_id>/recebimentos/<int:receipt_id>/editar", methods=["POST"])
def editar_recebimento_invoice(invoice_id, receipt_id):
    conn = db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        summary = invoice_summary(conn, invoice_id)
        receipt = conn.execute("SELECT * FROM recebimentos_invoice WHERE id=? AND invoice_id=?", (receipt_id, invoice_id)).fetchone()
        if not summary or not receipt:
            raise ValueError("Recebimento não encontrado.")
        amount = parse_number(request.form.get("valor_moeda"))
        if decimal_value(amount) <= 0:
            raise ValueError("O valor do recebimento deve ser maior que zero.")
        new_received = summary["total_recebido"] - decimal_value(receipt["valor_moeda"]) + decimal_value(amount)
        validate_invoice_balances(summary, replacement_recebido=new_received)
        bank = resolve_counterparty(conn, request.form.get("banco_credito_id"))
        data_credito = parse_date(request.form.get("data_credito"))
        if not data_credito:
            raise ValueError("A data de crédito é obrigatória.")
        conn.execute("""
            UPDATE recebimentos_invoice SET banco_credito_id=?,data_credito=?,valor_moeda=?,documento=?,observacao=?
            WHERE id=? AND invoice_id=?
        """, (bank["id"] if bank else None, data_credito, amount,
              (request.form.get("documento") or "").strip() or None,
              (request.form.get("observacao") or "").strip() or None, receipt_id, invoice_id))
        refresh_invoice_status(conn, invoice_id); sync_contracts_for_invoice(conn, invoice_id); conn.commit(); flash("Recebimento atualizado com sucesso.", "success")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc) if isinstance(exc, ValueError) else "Não foi possível atualizar o recebimento.", "danger")
    finally:
        conn.close()
    return redirect(url_for("detalhe_invoice", invoice_id=invoice_id))

@app.route("/invoice/<int:invoice_id>/recebimentos/<int:receipt_id>/excluir", methods=["POST"])
def excluir_recebimento_invoice(invoice_id, receipt_id):
    conn = db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        summary = invoice_summary(conn, invoice_id)
        receipt = conn.execute("SELECT valor_moeda FROM recebimentos_invoice WHERE id=? AND invoice_id=?", (receipt_id, invoice_id)).fetchone()
        if not summary or not receipt:
            raise ValueError("Recebimento não encontrado.")
        new_received = summary["total_recebido"] - decimal_value(receipt["valor_moeda"])
        validate_invoice_balances(summary, replacement_recebido=new_received)
        conn.execute("DELETE FROM recebimentos_invoice WHERE id=? AND invoice_id=?", (receipt_id, invoice_id))
        refresh_invoice_status(conn, invoice_id); sync_contracts_for_invoice(conn, invoice_id); conn.commit(); flash("Recebimento excluído com sucesso.", "success")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc) if isinstance(exc, ValueError) else "Não foi possível excluir o recebimento.", "danger")
    finally:
        conn.close()
    return redirect(url_for("detalhe_invoice", invoice_id=invoice_id))

@app.route("/invoice/<int:invoice_id>/cambio", methods=["POST"])
def adicionar_cambio_invoice(invoice_id):
    conn = db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        summary = invoice_summary(conn, invoice_id)
        if not summary:
            raise ValueError("Invoice não encontrada.")
        amount = parse_number(request.form.get("valor_alocado"))
        if decimal_value(amount) <= 0:
            raise ValueError("O valor de câmbio deve ser maior que zero.")
        validate_invoice_balances(summary, extra_cambio=amount)
        metadata = contract_metadata_from_form(request.form, conn)
        metadata["contrato_id"] = request.form.get("contrato_id")
        contract_id = contract_for_invoice(conn, summary, metadata)
        link = conn.execute("SELECT id FROM invoice_contrato_cambio WHERE invoice_id=? AND contrato_id=?", (invoice_id, contract_id)).fetchone()
        if link:
            conn.execute("UPDATE invoice_contrato_cambio SET valor_alocado=valor_alocado+?,observacao=? WHERE id=?",
                         (amount, metadata.get("observacao"), link["id"]))
        else:
            conn.execute("INSERT INTO invoice_contrato_cambio(invoice_id,contrato_id,valor_alocado,observacao) VALUES (?,?,?,?)",
                         (invoice_id, contract_id, amount, metadata.get("observacao")))
        sync_contract_cache(conn, contract_id); refresh_invoice_status(conn, invoice_id)
        conn.commit(); flash("Câmbio vinculado à Invoice com sucesso.", "success")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc) if isinstance(exc, ValueError) else "Não foi possível vincular o câmbio.", "danger")
    finally:
        conn.close()
    return redirect(url_for("detalhe_invoice", invoice_id=invoice_id))

@app.route("/invoice/<int:invoice_id>/cambio/<int:link_id>/editar", methods=["POST"])
def editar_cambio_invoice(invoice_id, link_id):
    conn = db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        summary = invoice_summary(conn, invoice_id)
        link = conn.execute("SELECT * FROM invoice_contrato_cambio WHERE id=? AND invoice_id=?", (link_id, invoice_id)).fetchone()
        if not summary or not link:
            raise ValueError("Vínculo de câmbio não encontrado.")
        amount = parse_number(request.form.get("valor_alocado"))
        if decimal_value(amount) <= 0:
            raise ValueError("O valor de câmbio deve ser maior que zero.")
        new_total = summary["total_cambio"] - decimal_value(link["valor_alocado"]) + decimal_value(amount)
        validate_invoice_balances(summary, replacement_cambio=new_total)
        conn.execute("UPDATE invoice_contrato_cambio SET valor_alocado=?,observacao=? WHERE id=?",
                     (amount, (request.form.get("observacao") or "").strip() or None, link_id))
        sync_contract_cache(conn, link["contrato_id"]); refresh_invoice_status(conn, invoice_id)
        conn.commit(); flash("Alocação de câmbio atualizada com sucesso.", "success")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc) if isinstance(exc, ValueError) else "Não foi possível atualizar o câmbio.", "danger")
    finally:
        conn.close()
    return redirect(url_for("detalhe_invoice", invoice_id=invoice_id))

@app.route("/invoice/<int:invoice_id>/cambio/<int:link_id>/excluir", methods=["POST"])
def excluir_cambio_invoice(invoice_id, link_id):
    conn = db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        link = conn.execute("SELECT * FROM invoice_contrato_cambio WHERE id=? AND invoice_id=?", (link_id, invoice_id)).fetchone()
        if not link:
            raise ValueError("Vínculo de câmbio não encontrado.")
        conn.execute("DELETE FROM invoice_contrato_cambio WHERE id=?", (link_id,))
        sync_contract_cache(conn, link["contrato_id"])
        if not conn.execute("SELECT 1 FROM invoice_contrato_cambio WHERE contrato_id=?", (link["contrato_id"],)).fetchone():
            if conn.execute("SELECT 1 FROM due_contratos WHERE contrato_id=?", (link["contrato_id"],)).fetchone():
                raise ValueError("O Contrato Câmbio possui vínculo direto com DU-E e não pode ficar sem Invoice.")
            conn.execute("DELETE FROM contratos WHERE id=?", (link["contrato_id"],))
        refresh_invoice_status(conn, invoice_id); conn.commit(); flash("Vínculo de câmbio excluído com sucesso.", "success")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc) if isinstance(exc, ValueError) else "Não foi possível excluir o câmbio.", "danger")
    finally:
        conn.close()
    return redirect(url_for("detalhe_invoice", invoice_id=invoice_id))

@app.route("/invoice/<int:invoice_id>/due", methods=["POST"])
def adicionar_due_invoice(invoice_id):
    conn = db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        summary = invoice_summary(conn, invoice_id)
        due_id = form_record_id(request.form.get("due_id"))
        due = conn.execute("SELECT * FROM dues WHERE id=?", (due_id,)).fetchone() if due_id else None
        if not summary or not due:
            raise ValueError("Invoice ou DU-E não encontrada.")
        if due["moeda"] != summary["moeda"]:
            raise ValueError("A moeda da DU-E deve ser igual à moeda da Invoice.")
        amount = parse_number(request.form.get("valor_vinculado"))
        if decimal_value(amount) <= 0:
            raise ValueError("O valor do vínculo deve ser maior que zero.")
        current = conn.execute("SELECT valor_vinculado FROM due_invoice WHERE due_id=? AND invoice_id=?", (due_id, invoice_id)).fetchone()
        existing = decimal_value(current["valor_vinculado"]) if current else Decimal("0")
        total_due_invoice = decimal_value(conn.execute("SELECT COALESCE(SUM(valor_vinculado),0) FROM due_invoice WHERE invoice_id=?", (invoice_id,)).fetchone()[0])
        if current:
            total_due_invoice -= existing
        ensure_non_negative_balance(summary["total_cambio"] - total_due_invoice - decimal_value(amount),
                                    "O vínculo com DU-E não pode ultrapassar o total de câmbio da Invoice.")
        conn.execute("""
            INSERT INTO due_invoice(due_id,invoice_id,valor_vinculado,observacao) VALUES (?,?,?,?)
            ON CONFLICT(due_id,invoice_id) DO UPDATE SET valor_vinculado=excluded.valor_vinculado,
                observacao=excluded.observacao
        """, (due_id, invoice_id, amount, (request.form.get("observacao") or "").strip() or None))
        conn.commit(); flash("Invoice vinculada à DU-E para rastreabilidade.", "success")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc) if isinstance(exc, ValueError) else "Não foi possível vincular a DU-E.", "danger")
    finally:
        conn.close()
    return redirect(url_for("detalhe_invoice", invoice_id=invoice_id))

@app.route("/invoice/<int:invoice_id>/due/<int:link_id>/excluir", methods=["POST"])
def excluir_due_invoice(invoice_id, link_id):
    conn = db()
    try:
        conn.execute("DELETE FROM due_invoice WHERE id=? AND invoice_id=?", (link_id, invoice_id)); conn.commit()
        flash("Vínculo Invoice↔DU-E excluído com sucesso.", "success")
    except sqlite3.Error:
        conn.rollback(); flash("Não foi possível excluir o vínculo com a DU-E.", "danger")
    finally:
        conn.close()
    return redirect(url_for("detalhe_invoice", invoice_id=invoice_id))

@app.route("/invoice/<int:invoice_id>/due/<int:link_id>/editar", methods=["POST"])
def editar_due_invoice(invoice_id, link_id):
    conn = db()
    try:
        conn.execute("BEGIN IMMEDIATE")
        summary = invoice_summary(conn, invoice_id)
        link = conn.execute("SELECT * FROM due_invoice WHERE id=? AND invoice_id=?", (link_id, invoice_id)).fetchone()
        if not summary or not link:
            raise ValueError("Vínculo Invoice↔DU-E não encontrado.")
        amount = parse_number(request.form.get("valor_vinculado"))
        if decimal_value(amount) <= 0:
            raise ValueError("O valor do vínculo deve ser maior que zero.")
        linked = decimal_value(conn.execute("SELECT COALESCE(SUM(valor_vinculado),0) FROM due_invoice WHERE invoice_id=?", (invoice_id,)).fetchone()[0])
        linked = linked - decimal_value(link["valor_vinculado"]) + decimal_value(amount)
        ensure_non_negative_balance(summary["total_cambio"] - linked,
                                    "O vínculo com DU-E não pode ultrapassar o total de câmbio da Invoice.")
        conn.execute("UPDATE due_invoice SET valor_vinculado=?,observacao=? WHERE id=?",
                     (amount, (request.form.get("observacao") or "").strip() or None, link_id))
        conn.commit(); flash("Vínculo Invoice↔DU-E atualizado com sucesso.", "success")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc) if isinstance(exc, ValueError) else "Não foi possível atualizar o vínculo.", "danger")
    finally:
        conn.close()
    return redirect(url_for("detalhe_invoice", invoice_id=invoice_id))

@app.route("/invoices/<int:invoice_id>/saldo")
def saldo_invoice(invoice_id):
    conn = db(); summary = invoice_summary(conn, invoice_id); conn.close()
    if not summary:
        return jsonify({"error": "Invoice não encontrada."}), 404
    return jsonify({
        "id": summary["id"], "numero_invoice": summary["numero_invoice"], "moeda": summary["moeda"],
        "valor_invoice": float(summary["valor_moeda"]), "total_recebido": float(summary["total_recebido"]),
        "saldo_recebimento": float(summary["saldo_recebimento"]), "total_cambio": float(summary["total_cambio"]),
        "saldo_cambio": float(summary["saldo_cambio"]), "status": summary["status"],
        "status_label": INVOICE_STATUS_LABELS[summary["status"]],
    })

def invoice_import_preview_context(payload):
    groups = {}
    for row in payload["rows"]:
        groups.setdefault((row["cnpj"], row["numero_invoice"], row["tipo_documento"]), []).append(row)
    return [{"group_id": f"g{index}", "key": key, "rows": rows,
             "existing": payload.get("existing_by_key", {}).get("|".join(map(str, key)))}
            for index, (key, rows) in enumerate(groups.items())]

def revalidate_invoice_import_stage(conn, payload):
    expected = payload.get("existing_by_key", {})
    current = invoice_identity_rows(conn, payload.get("rows", []))
    for key, snapshot in expected.items():
        if invoice_import_snapshot(current.get(tuple(key.split("|")))) != snapshot:
            raise ValueError("Uma Invoice foi alterada desde a prévia. Analise a planilha novamente.")

@app.route("/invoices/importar", methods=["POST"])
def importar_invoices():
    conn = None
    try:
        import pandas as pd
        arquivo = request.files.get("arquivo")
        if not arquivo or not arquivo.filename:
            flash("Selecione um arquivo Excel de Invoices.", "danger")
            return redirect(url_for("lista_invoices"))
        df = pd.read_excel(arquivo)
        df.columns = normalize_invoice_import_columns(df.columns)
        rows = prepare_invoice_import_rows(df, pd)
        conn = db()
        resolve_invoice_import_companies(conn, rows)
        bank_suggestions = resolve_invoice_import_banks(conn, rows)
        client_suggestions = resolve_invoice_import_clients(conn, rows)
        competencia_suggestions = resolve_invoice_import_competencies(conn, rows)
        existing = invoice_identity_rows(conn, rows)
        snapshots = {}
        for key, item in existing.items():
            snapshots["|".join(map(str, key))] = invoice_import_snapshot(item)
        payload = {"version": 1, "rows": rows, "existing_by_key": snapshots,
                   "bank_suggestions": bank_suggestions,
                   "client_suggestions": client_suggestions,
                   "competencia_suggestions": competencia_suggestions}
        save_invoice_stage(payload)
        return render_template("invoice_import_preview.html", groups=invoice_import_preview_context(payload),
                               client_suggestions=client_suggestions, countries=CLIENTE_PAISES_ORDENADOS,
                               clientes=clientes_for_form(conn),
                               bank_suggestions=bank_suggestions, contrapartes=contrapartes_for_form(conn),
                               competencia_suggestions=competencia_suggestions,
                               total_rows=len(rows), stage_token=session.get("invoice_import_stage"), error=None)
    except ValueError as exc:
        flash(str(exc), "danger")
    except Exception as exc:
        flash(f"Erro na análise da importação: {exc}", "danger")
    finally:
        if conn is not None:
            conn.close()
    return redirect(url_for("lista_invoices"))

@app.route("/invoices/importar/confirmar", methods=["POST"])
def confirmar_importacao_invoices():
    payload = None; conn = None
    try:
        payload = load_invoice_stage(request.form.get("stage_token"))
        conn = db(); conn.execute("BEGIN IMMEDIATE")
        revalidate_invoice_import_stage(conn, payload)
        existing_keys = {key for key, value in payload.get("existing_by_key", {}).items() if value}
        selected_rows = []
        groups = invoice_import_preview_context(payload)
        for group in groups:
            key = "|".join(map(str, group["key"]))
            if key in existing_keys and request.form.get(f"existing_action_{group['group_id']}") == "discard":
                continue
            selected_rows.extend(group["rows"])
        country_overrides = {}
        client_overrides = {}
        bank_overrides = {}
        selected_bank_keys = {
            invoice_import_bank_key(field, row.get(field))
            for row in selected_rows
            for field in ("banco_credito", "banco_liquidacao")
            if row.get(field)
        }
        for suggestion in payload.get("bank_suggestions", []):
            if suggestion["key"] not in selected_bank_keys:
                continue
            selected_bank_id = request.form.get(f"banco_existente_{suggestion['suggestion_id']}")
            if not selected_bank_id:
                raise ValueError(
                    f"Selecione um banco cadastrado para {suggestion['nome']} ({suggestion['tipo']})."
                )
            bank_overrides[suggestion["key"]] = selected_bank_id
            suggestion["contraparte_id"] = selected_bank_id
        selected_client_keys = {row.get("cliente_import_key") for row in selected_rows}
        for suggestion in payload.get("client_suggestions", []):
            if suggestion["key"] not in selected_client_keys:
                continue
            selected_client_id = request.form.get(f"cliente_existente_{suggestion['suggestion_id']}")
            if selected_client_id:
                client_overrides[suggestion["key"]] = selected_client_id
                suggestion["cliente_existente_id"] = selected_client_id
                continue
            country = normalize_import_client_country(
                request.form.get(f"cliente_novo_pais_{suggestion['suggestion_id']}") or suggestion.get("pais")
            )
            if not country:
                raise ValueError(
                    f"Informe o País para o novo cliente sugerido {suggestion['nome']}."
                )
            country_overrides[suggestion["key"]] = country
            suggestion["cliente_novo_pais"] = country
        competency_overrides = {}
        selected_competency_keys = {row.get("competencia_import_key") for row in selected_rows}
        for suggestion in payload.get("competencia_suggestions", []):
            if suggestion["key"] not in selected_competency_keys:
                continue
            if request.form.get(f"competencia_nova_confirmar_{suggestion['suggestion_id']}") != "1":
                raise ValueError(
                    f"Confirme o cadastro da competência {suggestion['descricao']} "
                    f"para a empresa {suggestion.get('empresa_apelido') or suggestion.get('empresa_nome')}."
                )
            competency_overrides[suggestion["key"]] = {
                "descricao": request.form.get(f"competencia_nova_descricao_{suggestion['suggestion_id']}"),
                "data_inicial": request.form.get(f"competencia_nova_data_inicial_{suggestion['suggestion_id']}"),
                "data_final": request.form.get(f"competencia_nova_data_final_{suggestion['suggestion_id']}"),
            }
        result = apply_invoice_import_rows(
            conn, selected_rows, country_overrides=country_overrides,
            client_overrides=client_overrides,
            competency_overrides=competency_overrides, bank_overrides=bank_overrides,
        )
        conn.commit(); remove_invoice_stage(request.form.get("stage_token"))
        flash(f"Importação concluída: {result['inserted']} Invoice(s) inserida(s) e {result['updated']} atualizada(s).", "success")
        return redirect(url_for("lista_invoices"))
    except (ValueError, sqlite3.Error) as exc:
        if conn is not None: conn.rollback()
        message = str(exc) if isinstance(exc, ValueError) else "Não foi possível concluir a importação."
        if payload is not None:
            return render_template("invoice_import_preview.html", groups=invoice_import_preview_context(payload),
                                   client_suggestions=payload.get("client_suggestions", []),
                                   clientes=clientes_for_form(conn),
                                   bank_suggestions=payload.get("bank_suggestions", []),
                                   contrapartes=contrapartes_for_form(conn),
                                   competencia_suggestions=payload.get("competencia_suggestions", []),
                                   countries=CLIENTE_PAISES_ORDENADOS, total_rows=len(payload["rows"]),
                                   stage_token=session.get("invoice_import_stage"), error=message), 400
        flash(message, "danger")
    finally:
        if conn is not None: conn.close()
    return redirect(url_for("lista_invoices"))

@app.route("/invoices/importar/cancelar", methods=["POST"])
def cancelar_importacao_invoices():
    remove_invoice_stage(request.form.get("stage_token")); flash("Importação cancelada.", "success")
    return redirect(url_for("lista_invoices"))

@app.route("/invoices/modelo")
def modelo_invoices():
    import pandas as pd
    columns = ["empresa", "invoice", "contrato_comercial", "competencia", "tipo",
               "banco_credito", "banco_liquidacao", "contrato_cambio", "cliente", "emissao",
               "data_credito", "data_fechamento", "data_liquidacao", "moeda", "valor_moeda",
               "taxa_cambio", "valor_brl", "status"]
    output = io.BytesIO()
    conn = db()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            pd.DataFrame(columns=columns).to_excel(writer, index=False, sheet_name="Invoices")
            write_excel_model_orientations(writer, conn, pd)
    finally:
        conn.close()
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="modelo_invoices.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
